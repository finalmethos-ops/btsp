from dataclasses import dataclass
from io import BytesIO
from secrets import token_urlsafe
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.catalog import CatalogVendor
from app.models.identity import User
from app.schemas.user_admin import UserCreate
from app.services.user_admin_service import create_user

MAX_USER_ROWS = 2_000
REQUIRED_HEADERS = {
    "Vendors": ("Vendor", "Name", "E-mail"),
    "Buddys": ("Entity", "Approval Rights", "Name", "Email"),
    "Operations": ("Role", "Position", "Name", "E-mail"),
}
OPERATIONS_ROLE_CODES = {
    "MASTER ADMIN": "SYSTEM_ADMIN",
    "ADMIN": "ADMIN",
    "PURCHASING": "PURCHASING",
    "RECONCILIATION": "RECONCILIATION",
    "EXECUTIVE": "EXECUTIVE",
}


class UserListingImportError(ValueError):
    pass


@dataclass
class ImportedUser:
    email: str
    display_name: str
    role_codes: set[str]
    vendor_codes: set[str]
    entity_code: str | None = None
    region_code: str | None = None


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _email(value: Any, sheet: str, row_number: int) -> str:
    email = _text(value).lower()
    if not email or "@" not in email or email.startswith("@") or email.endswith("@"):
        raise UserListingImportError(f"Invalid email on {sheet} row {row_number}")
    return email


def _region(value: Any) -> str | None:
    text = _text(value).upper()
    if not text:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    return "ALL_STORES" if text == "ALL STORES" else text


def _vendor_code(name: str) -> str:
    code = "".join(character if character.isalnum() else "-" for character in name.upper())
    return "-".join(part for part in code.split("-") if part)[:64]


def _worksheet_rows(workbook: Any, sheet: str) -> list[tuple[int, tuple[Any, ...]]]:
    if sheet not in workbook.sheetnames:
        raise UserListingImportError(f"Workbook is missing the {sheet} sheet")
    rows = workbook[sheet].iter_rows(values_only=True)
    try:
        headers = tuple(_text(value) for value in next(rows))
    except StopIteration as exc:
        raise UserListingImportError(f"{sheet} sheet is empty") from exc
    if headers[: len(REQUIRED_HEADERS[sheet])] != REQUIRED_HEADERS[sheet]:
        raise UserListingImportError(f"{sheet} sheet headers do not match the approved template")
    return [
        (row_number, values)
        for row_number, values in enumerate(rows, start=2)
        if any(value not in (None, "") for value in values)
    ]


def parse_user_listing(
    content: bytes,
    db: Session,
    create_missing_vendors: bool = False,
) -> list[ImportedUser]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise UserListingImportError("File is not a readable Excel workbook") from exc

    records: dict[str, ImportedUser] = {}

    def get_user(email: str, display_name: str) -> ImportedUser:
        record = records.get(email)
        if record is None:
            record = ImportedUser(
                email=email,
                display_name=display_name,
                role_codes=set(),
                vendor_codes=set(),
            )
            records[email] = record
        return record

    vendor_by_name = {
        vendor.name.strip().casefold(): vendor.vendor_code
        for vendor in db.scalars(select(CatalogVendor).where(CatalogVendor.is_active.is_(True)))
    }
    for row_number, values in _worksheet_rows(workbook, "Vendors"):
        if row_number > MAX_USER_ROWS:
            raise UserListingImportError("Vendor sheet exceeds the allowed row limit")
        padded = values + (None,) * 3
        vendor_name, display_name = _text(padded[0]), _text(padded[1])
        email = _email(padded[2], "Vendors", row_number)
        vendor_code = vendor_by_name.get(vendor_name.casefold())
        if vendor_code is None:
            if not create_missing_vendors:
                raise UserListingImportError(
                    f"No active BTSP vendor matches '{vendor_name}' on Vendors row {row_number}"
                )
            vendor_code = _vendor_code(vendor_name)
            conflict = db.scalar(
                select(CatalogVendor).where(CatalogVendor.vendor_code == vendor_code)
            )
            if conflict is not None and conflict.name.strip().casefold() != vendor_name.casefold():
                raise UserListingImportError(
                    f"Vendor code collision for '{vendor_name}' on Vendors row {row_number}"
                )
            if conflict is None:
                db.add(
                    CatalogVendor(
                        vendor_code=vendor_code,
                        name=vendor_name,
                        is_active=True,
                        source_file="user_listing_import",
                    )
                )
                db.flush()
            vendor_by_name[vendor_name.casefold()] = vendor_code
        record = get_user(email, display_name)
        record.role_codes.add("VENDOR")
        record.vendor_codes.add(vendor_code)

    for row_number, values in _worksheet_rows(workbook, "Buddys"):
        if row_number > MAX_USER_ROWS:
            raise UserListingImportError("Buddys sheet exceeds the allowed row limit")
        padded = values + (None,) * 4
        entity_code, approval_rights, display_name = (
            _text(padded[0]).upper(),
            _region(padded[1]),
            _text(padded[2]),
        )
        email = _email(padded[3], "Buddys", row_number)
        record = get_user(email, display_name)
        record.role_codes.add("FRANCHISE_OPERATOR")
        record.entity_code = entity_code or record.entity_code
        record.region_code = approval_rights or record.region_code

    for row_number, values in _worksheet_rows(workbook, "Operations"):
        if row_number > MAX_USER_ROWS:
            raise UserListingImportError("Operations sheet exceeds the allowed row limit")
        padded = values + (None,) * 4
        workbook_role, display_name = _text(padded[0]).upper(), _text(padded[2])
        email = _email(padded[3], "Operations", row_number)
        role_code = OPERATIONS_ROLE_CODES.get(workbook_role)
        if role_code is None:
            raise UserListingImportError(
                f"Unsupported Operations role '{workbook_role}' on row {row_number}"
            )
        record = get_user(email, display_name)
        record.role_codes.add(role_code)

    return sorted(records.values(), key=lambda record: record.email)


def import_user_listing(content: bytes, db: Session) -> tuple[int, int]:
    records = parse_user_listing(content, db, create_missing_vendors=True)
    existing_emails = set(db.scalars(select(User.email)).all())
    created = 0
    skipped = 0
    for record in records:
        if record.email in existing_emails:
            skipped += 1
            continue
        # Imported accounts are inactive until an administrator assigns a
        # password and deliberately activates the account in User Management.
        create_user(
            db,
            UserCreate(
                email=record.email,
                display_name=record.display_name,
                password=token_urlsafe(32),
                entity_code=record.entity_code,
                region_code=record.region_code,
                vendor_codes=sorted(record.vendor_codes),
                is_active=False,
                password_change_required=True,
                role_codes=sorted(record.role_codes),
            ),
        )
        created += 1
    return created, skipped
