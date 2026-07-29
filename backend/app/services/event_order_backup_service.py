import re
from datetime import UTC, datetime
from decimal import Decimal
from hashlib import sha256
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models.catalog import CatalogProduct
from app.models.event_management import (
    EventEntityOrder,
    EventOrderBackupArtifact,
    EventProductSlide,
    ManagedEvent,
    ManagedSubEvent,
)
from app.models.identity import User
from app.models.purchase_order import PurchaseOrder, PurchaseOrderSource
from app.models.purchasing import PurchaseRequest
from app.models.store import Store
from app.services.spreadsheet_security import spreadsheet_safe_row

HEADER_FILL = PatternFill("solid", fgColor="123A73")
HEADER_FONT = Font(bold=True, color="FFFFFF")
MONEY_FORMAT = "$#,##0.00"
ORDER_HEADERS = [
    "channel",
    "source_order_id",
    "order_number",
    "sub_event",
    "entity",
    "region",
    "store",
    "vendor",
    "model_number",
    "product_name",
    "quantity",
    "unit_cost",
    "line_total",
    "delivery_start",
    "delivery_end",
    "order_status",
    "review_status",
    "placed_by",
    "placed_at",
    "updated_at",
]
ORDER_WIDTHS = [20, 38, 38, 28, 18, 16, 14, 18, 22, 36, 12, 14, 16, 16, 16, 22, 18, 30, 24, 24]


def event_order_backup_filename(event: ManagedEvent) -> str:
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "-", event.name).strip("-") or "event"
    return f"{safe_name}-all-orders.xlsx"


def _datetime_text(value) -> str:
    return value.isoformat() if value else ""


def _sheet_name(entity: str, used: set[str]) -> str:
    clean = "".join("-" if character in "[]:*?/\\" else character for character in entity)
    base = f"Entity {clean.strip() or 'Unassigned'}"[:31]
    name = base
    sequence = 2
    while name.casefold() in used:
        suffix = f" {sequence}"
        name = f"{base[:31-len(suffix)]}{suffix}"
        sequence += 1
    used.add(name.casefold())
    return name


def _variants(order: EventEntityOrder, slide: EventProductSlide):
    variants = {str(item["model_number"]): item for item in (slide.product_variants or [])}
    if variants and order.variant_quantities:
        return [
            (
                model,
                str(variants[model].get("name") or slide.name),
                quantity,
                Decimal(str(variants[model]["event_unit_cost"])),
            )
            for model, quantity in order.variant_quantities.items()
            if quantity > 0 and model in variants
        ]
    return [(slide.model_number, slide.name, order.quantity, order.unit_cost)]


def _style_table(sheet, widths: list[int], money_columns: tuple[int, ...] = ()) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for column in money_columns:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, column).number_format = MONEY_FORMAT


def _append_safe(sheet, values) -> None:
    sheet.append(spreadsheet_safe_row(values))


def export_event_order_backup(db: Session, event_id: str) -> tuple[ManagedEvent, bytes] | None:
    event = db.get(ManagedEvent, event_id)
    if event is None:
        return None

    live_rows = db.execute(
        select(EventEntityOrder, EventProductSlide, ManagedSubEvent, User)
        .join(EventProductSlide, EventProductSlide.id == EventEntityOrder.slide_id)
        .join(ManagedSubEvent, ManagedSubEvent.id == EventEntityOrder.sub_event_id)
        .join(User, User.id == EventEntityOrder.user_id)
        .where(EventEntityOrder.event_id == event_id)
        .order_by(ManagedSubEvent.starts_at, EventEntityOrder.submitted_at)
    ).all()
    requests = (
        db.scalars(
            select(PurchaseRequest)
            .options(selectinload(PurchaseRequest.line_items))
            .where(PurchaseRequest.workflow_code == "VENDOR_ORDER")
            .order_by(PurchaseRequest.created_at)
        )
        .unique()
        .all()
    )
    event_requests = [
        request for request in requests if request.context.get("event_id") == event_id
    ]
    event_request_ids = {request.id for request in event_requests}
    purchase_orders = (
        (
            db.scalars(
                select(PurchaseOrder)
                .options(selectinload(PurchaseOrder.sources), selectinload(PurchaseOrder.lines))
                .where(
                    PurchaseOrder.sources.any(
                        PurchaseOrderSource.purchase_request_id.in_(event_request_ids)
                    )
                )
                .order_by(PurchaseOrder.created_at)
            )
            .unique()
            .all()
        )
        if event_request_ids
        else []
    )
    buy_fair_requests = [
        request
        for request in event_requests
        if request.context.get("source") == "event_vendor_buy_fair"
    ]
    product_codes = {
        line.product_code for request in buy_fair_requests for line in request.line_items
    }
    product_codes.update(
        line.product_code
        for order in purchase_orders
        for line in order.lines
        if line.source_request_id in event_request_ids
    )
    products = (
        {
            product.product_code: product
            for product in db.scalars(
                select(CatalogProduct).where(CatalogProduct.product_code.in_(product_codes))
            ).all()
        }
        if product_codes
        else {}
    )
    stores = {store.store_number: store for store in db.scalars(select(Store)).all()}
    stores_by_entity: dict[str, list[Store]] = {}
    for store in stores.values():
        if store.entity_code:
            stores_by_entity.setdefault(store.entity_code, []).append(store)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Event Summary"
    _append_safe(summary, ["Event order backup", event.name])
    _append_safe(summary, ["Event ID", event.id])
    _append_safe(
        summary,
        ["Event dates", f"{event.starts_at.isoformat()} through {event.ends_at.isoformat()}"],
    )
    _append_safe(
        summary,
        [
            "Location",
            ", ".join(value for value in [event.venue_name, event.city, event.state_code] if value),
        ],
    )
    _append_safe(summary, ["Generated UTC", datetime.now(UTC).isoformat()])
    _append_safe(summary, [])
    _append_safe(summary, ["Channel", "Orders", "Units", "Spend"])

    order_records: list[list[object]] = []

    live_units = 0
    live_spend = Decimal("0")
    for order, slide, sub_event, user in live_rows:
        store = stores.get(user.home_store_number or "")
        if store is None and len(stores_by_entity.get(order.entity_code, [])) == 1:
            store = stores_by_entity[order.entity_code][0]
        for model, name, quantity, unit_cost in _variants(order, slide):
            total = unit_cost * quantity
            live_units += quantity
            live_spend += total
            order_records.append(
                [
                    "Live presentation",
                    order.id,
                    "",
                    sub_event.name,
                    order.entity_code,
                    store.region_code if store else "",
                    store.store_number if store else "",
                    slide.vendor_code,
                    model,
                    name,
                    quantity,
                    unit_cost,
                    total,
                    order.requested_delivery_start,
                    order.requested_delivery_end,
                    order.status,
                    order.review_status,
                    user.email,
                    _datetime_text(order.submitted_at),
                    _datetime_text(order.updated_at),
                ]
            )

    buy_fair_units = Decimal("0")
    buy_fair_spend = Decimal("0")
    for request in buy_fair_requests:
        store = stores.get(request.store_number)
        entity_code = (store.entity_code if store else None) or str(
            request.context.get("entity_code") or "Unassigned"
        )
        for line in request.line_items:
            product = products.get(line.product_code)
            buy_fair_units += line.quantity
            buy_fair_spend += line.extended_amount
            order_records.append(
                [
                    "Vendor buy fair",
                    request.id,
                    request.order_number,
                    str(
                        request.context.get("sub_event_name")
                        or request.context.get("sub_event_id")
                        or ""
                    ),
                    entity_code,
                    store.region_code if store else "",
                    request.store_number,
                    request.vendor_code,
                    product.model_number if product and product.model_number else line.product_code,
                    line.product_name,
                    line.quantity,
                    line.unit_price,
                    line.extended_amount,
                    line.requested_delivery_date or request.expected_delivery_date,
                    "",
                    request.status,
                    "",
                    request.created_by,
                    _datetime_text(request.created_at),
                    _datetime_text(request.updated_at),
                ]
            )

    order_records.sort(
        key=lambda row: tuple(str(row[index] or "").casefold() for index in (7, 5, 6, 2, 8))
    )
    all_orders = workbook.create_sheet("All Order Lines")
    _append_safe(all_orders, ORDER_HEADERS)
    for record in order_records:
        _append_safe(all_orders, record)

    _append_safe(summary, ["Live presentation", len(live_rows), live_units, live_spend])
    _append_safe(
        summary, ["Vendor buy fair", len(buy_fair_requests), buy_fair_units, buy_fair_spend]
    )
    _append_safe(
        summary,
        [
            "Combined",
            len(live_rows) + len(buy_fair_requests),
            live_units + buy_fair_units,
            live_spend + buy_fair_spend,
        ],
    )
    event_po_lines = [
        line
        for order in purchase_orders
        for line in order.lines
        if line.source_request_id in event_request_ids
    ]
    _append_safe(
        summary,
        [
            "Generated purchase orders",
            len(purchase_orders),
            sum((line.quantity for line in event_po_lines), Decimal("0")),
            sum((line.extended_amount for line in event_po_lines), Decimal("0")),
        ],
    )
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 48
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 16
    for cell in summary[7]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in range(8, 12):
        summary.cell(row, 4).number_format = MONEY_FORMAT

    _style_table(
        all_orders,
        ORDER_WIDTHS,
        (12, 13),
    )

    used_sheet_names = {name.casefold() for name in workbook.sheetnames}
    entities = sorted({str(row[4] or "Unassigned") for row in order_records}, key=str.casefold)
    for entity_code in entities:
        entity_sheet = workbook.create_sheet(_sheet_name(entity_code, used_sheet_names))
        _append_safe(entity_sheet, ORDER_HEADERS)
        for record in order_records:
            if str(record[4] or "Unassigned") == entity_code:
                _append_safe(entity_sheet, record)
        _style_table(entity_sheet, ORDER_WIDTHS, (12, 13))

    handoff = workbook.create_sheet("Purchasing Handoff")
    _append_safe(
        handoff,
        [
            "purchase_request_id",
            "order_number",
            "source",
            "store",
            "vendor",
            "status",
            "units",
            "subtotal",
            "total",
            "release_batch_id",
            "created_by",
            "created_at",
        ],
    )
    for request in event_requests:
        _append_safe(
            handoff,
            [
                request.id,
                request.order_number,
                request.context.get("source", ""),
                request.store_number,
                request.vendor_code,
                request.status,
                sum((line.quantity for line in request.line_items), Decimal("0")),
                request.subtotal,
                request.total,
                request.context.get("release_batch_id", ""),
                request.created_by,
                _datetime_text(request.created_at),
            ],
        )
    _style_table(handoff, [38, 42, 28, 14, 18, 24, 12, 16, 16, 38, 30, 24], (8, 9))

    po_sheet = workbook.create_sheet("Purchase Orders")
    _append_safe(
        po_sheet,
        [
            "purchase_order_id",
            "po_number",
            "status",
            "entity",
            "region",
            "store",
            "vendor",
            "model_number",
            "product_name",
            "quantity",
            "received_quantity",
            "unit_price",
            "line_total",
            "expected_delivery",
            "vendor_eta",
            "source_request_id",
            "created_by",
            "created_at",
            "updated_at",
        ],
    )
    po_records: list[list[object]] = []
    for order in purchase_orders:
        for line in order.lines:
            if line.source_request_id not in event_request_ids:
                continue
            store = stores.get(line.store_number)
            product = products.get(line.product_code)
            po_records.append(
                [
                    order.id,
                    order.po_number,
                    order.status,
                    store.entity_code if store and store.entity_code else "Unassigned",
                    store.region_code if store else "",
                    line.store_number,
                    order.vendor_code,
                    product.model_number if product and product.model_number else line.product_code,
                    line.product_name,
                    line.quantity,
                    line.received_quantity,
                    line.unit_price,
                    line.extended_amount,
                    order.expected_delivery_date,
                    order.vendor_eta,
                    line.source_request_id,
                    order.created_by,
                    _datetime_text(order.created_at),
                    _datetime_text(order.updated_at),
                ]
            )
    po_records.sort(
        key=lambda row: tuple(str(row[index] or "").casefold() for index in (6, 4, 5, 1, 7))
    )
    for record in po_records:
        _append_safe(po_sheet, record)
    _style_table(
        po_sheet,
        [38, 24, 24, 18, 16, 14, 18, 22, 36, 12, 16, 14, 16, 18, 18, 38, 30, 24, 24],
        (12, 13),
    )

    output = BytesIO()
    workbook.save(output)
    return event, output.getvalue()


def archive_event_order_backup(
    db: Session, event_id: str, actor: str
) -> EventOrderBackupArtifact | None:
    existing = db.scalar(
        select(EventOrderBackupArtifact).where(EventOrderBackupArtifact.event_id == event_id)
    )
    if existing is not None:
        return existing
    export = export_event_order_backup(db, event_id)
    if export is None:
        return None
    event, content = export
    artifact = EventOrderBackupArtifact(
        event_id=event.id,
        filename=event_order_backup_filename(event),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=content,
        size_bytes=len(content),
        sha256=sha256(content).hexdigest(),
        created_by=actor,
    )
    db.add(artifact)
    db.flush()
    return artifact


def get_archived_event_order_backup(db: Session, event_id: str) -> EventOrderBackupArtifact | None:
    return db.scalar(
        select(EventOrderBackupArtifact).where(EventOrderBackupArtifact.event_id == event_id)
    )
