import csv
import re
import warnings
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any

# PyMuPDF 1.26.3 emits harmless CPython 3.12 deprecation warnings while its
# native Swig types initialize. Scope the suppression to this dependency only.
with warnings.catch_warnings():
    warnings.simplefilter("ignore", DeprecationWarning)
    import pymupdf  # noqa: E402
from openpyxl import load_workbook
from sqlalchemy import case, func, select, tuple_
from sqlalchemy.orm import Session

from app.auth.permissions import user_has_permission
from app.models.catalog import CatalogVendor
from app.models.event_management import (
    EventMembership,
    EventStaffTask,
    EventVendorBooth,
    ManagedEvent,
    ManagedSubEvent,
    VendorHallAuditLog,
    VendorHallBooth,
    VendorHallBoothCheckin,
    VendorHallEvent,
    VendorHallException,
    VendorHallFloorMap,
    VendorHallInventoryImport,
    VendorHallInventoryItem,
    VendorHallItemAttachment,
    VendorHallItemCheckin,
)
from app.models.identity import Role, User
from app.models.notification import NotificationEvent
from app.schemas.vendor_hall import (
    VendorHallBoothCheckinResponse,
    VendorHallBoothCheckinWrite,
    VendorHallBoothMapPositionWrite,
    VendorHallBoothResponse,
    VendorHallBoothStaffAssignmentWrite,
    VendorHallEventResponse,
    VendorHallEventWrite,
    VendorHallFloorMapResponse,
    VendorHallFloorMapStatusResponse,
    VendorHallFloorMapWrite,
    VendorHallInventoryImportResponse,
    VendorHallInventoryItemResponse,
    VendorHallInventoryItemWrite,
    VendorHallInventorySplitWrite,
    VendorHallItemAttachmentResponse,
    VendorHallItemCheckinResponse,
    VendorHallItemCheckinWrite,
    VendorHallSummaryResponse,
)
from app.services.event_access_service import event_operations_are_locked
from app.services.spreadsheet_security import spreadsheet_safe_cell
from app.services.upload_validation import content_matches_declared_type


class VendorHallError(ValueError):
    pass


class VendorHallAccessError(PermissionError):
    pass


def _assert_event_open(db: Session, event_id: str) -> None:
    if event_operations_are_locked(db, event_id):
        raise VendorHallError(
            "Vendor hall is locked because the event is cancelled or settlement is closed"
        )


FULL_HALL_PERMISSIONS = {
    "vendor_hall.manage",
    "vendor_hall.staff.checkin",
    "vendor_hall.export",
    "vendor_hall.map.manage",
}
EXCEPTION_ITEM_STATUSES = {"damaged", "not_in_booth", "quantity_mismatch"}
VENDOR_HALL_EXPORT_REPORTS = {
    "full-inventory",
    "available-for-sale",
    "damaged-items",
    "missing-items",
    "vendor-summary",
    "booth-completion",
    "staff-checkin-log",
}


def _notify_booth_inspection_assignment(
    db: Session,
    booth: VendorHallBooth,
    membership_id: str,
    actor: str,
) -> None:
    membership = db.get(EventMembership, membership_id)
    recipient = db.get(User, membership.user_id) if membership else None
    if recipient is None:
        return
    db.add(
        NotificationEvent(
            template_code="VENDOR_HALL_INSPECTION_READY",
            workflow_code="EVENTS",
            event_type="vendor_hall.inspection_ready",
            entity_type="vendor_hall_booth",
            entity_id=booth.id,
            actor=actor,
            channel="in_app",
            recipient_strategy="static_recipients",
            resolved_recipients=[recipient.email],
            subject=f"Inspection ready: {booth.booth_name}",
            body=(
                f"Vendor booth {booth.booth_number or 'TBD'} ({booth.booth_name}) "
                "is ready for inventory inspection."
            ),
            status="queued",
        )
    )


MAX_FLOOR_MAP_BYTES = 20 * 1024 * 1024
MAX_FLOOR_MAP_PAGES = 25
MAX_FLOOR_MAP_TEXT_FRAGMENTS = 50_000
MAX_INVENTORY_IMPORT_BYTES = 2 * 1024 * 1024
MAX_INVENTORY_IMPORT_ROWS = 10_000
MAX_INVENTORY_ATTACHMENT_BYTES = 8 * 1024 * 1024
ALLOWED_ATTACHMENT_TYPES = {"photo", "spec_sheet", "other"}
ALLOWED_ATTACHMENT_CONTENT_TYPES = {
    "image/jpeg",
    "image/png",
    "image/webp",
    "application/pdf",
}


def _event_response(db: Session, hall: VendorHallEvent) -> VendorHallEventResponse:
    event = db.get(ManagedEvent, hall.event_id)
    sub_event = db.get(ManagedSubEvent, hall.sub_event_id) if hall.sub_event_id else None
    return VendorHallEventResponse(
        id=hall.id,
        event_id=hall.event_id,
        event_name=event.name if event else "Event",
        sub_event_id=hall.sub_event_id,
        sub_event_name=sub_event.name if sub_event else None,
        status=hall.status,
        opens_at=hall.opens_at,
        vendor_submission_deadline=hall.vendor_submission_deadline,
        staff_checkin_opens_at=hall.staff_checkin_opens_at,
        staff_checkin_deadline=hall.staff_checkin_deadline,
        allow_vendor_edits_after_submission=hall.allow_vendor_edits_after_submission,
        require_staff_checkin=hall.require_staff_checkin,
        created_at=hall.created_at,
        updated_at=hall.updated_at,
    )


def _booth_response(db: Session, booth: VendorHallBooth) -> VendorHallBoothResponse:
    return _booth_responses(db, [booth])[0]


def _booth_responses(
    db: Session,
    booths: list[VendorHallBooth],
) -> list[VendorHallBoothResponse]:
    """Build booth responses with a fixed number of queries, regardless of hall size."""
    if not booths:
        return []

    booth_ids = {booth.id for booth in booths}
    event_names = dict(
        db.execute(
            select(ManagedEvent.id, ManagedEvent.name).where(
                ManagedEvent.id.in_({booth.event_id for booth in booths})
            )
        ).all()
    )
    vendor_names = dict(
        db.execute(
            select(CatalogVendor.vendor_code, CatalogVendor.name).where(
                CatalogVendor.vendor_code.in_({booth.vendor_code for booth in booths})
            )
        ).all()
    )
    assigned_staff_ids = {
        booth.assigned_staff_membership_id for booth in booths if booth.assigned_staff_membership_id
    }
    staff_names = (
        dict(
            db.execute(
                select(EventMembership.id, User.display_name)
                .join(User, User.id == EventMembership.user_id)
                .where(EventMembership.id.in_(assigned_staff_ids))
            ).all()
        )
        if assigned_staff_ids
        else {}
    )
    inventory_counts = {
        booth_id: (inventory_count, available_count or 0)
        for booth_id, inventory_count, available_count in db.execute(
            select(
                VendorHallInventoryItem.vendor_hall_booth_id,
                func.count(),
                func.sum(case((VendorHallInventoryItem.available_for_sale.is_(True), 1), else_=0)),
            )
            .where(VendorHallInventoryItem.vendor_hall_booth_id.in_(booth_ids))
            .group_by(VendorHallInventoryItem.vendor_hall_booth_id)
        ).all()
    }
    exception_counts = dict(
        db.execute(
            # Multiple validation attempts can create several exception audit
            # rows for one item. The booth summary represents affected items,
            # so count each inventory item only once.
            select(
                VendorHallException.vendor_hall_booth_id,
                func.count(func.distinct(VendorHallException.inventory_item_id)),
            )
            .where(
                VendorHallException.vendor_hall_booth_id.in_(booth_ids),
                VendorHallException.status == "open",
            )
            .group_by(VendorHallException.vendor_hall_booth_id)
        ).all()
    )
    return [
        _booth_response_from_context(
            booth,
            event_name=event_names.get(booth.event_id, "Event"),
            vendor_name=vendor_names.get(booth.vendor_code),
            inventory_count=inventory_counts.get(booth.id, (0, 0))[0],
            available_count=inventory_counts.get(booth.id, (0, 0))[1],
            exceptions_count=exception_counts.get(booth.id, 0),
            assigned_staff_display_name=staff_names.get(booth.assigned_staff_membership_id),
        )
        for booth in booths
    ]


def _booth_response_from_context(
    booth: VendorHallBooth,
    *,
    event_name: str,
    vendor_name: str | None,
    inventory_count: int,
    available_count: int,
    exceptions_count: int,
    assigned_staff_display_name: str | None = None,
) -> VendorHallBoothResponse:
    return VendorHallBoothResponse(
        id=booth.id,
        vendor_hall_event_id=booth.vendor_hall_event_id,
        event_vendor_booth_id=booth.event_vendor_booth_id,
        assigned_staff_membership_id=booth.assigned_staff_membership_id,
        assigned_staff_display_name=assigned_staff_display_name,
        event_id=booth.event_id,
        event_name=event_name,
        vendor_code=booth.vendor_code,
        vendor_name=vendor_name,
        booth_number=booth.booth_number,
        booth_name=booth.booth_name,
        floor_map_zone=booth.floor_map_zone,
        map_x=booth.map_x,
        map_y=booth.map_y,
        map_width=booth.map_width,
        map_height=booth.map_height,
        map_manually_adjusted=booth.map_manually_adjusted,
        status=booth.status,
        submitted_at=booth.submitted_at,
        checkin_started_at=booth.checkin_started_at,
        checkin_completed_at=booth.checkin_completed_at,
        exceptions_count=exceptions_count,
        available_for_sale_count=available_count,
        inventory_count=inventory_count,
        updated_at=booth.updated_at,
    )


def _inventory_item_response(
    item: VendorHallInventoryItem,
    attachments: list[VendorHallItemAttachment] | None = None,
    *,
    validated: bool = False,
) -> VendorHallInventoryItemResponse:
    return VendorHallInventoryItemResponse(
        id=item.id,
        vendor_hall_booth_id=item.vendor_hall_booth_id,
        event_id=item.event_id,
        vendor_code=item.vendor_code,
        model_number=item.model_number,
        serial_number=item.serial_number,
        item_name=item.item_name,
        description=item.description,
        quantity_expected=item.quantity_expected,
        quantity_checked_in=item.quantity_checked_in,
        unit_price=item.unit_price,
        currency=item.currency,
        condition=item.condition,
        status=item.status,
        available_for_sale=item.available_for_sale,
        sell_to_buddys_price=item.sell_to_buddys_price,
        notes=item.notes,
        vendor_notes=item.vendor_notes,
        staff_notes=item.staff_notes,
        validated=validated,
        attachments=[_attachment_response(attachment) for attachment in attachments or []],
        created_at=item.created_at,
        updated_at=item.updated_at,
    )


def _item_checkin_response(checkin: VendorHallItemCheckin) -> VendorHallItemCheckinResponse:
    return VendorHallItemCheckinResponse(
        id=checkin.id,
        inventory_item_id=checkin.inventory_item_id,
        vendor_hall_booth_id=checkin.vendor_hall_booth_id,
        status=checkin.status,
        quantity_checked=checkin.quantity_checked,
        condition=checkin.condition,
        damage_notes=checkin.damage_notes,
        exception_notes=checkin.exception_notes,
        checked_by=checkin.checked_by,
        checked_at=checkin.checked_at,
    )


def _booth_checkin_response(checkin: VendorHallBoothCheckin) -> VendorHallBoothCheckinResponse:
    return VendorHallBoothCheckinResponse(
        id=checkin.id,
        vendor_hall_booth_id=checkin.vendor_hall_booth_id,
        status=checkin.status,
        started_by=checkin.started_by,
        started_at=checkin.started_at,
        completed_by=checkin.completed_by,
        completed_at=checkin.completed_at,
        completion_percentage=checkin.completion_percentage,
        items_expected=checkin.items_expected,
        items_checked=checkin.items_checked,
        exceptions_count=checkin.exceptions_count,
        notes=checkin.notes,
    )


def _inventory_import_response(
    inventory_import: VendorHallInventoryImport,
) -> VendorHallInventoryImportResponse:
    return VendorHallInventoryImportResponse(
        id=inventory_import.id,
        vendor_hall_booth_id=inventory_import.vendor_hall_booth_id,
        filename=inventory_import.filename,
        content_type=inventory_import.content_type,
        row_count=inventory_import.row_count,
        accepted_count=inventory_import.accepted_count,
        rejected_count=inventory_import.rejected_count,
        status=inventory_import.status,
        error_summary=inventory_import.error_summary,
        uploaded_by=inventory_import.uploaded_by,
        uploaded_at=inventory_import.uploaded_at,
        completed_at=inventory_import.completed_at,
    )


def _attachment_response(
    attachment: VendorHallItemAttachment,
) -> VendorHallItemAttachmentResponse:
    return VendorHallItemAttachmentResponse(
        id=attachment.id,
        inventory_item_id=attachment.inventory_item_id,
        attachment_type=attachment.attachment_type,
        filename=attachment.filename,
        content_type=attachment.content_type,
        uploaded_by=attachment.uploaded_by,
        uploaded_at=attachment.uploaded_at,
    )


def _floor_map_response(floor_map: VendorHallFloorMap) -> VendorHallFloorMapResponse:
    return VendorHallFloorMapResponse(
        id=floor_map.id,
        vendor_hall_event_id=floor_map.vendor_hall_event_id,
        name=floor_map.name,
        has_image=floor_map.image_content is not None,
        layout_json=floor_map.layout_json or {},
        uploaded_by=floor_map.uploaded_by,
        uploaded_at=floor_map.uploaded_at,
        is_active=floor_map.is_active,
    )


def _get_booth(db: Session, booth_id: str) -> VendorHallBooth | None:
    return db.get(VendorHallBooth, booth_id)


def _user_has_full_hall_access(user: User) -> bool:
    return any(user_has_permission(user, permission) for permission in FULL_HALL_PERMISSIONS)


def _is_elevated_staff(user: User) -> bool:
    role_codes = {role.workflow_code for role in user.roles if role.workflow_code}
    return bool(role_codes & {"ADMIN", "SYSTEM_ADMIN"})


def _is_scoped_event_staff(db: Session, user: User, event_id: str) -> bool:
    role_codes = {role.workflow_code for role in user.roles if role.workflow_code}
    if _is_elevated_staff(user):
        return False
    return (
        "EVENT_STAFF" in role_codes
        or db.scalar(
            select(EventMembership.id).where(
                EventMembership.event_id == event_id,
                EventMembership.user_id == user.id,
                EventMembership.membership_type == "staff",
                EventMembership.is_active.is_(True),
            )
        )
        is not None
    )


def _user_is_booth_vendor(db: Session, user: User, booth: VendorHallBooth) -> bool:
    if user.vendor_code and user.vendor_code == booth.vendor_code:
        return True
    return (
        db.scalar(
            select(EventMembership.id).where(
                EventMembership.event_id == booth.event_id,
                EventMembership.user_id == user.id,
                EventMembership.vendor_code == booth.vendor_code,
                EventMembership.membership_type == "vendor",
                EventMembership.is_active.is_(True),
            )
        )
        is not None
    )


def _assert_can_view_booth(db: Session, user: User, booth: VendorHallBooth) -> None:
    if _user_has_full_hall_access(user) or _user_is_booth_vendor(db, user, booth):
        return
    raise VendorHallAccessError("You do not have access to this vendor hall booth")


def _assert_can_manage_vendor_inventory(db: Session, user: User, booth: VendorHallBooth) -> None:
    if _is_scoped_event_staff(db, user, booth.event_id):
        raise VendorHallAccessError("Onsite staff can only validate assigned booth inventory")
    if user_has_permission(user, "vendor_hall.manage"):
        return
    if user_has_permission(user, "vendor_hall.vendor.manage") and _user_is_booth_vendor(
        db, user, booth
    ):
        return
    raise VendorHallAccessError("You cannot manage inventory for this booth")


def _assert_can_checkin_booth(user: User) -> None:
    if user_has_permission(user, "vendor_hall.manage") or user_has_permission(
        user, "vendor_hall.staff.checkin"
    ):
        return
    raise VendorHallAccessError("You cannot perform vendor hall check-in")


def _assert_booth_editable(booth: VendorHallBooth) -> None:
    """Completed booths are immutable; reopening requires an admin workflow."""
    # A booth completed with exceptions is intentionally stored as
    # ``exceptions_present``; the completion timestamp is therefore the
    # authoritative lock signal for both clean and exception completions.
    if booth.checkin_completed_at is not None or booth.status in {
        "fully_checked_in",
        "admin_reviewed",
    }:
        raise VendorHallAccessError("This booth is complete and locked for editing")


def configure_vendor_hall(
    db: Session,
    event_id: str,
    payload: VendorHallEventWrite,
    actor: str,
) -> VendorHallEventResponse | None:
    event = db.get(ManagedEvent, event_id)
    if event is None:
        return None
    _assert_event_open(db, event_id)
    if payload.sub_event_id:
        sub_event = db.get(ManagedSubEvent, payload.sub_event_id)
        if sub_event is None or sub_event.event_id != event_id:
            raise VendorHallError("Sub-event does not belong to this event")
    if payload.status == "closed":
        incomplete = (
            db.scalar(
                select(func.count(VendorHallBooth.id)).where(
                    VendorHallBooth.event_id == event_id,
                    VendorHallBooth.status.not_in(("fully_checked_in", "admin_reviewed", "closed")),
                )
            )
            or 0
        )
        if incomplete:
            raise VendorHallError(
                f"Vendor hall cannot close until all booths are complete ({incomplete} remaining)"
            )
    hall = db.scalar(
        select(VendorHallEvent).where(
            VendorHallEvent.event_id == event_id,
            VendorHallEvent.sub_event_id == payload.sub_event_id,
        )
    )
    if hall is None:
        hall = VendorHallEvent(event_id=event_id, created_by=actor)
        db.add(hall)
    for field, value in payload.model_dump().items():
        setattr(hall, field, value)
    db.commit()
    db.refresh(hall)
    _audit(db, event_id, hall.id, None, "vendor_hall.configured", actor, payload.model_dump())
    return _event_response(db, hall)


def force_close_vendor_hall(
    db: Session,
    event_id: str,
    actor: str,
) -> VendorHallEventResponse | None:
    event = db.get(ManagedEvent, event_id)
    if event is None:
        return None
    hall = get_vendor_hall(db, event_id)
    if hall is None:
        raise VendorHallError("Vendor hall is not configured")
    hall.status = "closed"
    db.commit()
    db.refresh(hall)
    _audit(
        db,
        event_id,
        hall.id,
        None,
        "vendor_hall.force_closed",
        actor,
        {"override": True},
    )
    return _event_response(db, hall)


def get_vendor_hall(db: Session, event_id: str) -> VendorHallEvent | None:
    return db.scalar(select(VendorHallEvent).where(VendorHallEvent.event_id == event_id))


def sync_vendor_hall_booths(
    db: Session, event_id: str, actor: str
) -> list[VendorHallBoothResponse] | None:
    _assert_event_open(db, event_id)
    hall = get_vendor_hall(db, event_id)
    if hall is None:
        event = db.get(ManagedEvent, event_id)
        if event is None:
            return None
        hall = VendorHallEvent(event_id=event_id, created_by=actor)
        db.add(hall)
        db.flush()
    event_booths = db.scalars(
        select(EventVendorBooth)
        .where(EventVendorBooth.event_id == event_id)
        .order_by(EventVendorBooth.booth_name)
    ).all()
    existing_by_event_booth = {
        booth.event_vendor_booth_id: booth
        for booth in db.scalars(
            select(VendorHallBooth).where(VendorHallBooth.event_id == event_id)
        ).all()
        if booth.event_vendor_booth_id
    }
    for event_booth in event_booths:
        booth = existing_by_event_booth.get(event_booth.id)
        if booth is None:
            booth = VendorHallBooth(
                vendor_hall_event_id=hall.id,
                event_vendor_booth_id=event_booth.id,
                event_id=event_id,
                vendor_code=event_booth.vendor_code,
                booth_number=event_booth.booth_number or "",
                booth_name=event_booth.booth_name,
            )
            db.add(booth)
        else:
            booth.vendor_hall_event_id = hall.id
            booth.vendor_code = event_booth.vendor_code
            booth.booth_number = event_booth.booth_number or ""
            booth.booth_name = event_booth.booth_name
    db.flush()
    active_floor_map = db.scalar(
        select(VendorHallFloorMap)
        .where(
            VendorHallFloorMap.vendor_hall_event_id == hall.id,
            VendorHallFloorMap.is_active.is_(True),
            VendorHallFloorMap.image_content.is_not(None),
        )
        .order_by(VendorHallFloorMap.uploaded_at.desc())
    )
    if (
        active_floor_map is not None
        and active_floor_map.image_content is not None
        and active_floor_map.image_content_type in {"application/pdf", "application/x-pdf"}
    ):
        active_floor_map.layout_json = _scan_floor_map_pdf_against_booths(
            db,
            event_id,
            active_floor_map.image_content,
        )
    db.commit()
    _audit(
        db,
        event_id,
        hall.id,
        None,
        "vendor_hall.booths_synced",
        actor,
        {"booth_count": len(event_booths)},
    )
    return list_vendor_hall_booths(db, event_id)


def list_vendor_hall_booths(db: Session, event_id: str) -> list[VendorHallBoothResponse] | None:
    if db.get(ManagedEvent, event_id) is None:
        return None
    booths = db.scalars(
        select(VendorHallBooth)
        .where(VendorHallBooth.event_id == event_id)
        .order_by(VendorHallBooth.booth_number, VendorHallBooth.booth_name)
    ).all()
    return _booth_responses(db, booths)


def my_vendor_hall_booths(db: Session, user: User) -> list[VendorHallBoothResponse]:
    memberships = db.scalars(
        select(EventMembership).where(
            EventMembership.user_id == user.id,
            EventMembership.membership_type == "vendor",
            EventMembership.is_active.is_(True),
        )
    ).all()
    vendor_scopes = {
        (membership.event_id, vendor_code)
        for membership in memberships
        for vendor_code in (
            set(membership.vendor_codes or [])
            | ({membership.vendor_code} if membership.vendor_code else set())
        )
    }
    booths: list[VendorHallBooth] = (
        list(
            db.scalars(
                select(VendorHallBooth).where(
                    tuple_(VendorHallBooth.event_id, VendorHallBooth.vendor_code).in_(vendor_scopes)
                )
            ).all()
        )
        if vendor_scopes
        else []
    )
    staff_membership_ids = {
        membership.id
        for membership in db.scalars(
            select(EventMembership).where(
                EventMembership.user_id == user.id,
                EventMembership.membership_type == "staff",
                EventMembership.is_active.is_(True),
            )
        ).all()
    }
    if staff_membership_ids:
        booths.extend(
            db.scalars(
                select(VendorHallBooth).where(
                    VendorHallBooth.assigned_staff_membership_id.in_(staff_membership_ids)
                )
            ).all()
        )
    unique = {booth.id: booth for booth in booths}
    sorted_booths = sorted(unique.values(), key=lambda item: (item.booth_number, item.booth_name))
    return _booth_responses(db, sorted_booths)


def vendor_hall_summary(db: Session, event_id: str) -> VendorHallSummaryResponse | None:
    event = db.get(ManagedEvent, event_id)
    if event is None:
        return None
    hall = get_vendor_hall(db, event_id)
    if hall is None:
        return VendorHallSummaryResponse(
            event_id=event_id,
            event_name=event.name,
            vendor_hall_event_id=None,
            booth_total=0,
            inventory_submitted=0,
            checkin_in_progress=0,
            fully_checked_in=0,
            exceptions_present=0,
            closed=0,
            completion_percentage=Decimal("0.00"),
            inventory_item_total=0,
            inventory_items_checked=0,
            inventory_completion_percentage=Decimal("0.00"),
            closeout_ready=False,
            vendors_not_submitted=[],
        )
    booths = db.scalars(select(VendorHallBooth).where(VendorHallBooth.event_id == event_id)).all()
    counts = {
        status: 0
        for status in (
            "inventory_submitted",
            "ready_for_inspection",
            "checkin_in_progress",
            "fully_checked_in",
            "exceptions_present",
            "closed",
        )
    }
    for booth in booths:
        if booth.status in counts:
            counts[booth.status] += 1
    complete = counts["fully_checked_in"] + counts["closed"]
    completion = Decimal("0.00")
    if booths:
        completion = (Decimal(complete) / Decimal(len(booths)) * Decimal(100)).quantize(
            Decimal("0.01")
        )
    item_total = (
        db.scalar(
            select(func.count(VendorHallInventoryItem.id)).where(
                VendorHallInventoryItem.event_id == event_id
            )
        )
        or 0
    )
    item_checked = (
        db.scalar(
            select(func.count(VendorHallInventoryItem.id)).where(
                VendorHallInventoryItem.event_id == event_id,
                VendorHallInventoryItem.status.in_(("checked_in", "purchased", "removed")),
            )
        )
        or 0
    )
    item_completion = (
        (Decimal(item_checked) / Decimal(item_total) * Decimal(100)).quantize(Decimal("0.01"))
        if item_total
        else Decimal("0.00")
    )
    return VendorHallSummaryResponse(
        event_id=event_id,
        event_name=event.name,
        vendor_hall_event_id=hall.id,
        booth_total=len(booths),
        inventory_submitted=counts["inventory_submitted"],
        checkin_in_progress=counts["checkin_in_progress"] + counts["ready_for_inspection"],
        fully_checked_in=counts["fully_checked_in"],
        exceptions_present=counts["exceptions_present"],
        closed=counts["closed"],
        completion_percentage=completion,
        inventory_item_total=item_total,
        inventory_items_checked=item_checked,
        inventory_completion_percentage=item_completion,
        closeout_ready=bool(booths) and complete == len(booths) and item_checked == item_total,
        vendors_not_submitted=_booth_responses(
            db, [booth for booth in booths if booth.status == "draft"]
        ),
    )


def list_booth_inventory(
    db: Session, booth_id: str, user: User
) -> list[VendorHallInventoryItemResponse] | None:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return None
    _assert_can_view_booth(db, user, booth)
    items = list(
        db.scalars(
            select(VendorHallInventoryItem)
            .where(VendorHallInventoryItem.vendor_hall_booth_id == booth_id)
            .order_by(VendorHallInventoryItem.item_name, VendorHallInventoryItem.model_number)
        ).all()
    )
    attachments_by_item: dict[str, list[VendorHallItemAttachment]] = {item.id: [] for item in items}
    validated_item_ids: set[str] = set()
    if items:
        validated_item_ids = set(
            db.scalars(
                select(VendorHallItemCheckin.inventory_item_id).where(
                    VendorHallItemCheckin.inventory_item_id.in_([item.id for item in items])
                )
            ).all()
        )
        attachments = db.scalars(
            select(VendorHallItemAttachment)
            .where(VendorHallItemAttachment.inventory_item_id.in_([item.id for item in items]))
            .order_by(VendorHallItemAttachment.uploaded_at, VendorHallItemAttachment.filename)
        ).all()
        for attachment in attachments:
            attachments_by_item[attachment.inventory_item_id].append(attachment)
    return [
        _inventory_item_response(
            item,
            attachments_by_item[item.id],
            validated=item.id in validated_item_ids,
        )
        for item in items
    ]


def inventory_item_attachment_content(
    db: Session,
    booth_id: str,
    item_id: str,
    attachment_id: str,
    user: User,
) -> tuple[str, str, bytes] | None:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return None
    _assert_can_view_booth(db, user, booth)
    item = db.get(VendorHallInventoryItem, item_id)
    if item is None or item.vendor_hall_booth_id != booth_id:
        return None
    attachment = db.get(VendorHallItemAttachment, attachment_id)
    if attachment is None or attachment.inventory_item_id != item_id:
        return None
    return attachment.filename, attachment.content_type, attachment.content


def remove_inventory_item_attachment(
    db: Session,
    booth_id: str,
    item_id: str,
    attachment_id: str,
    user: User,
) -> bool:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return False
    _assert_event_open(db, booth.event_id)
    _assert_booth_editable(booth)
    _assert_can_manage_vendor_inventory(db, user, booth)
    item = db.get(VendorHallInventoryItem, item_id)
    if item is None or item.vendor_hall_booth_id != booth_id:
        return False
    attachment = db.get(VendorHallItemAttachment, attachment_id)
    if attachment is None or attachment.inventory_item_id != item_id:
        return False
    filename = attachment.filename
    db.delete(attachment)
    _audit(
        db,
        booth.event_id,
        booth.vendor_hall_event_id,
        booth.id,
        "vendor_hall.inventory_item.attachment_deleted",
        user.email,
        {
            "item_id": item.id,
            "attachment_id": attachment_id,
            "filename": filename,
        },
        inventory_item_id=item.id,
    )
    return True


def create_booth_inventory_item(
    db: Session,
    booth_id: str,
    payload: VendorHallInventoryItemWrite,
    user: User,
) -> VendorHallInventoryItemResponse | None:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return None
    _assert_event_open(db, booth.event_id)
    _assert_booth_editable(booth)
    _assert_can_manage_vendor_inventory(db, user, booth)
    item = VendorHallInventoryItem(
        vendor_hall_booth_id=booth.id,
        event_id=booth.event_id,
        vendor_code=booth.vendor_code,
        created_by=user.email,
        **payload.model_dump(),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    _audit(
        db,
        booth.event_id,
        booth.vendor_hall_event_id,
        booth.id,
        "vendor_hall.inventory_item.created",
        user.email,
        {"item_id": item.id, "item_name": item.item_name},
        inventory_item_id=item.id,
    )
    return _inventory_item_response(item)


def update_booth_inventory_item(
    db: Session,
    booth_id: str,
    item_id: str,
    payload: VendorHallInventoryItemWrite,
    user: User,
) -> VendorHallInventoryItemResponse | None:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return None
    _assert_event_open(db, booth.event_id)
    _assert_booth_editable(booth)
    _assert_can_manage_vendor_inventory(db, user, booth)
    item = db.get(VendorHallInventoryItem, item_id)
    if item is None or item.vendor_hall_booth_id != booth_id:
        return None
    for field, value in payload.model_dump().items():
        setattr(item, field, value)
    db.commit()
    db.refresh(item)
    _audit(
        db,
        booth.event_id,
        booth.vendor_hall_event_id,
        booth.id,
        "vendor_hall.inventory_item.updated",
        user.email,
        {"item_id": item.id, "item_name": item.item_name},
        inventory_item_id=item.id,
    )
    return _inventory_item_response(item)


def update_booth_inventory_item_staff(
    db: Session,
    booth_id: str,
    item_id: str,
    payload,
    user: User,
) -> VendorHallInventoryItemResponse | None:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return None
    _assert_event_open(db, booth.event_id)
    _assert_booth_editable(booth)
    _assert_can_checkin_booth(user)
    item = db.get(VendorHallInventoryItem, item_id)
    if item is None or item.vendor_hall_booth_id != booth_id:
        return None
    item.quantity_checked_in = payload.quantity_checked_in
    item.condition = payload.condition
    item.staff_notes = payload.staff_notes
    db.commit()
    db.refresh(item)
    return _inventory_item_response(item)


def split_booth_inventory_item(
    db: Session,
    booth_id: str,
    item_id: str,
    payload: VendorHallInventorySplitWrite,
    user: User,
) -> VendorHallInventoryItemResponse | None:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return None
    _assert_event_open(db, booth.event_id)
    _assert_booth_editable(booth)
    _assert_can_manage_vendor_inventory(db, user, booth)
    item = db.get(VendorHallInventoryItem, item_id)
    if item is None or item.vendor_hall_booth_id != booth_id:
        return None
    if payload.split_quantity >= item.quantity_expected:
        raise VendorHallError("Split quantity must be less than the original quantity")
    item.quantity_expected -= payload.split_quantity
    item.quantity_checked_in = min(item.quantity_checked_in, item.quantity_expected)
    split_item = VendorHallInventoryItem(
        vendor_hall_booth_id=booth.id,
        event_id=booth.event_id,
        vendor_code=booth.vendor_code,
        created_by=user.email,
        model_number=item.model_number,
        serial_number=None,
        item_name=item.item_name,
        description=item.description,
        quantity_expected=payload.split_quantity,
        unit_price=item.unit_price,
        currency=item.currency,
        condition="damaged" if payload.status == "damaged" else item.condition,
        status=payload.status,
        available_for_sale=item.available_for_sale,
        sell_to_buddys_price=item.sell_to_buddys_price,
        notes=payload.notes or item.notes,
        vendor_notes=payload.notes or item.vendor_notes,
    )
    db.add(split_item)
    db.commit()
    db.refresh(split_item)
    _audit(
        db,
        booth.event_id,
        booth.vendor_hall_event_id,
        booth.id,
        "vendor_hall.inventory_item.split",
        user.email,
        {
            "item_id": item.id,
            "split_item_id": split_item.id,
            "split_quantity": payload.split_quantity,
        },
        inventory_item_id=item.id,
    )
    return _inventory_item_response(split_item)


def submit_booth_inventory(
    db: Session, booth_id: str, user: User
) -> VendorHallBoothResponse | None:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return None
    _assert_event_open(db, booth.event_id)
    _assert_booth_editable(booth)
    _assert_can_manage_vendor_inventory(db, user, booth)
    booth.status = "inventory_submitted"
    booth.submitted_at = datetime.now(UTC)
    booth.submitted_by = user.email
    db.commit()
    db.refresh(booth)
    _audit(
        db,
        booth.event_id,
        booth.vendor_hall_event_id,
        booth.id,
        "vendor_hall.inventory.submitted",
        user.email,
        {"booth_id": booth.id},
    )
    return _booth_response(db, booth)


def mark_booth_ready_for_inspection(
    db: Session, booth_id: str, user: User
) -> VendorHallBoothResponse | None:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return None
    _assert_event_open(db, booth.event_id)
    _assert_booth_editable(booth)
    _assert_can_manage_vendor_inventory(db, user, booth)
    item_count = _booth_item_count(db, booth.id)
    if item_count == 0:
        raise VendorHallError("Booth inventory must contain at least one item before inspection")
    preliminary_complete = (
        db.scalar(
            select(func.count())
            .select_from(VendorHallInventoryItem)
            .where(
                VendorHallInventoryItem.vendor_hall_booth_id == booth.id,
                VendorHallInventoryItem.status.not_in(
                    ("checked_in", "damaged", "not_in_booth", "removed")
                ),
            )
        )
        or 0
    )
    if preliminary_complete:
        raise VendorHallError(
            "Every inventory item must be marked checked in, damaged, not in booth, "
            "or removed before inspection"
        )
    booth.status = "ready_for_inspection"
    booth.submitted_at = datetime.now(UTC)
    booth.submitted_by = user.email
    if booth.assigned_staff_membership_id:
        existing_task = db.scalar(
            select(EventStaffTask).where(
                EventStaffTask.vendor_hall_booth_id == booth.id,
                EventStaffTask.status.in_(("open", "in_progress", "blocked")),
            )
        )
        if existing_task is None:
            hall = db.get(VendorHallEvent, booth.vendor_hall_event_id)
            db.add(
                EventStaffTask(
                    event_id=booth.event_id,
                    sub_event_id=hall.sub_event_id if hall else None,
                    vendor_hall_booth_id=booth.id,
                    assigned_membership_id=booth.assigned_staff_membership_id,
                    title=f"Validate inventory: {booth.booth_name}",
                    description=(
                        f"Validate the vendor inventory for booth {booth.booth_number or 'TBD'}, "
                        "record discrepancies, and complete the booth inspection."
                    ),
                    priority="high",
                    status="open",
                    task_phase="pre_event",
                    created_by=user.email,
                )
            )
            _notify_booth_inspection_assignment(
                db, booth, booth.assigned_staff_membership_id, user.email
            )
    db.commit()
    db.refresh(booth)
    _audit(
        db,
        booth.event_id,
        booth.vendor_hall_event_id,
        booth.id,
        "vendor_hall.booth.ready_for_inspection",
        user.email,
        {"booth_id": booth.id},
    )
    return _booth_response(db, booth)


def start_booth_checkin(
    db: Session,
    booth_id: str,
    payload: VendorHallBoothCheckinWrite,
    user: User,
) -> VendorHallBoothCheckinResponse | None:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return None
    _assert_event_open(db, booth.event_id)
    _assert_booth_editable(booth)
    _assert_can_checkin_booth(user)
    booth.status = "checkin_in_progress"
    booth.checkin_started_at = booth.checkin_started_at or datetime.now(UTC)
    booth.checked_in_by = user.email
    checkin = VendorHallBoothCheckin(
        vendor_hall_booth_id=booth.id,
        status=booth.status,
        started_by=user.email,
        notes=payload.notes,
        items_expected=_booth_item_count(db, booth.id),
        items_checked=_booth_checked_count(db, booth.id),
        exceptions_count=_booth_exception_count(db, booth.id),
    )
    db.add(checkin)
    task = db.scalar(
        select(EventStaffTask).where(
            EventStaffTask.vendor_hall_booth_id == booth.id,
            EventStaffTask.status.in_(("open", "blocked")),
        )
    )
    if task is not None:
        task.status = "in_progress"
    db.commit()
    db.refresh(checkin)
    _audit(
        db,
        booth.event_id,
        booth.vendor_hall_event_id,
        booth.id,
        "vendor_hall.checkin.started",
        user.email,
        {"booth_id": booth.id},
    )
    return _booth_checkin_response(checkin)


def checkin_booth_inventory_item(
    db: Session,
    booth_id: str,
    item_id: str,
    payload: VendorHallItemCheckinWrite,
    user: User,
) -> VendorHallItemCheckinResponse | None:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return None
    _assert_event_open(db, booth.event_id)
    _assert_booth_editable(booth)
    _assert_can_checkin_booth(user)
    item = db.get(VendorHallInventoryItem, item_id)
    if item is None or item.vendor_hall_booth_id != booth_id:
        return None
    status = payload.status
    item.status = status
    item.quantity_checked_in = payload.quantity_checked
    if payload.condition is not None:
        item.condition = payload.condition
    item.staff_notes = payload.staff_notes
    checkin = VendorHallItemCheckin(
        inventory_item_id=item.id,
        vendor_hall_booth_id=booth.id,
        status=status,
        quantity_checked=payload.quantity_checked,
        condition=payload.condition,
        damage_notes=payload.damage_notes,
        exception_notes=payload.exception_notes,
        checked_by=user.email,
    )
    db.add(checkin)
    if status in EXCEPTION_ITEM_STATUSES:
        _create_item_exception(db, booth, item, status, payload, user.email)
        booth.status = "exceptions_present"
    elif booth.status == "draft":
        booth.status = "checkin_in_progress"
    db.commit()
    db.refresh(checkin)
    _audit(
        db,
        booth.event_id,
        booth.vendor_hall_event_id,
        booth.id,
        "vendor_hall.inventory_item.checked",
        user.email,
        {"item_id": item.id, "status": status, "quantity_checked": payload.quantity_checked},
        inventory_item_id=item.id,
    )
    return _item_checkin_response(checkin)


def complete_booth_checkin(
    db: Session,
    booth_id: str,
    payload: VendorHallBoothCheckinWrite,
    user: User,
) -> VendorHallBoothResponse | None:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return None
    _assert_event_open(db, booth.event_id)
    _assert_booth_editable(booth)
    _assert_can_checkin_booth(user)
    expected = _booth_item_count(db, booth.id)
    checked = _booth_checked_count(db, booth.id)
    exceptions = _booth_exception_count(db, booth.id)
    if expected == 0 or checked < expected:
        raise VendorHallError(
            "Every inventory item must be manually validated before completing booth check-in"
        )
    completion = Decimal("0.00")
    if expected:
        completion = (Decimal(checked) / Decimal(expected) * Decimal(100)).quantize(Decimal("0.01"))
    booth.status = "exceptions_present" if exceptions else "fully_checked_in"
    booth.checkin_completed_at = datetime.now(UTC)
    booth.checked_in_by = user.email
    checkin = VendorHallBoothCheckin(
        vendor_hall_booth_id=booth.id,
        status=booth.status,
        started_by=booth.checked_in_by or user.email,
        completed_by=user.email,
        completed_at=booth.checkin_completed_at,
        completion_percentage=completion,
        items_expected=expected,
        items_checked=checked,
        exceptions_count=exceptions,
        notes=payload.notes,
    )
    db.add(checkin)
    task = db.scalar(
        select(EventStaffTask).where(
            EventStaffTask.vendor_hall_booth_id == booth.id,
            EventStaffTask.status.in_(("open", "in_progress", "blocked")),
        )
    )
    if task is not None:
        task.status = "blocked" if exceptions else "done"
    if booth.status == "fully_checked_in":
        remaining = (
            db.scalar(
                select(func.count(VendorHallBooth.id)).where(
                    VendorHallBooth.event_id == booth.event_id,
                    VendorHallBooth.id != booth.id,
                    VendorHallBooth.status.not_in(("fully_checked_in", "admin_reviewed", "closed")),
                )
            )
            or 0
        )
        already_notified = (
            db.scalar(
                select(func.count(NotificationEvent.id)).where(
                    NotificationEvent.entity_type == "vendor_hall_event",
                    NotificationEvent.entity_id == booth.vendor_hall_event_id,
                    NotificationEvent.event_type == "vendor_hall.closeout_ready",
                )
            )
            or 0
        )
        if remaining == 0 and already_notified == 0:
            recipients = db.scalars(
                select(User)
                .join(User.roles)
                .where(
                    User.is_active.is_(True), Role.code.in_(("SYSTEM_ADMIN", "ADMIN", "EXECUTIVE"))
                )
            ).all()
            db.add(
                NotificationEvent(
                    template_code="VENDOR_HALL_CLOSEOUT_READY",
                    workflow_code="EVENTS",
                    event_type="vendor_hall.closeout_ready",
                    entity_type="vendor_hall_event",
                    entity_id=booth.vendor_hall_event_id,
                    actor=user.email,
                    channel="in_app",
                    recipient_strategy="static_recipients",
                    resolved_recipients=[recipient.email for recipient in recipients],
                    subject="Vendor Hall closeout ready",
                    body=(
                        "All vendor booths have completed inspection and are ready "
                        "for event closeout."
                    ),
                    status="queued",
                )
            )
    db.commit()
    db.refresh(booth)
    _audit(
        db,
        booth.event_id,
        booth.vendor_hall_event_id,
        booth.id,
        "vendor_hall.checkin.completed",
        user.email,
        {
            "booth_id": booth.id,
            "status": booth.status,
            "items_expected": expected,
            "items_checked": checked,
            "exceptions_count": exceptions,
        },
    )
    return _booth_response(db, booth)


def import_booth_inventory_csv(
    db: Session,
    booth_id: str,
    filename: str,
    content_type: str,
    content: bytes,
    user: User,
) -> VendorHallInventoryImportResponse | None:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return None
    _assert_event_open(db, booth.event_id)
    _assert_booth_editable(booth)
    _assert_can_manage_vendor_inventory(db, user, booth)
    is_excel = filename.lower().endswith((".xlsx", ".xlsm")) or content_type in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroEnabled.12",
    }
    is_csv = filename.lower().endswith(".csv") or content_type in {
        "text/csv",
        "application/csv",
        "application/vnd.ms-excel",
    }
    if not is_csv and not is_excel:
        raise VendorHallError("Inventory import file must be a CSV")
    if not content:
        raise VendorHallError("Inventory import file cannot be empty")
    if len(content) > MAX_INVENTORY_IMPORT_BYTES:
        raise VendorHallError("Inventory import file must be 2 MB or smaller")
    if is_excel:
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            worksheet = workbook.active
            values = worksheet.iter_rows(values_only=True)
            headers = [str(value).strip() if value is not None else "" for value in next(values)]
            if not any(headers):
                raise VendorHallError("Inventory import spreadsheet must include a header row")
            rows = [
                {
                    headers[index]: value
                    for index, value in enumerate(row)
                    if index < len(headers) and headers[index]
                }
                for row in values
            ]
        except VendorHallError:
            raise
        except Exception as exc:
            raise VendorHallError("Inventory import spreadsheet could not be read") from exc
    else:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise VendorHallError("Inventory import CSV must use UTF-8 encoding") from exc
        reader = csv.DictReader(StringIO(text))
        rows = list(reader)
        if reader.fieldnames is None:
            raise VendorHallError("Inventory import CSV must include a header row")
    if len(rows) > MAX_INVENTORY_IMPORT_ROWS:
        raise VendorHallError(
            f"Inventory import CSV cannot exceed {MAX_INVENTORY_IMPORT_ROWS:,} rows"
        )
    inventory_import = VendorHallInventoryImport(
        vendor_hall_booth_id=booth.id,
        filename=filename[:255],
        content_type=content_type[:128],
        uploaded_by=user.email,
        status="processing",
    )
    db.add(inventory_import)
    db.flush()
    accepted = 0
    rejected: list[str] = []
    row_count = 0
    for row_number, row in enumerate(rows, start=2):
        row_count += 1
        try:
            payload = _inventory_payload_from_import_row(row)
        except ValueError as exc:
            rejected.append(f"row {row_number}: {exc}")
            continue
        db.add(
            VendorHallInventoryItem(
                vendor_hall_booth_id=booth.id,
                event_id=booth.event_id,
                vendor_code=booth.vendor_code,
                source="import",
                source_import_id=inventory_import.id,
                created_by=user.email,
                **payload.model_dump(),
            )
        )
        accepted += 1
    inventory_import.row_count = row_count
    inventory_import.accepted_count = accepted
    inventory_import.rejected_count = len(rejected)
    inventory_import.status = "completed" if not rejected else "completed_with_errors"
    inventory_import.error_summary = "\n".join(rejected[:20]) or None
    inventory_import.completed_at = datetime.now(UTC)
    db.commit()
    db.refresh(inventory_import)
    _audit(
        db,
        booth.event_id,
        booth.vendor_hall_event_id,
        booth.id,
        "vendor_hall.inventory.imported",
        user.email,
        {
            "import_id": inventory_import.id,
            "row_count": row_count,
            "accepted_count": accepted,
            "rejected_count": len(rejected),
        },
    )
    return _inventory_import_response(inventory_import)


def attach_inventory_item_file(
    db: Session,
    booth_id: str,
    item_id: str,
    attachment_type: str,
    filename: str,
    content_type: str,
    content: bytes,
    user: User,
) -> VendorHallItemAttachmentResponse | None:
    booth = _get_booth(db, booth_id)
    if booth is None:
        return None
    _assert_event_open(db, booth.event_id)
    _assert_booth_editable(booth)
    _assert_can_manage_vendor_inventory(db, user, booth)
    item = db.get(VendorHallInventoryItem, item_id)
    if item is None or item.vendor_hall_booth_id != booth_id:
        return None
    if attachment_type not in ALLOWED_ATTACHMENT_TYPES:
        raise VendorHallError("Attachment type must be photo, spec_sheet, or other")
    if content_type not in ALLOWED_ATTACHMENT_CONTENT_TYPES:
        raise VendorHallError("Attachment must be JPEG, PNG, WebP, or PDF")
    if not content or len(content) > MAX_INVENTORY_ATTACHMENT_BYTES:
        raise VendorHallError("Attachment must be between 1 byte and 8 MB")
    if not content_matches_declared_type(content, content_type):
        raise VendorHallError("Attachment content does not match its declared type")
    attachment = VendorHallItemAttachment(
        inventory_item_id=item.id,
        attachment_type=attachment_type,
        filename=filename[:255],
        content_type=content_type,
        content=content,
        uploaded_by=user.email,
    )
    db.add(attachment)
    db.commit()
    db.refresh(attachment)
    _audit(
        db,
        booth.event_id,
        booth.vendor_hall_event_id,
        booth.id,
        "vendor_hall.inventory_item.attachment_uploaded",
        user.email,
        {"item_id": item.id, "attachment_id": attachment.id, "attachment_type": attachment_type},
        inventory_item_id=item.id,
    )
    return _attachment_response(attachment)


def save_vendor_hall_floor_map(
    db: Session,
    event_id: str,
    payload: VendorHallFloorMapWrite,
    user: User,
) -> VendorHallFloorMapResponse | None:
    _assert_event_open(db, event_id)
    hall = get_vendor_hall(db, event_id)
    if hall is None:
        if db.get(ManagedEvent, event_id) is None:
            return None
        hall = VendorHallEvent(event_id=event_id, created_by=user.email)
        db.add(hall)
        db.flush()
    floor_map = db.scalar(
        select(VendorHallFloorMap).where(
            VendorHallFloorMap.vendor_hall_event_id == hall.id,
            VendorHallFloorMap.is_active.is_(True),
        )
    )
    if floor_map is None:
        floor_map = VendorHallFloorMap(
            vendor_hall_event_id=hall.id,
            uploaded_by=user.email,
            name=payload.name,
        )
        db.add(floor_map)
    floor_map.name = payload.name
    floor_map.layout_json = payload.layout_json
    floor_map.is_active = payload.is_active
    floor_map.uploaded_by = user.email
    db.commit()
    db.refresh(floor_map)
    _audit(
        db,
        event_id,
        hall.id,
        None,
        "vendor_hall.floor_map.saved",
        user.email,
        {"floor_map_id": floor_map.id, "name": floor_map.name},
    )
    return _floor_map_response(floor_map)


def _floor_map_hall(db: Session, event_id: str, user: User) -> VendorHallEvent | None:
    hall = get_vendor_hall(db, event_id)
    if hall is None:
        if db.get(ManagedEvent, event_id) is None:
            return None
        hall = VendorHallEvent(event_id=event_id, created_by=user.email)
        db.add(hall)
        db.flush()
    return hall


def _map_label_matches(candidate: str | None, extracted: str) -> bool:
    candidate_parts = re.findall(r"[A-Z0-9]+", (candidate or "").upper())
    extracted_parts = re.findall(r"[A-Z0-9]+", extracted.upper())
    if not candidate_parts:
        return False
    width = len(candidate_parts)
    return any(
        extracted_parts[index : index + width] == candidate_parts
        for index in range(len(extracted_parts) - width + 1)
    )


def _pdf_text_positions(
    content: bytes,
) -> tuple[
    int,
    list[dict[str, float | int | str]],
    dict[int, tuple[float, float]],
]:
    try:
        document = pymupdf.open(stream=content, filetype="pdf")
    except Exception as exc:
        raise VendorHallError("Floor plan PDF could not be read") from exc
    try:
        if not document.page_count:
            raise VendorHallError("Floor plan PDF does not contain any pages")
        if document.page_count > MAX_FLOOR_MAP_PAGES:
            raise VendorHallError(f"Floor plan PDF cannot exceed {MAX_FLOOR_MAP_PAGES} pages")
        page_count = document.page_count
        fragments: list[dict[str, float | int | str]] = []
        page_dimensions: dict[int, tuple[float, float]] = {}
        for page_number, page in enumerate(document, 1):
            width = page.rect.width or 1
            height = page.rect.height or 1
            page_dimensions[page_number] = (float(width), float(height))
            for word in page.get_text("words"):
                label = " ".join(str(word[4]).split())
                if label:
                    fragments.append(
                        {
                            "text": label,
                            "x": float(word[0]) / width * 100,
                            "y": float(word[1]) / height * 100,
                            "page": page_number,
                        }
                    )
                    if len(fragments) > MAX_FLOOR_MAP_TEXT_FRAGMENTS:
                        raise VendorHallError(
                            "Floor plan PDF contains too much text to analyze safely"
                        )
    except VendorHallError:
        raise
    except Exception as exc:
        raise VendorHallError("Floor plan text and coordinates could not be scanned") from exc
    finally:
        document.close()
    return page_count, fragments, page_dimensions


def _pdf_booth_geometry(
    content: bytes,
    page_number: int,
    booths: list[VendorHallBooth],
) -> dict[str, tuple[float, float, float, float]]:
    """Estimate booth centers and dimensions from printed feet measurements."""
    try:
        document = pymupdf.open(stream=content, filetype="pdf")
    except Exception as exc:
        raise VendorHallError("Floor plan booth geometry could not be scanned") from exc
    try:
        page = document[page_number - 1]
        page_width = page.rect.width or 1
        page_height = page.rect.height or 1
        words = page.get_text("words")
        dimensions = [
            (int(match.group(1)), word)
            for word in words
            if (match := re.fullmatch(r"(\d+)'", str(word[4]))) and int(match.group(1)) >= 13
        ]
        geometry: dict[str, tuple[float, float, float, float]] = {}
        for booth in booths:
            if not booth.booth_number:
                continue
            number_word = next(
                (word for word in words if _map_label_matches(booth.booth_number, str(word[4]))),
                None,
            )
            if number_word is None:
                continue
            bx, by = float(number_word[0]), float(number_word[1])
            width_candidates = []
            for feet, word in dimensions:
                center_x = (float(word[0]) + float(word[2])) / 2
                center_y = (float(word[1]) + float(word[3])) / 2
                dx, dy = center_x - bx, by - center_y
                if 0 < dx < 115 and 35 < dy < 320:
                    width_candidates.append((abs(dx - feet * 1.65) + dy * 0.02, feet))
            if not width_candidates:
                continue
            width_feet = min(width_candidates)[1]
            expected_right = width_feet * 3.3 - 13
            height_candidates = []
            for feet, word in dimensions:
                center_x = (float(word[0]) + float(word[2])) / 2
                center_y = (float(word[1]) + float(word[3])) / 2
                dx, dy = center_x - bx, by - center_y
                if 15 < dx < 230 and -10 < dy < 145:
                    height_candidates.append(
                        (abs(dx - expected_right) + abs(dy - feet * 1.3) * 0.08, feet)
                    )
            if not height_candidates:
                continue
            height_feet = min(height_candidates)[1]
            edge_booth = bx < page_width * 0.15 or bx > page_width * 0.85
            has_nearby_fourteen = any(
                feet == 14
                and abs(((float(word[0]) + float(word[2])) / 2) - bx) < 120
                and abs(by - ((float(word[1]) + float(word[3])) / 2)) < 160
                for feet, word in dimensions
            )
            if edge_booth and has_nearby_fourteen:
                width_feet, height_feet = 26, 14
            width_points = width_feet * 3.3
            height_points = height_feet * 3.3
            bottom = float(number_word[3])
            left = max(0.0, bx - 3.0)
            geometry[booth.id] = (
                (left + width_points / 2) / page_width * 100,
                (bottom - height_points / 2) / page_height * 100,
                width_points / page_width * 100,
                height_points / page_height * 100,
            )
        return geometry
    except Exception as exc:
        raise VendorHallError("Floor plan booth geometry could not be scanned") from exc
    finally:
        document.close()


def _page_booth_matches(
    booths: list[VendorHallBooth],
    fragments: list[dict[str, float | int | str]],
    page_number: int,
) -> list[tuple[VendorHallBooth, dict[str, float | int | str], str]]:
    page_fragments = [item for item in fragments if item["page"] == page_number]
    matches: list[tuple[VendorHallBooth, dict[str, float | int | str], str]] = []
    used_fragments: set[int] = set()

    # Booth numbers are the most reliable anchors. PDF generators frequently
    # merge a whole row of vendor names into one text fragment whose coordinate
    # is the start of the row, not the individual vendor's booth.
    for booth in booths:
        match_index = next(
            (
                index
                for index, fragment in enumerate(page_fragments)
                if index not in used_fragments
                and _map_label_matches(booth.booth_number, str(fragment["text"]))
            ),
            None,
        )
        if match_index is not None:
            used_fragments.add(match_index)
            matches.append((booth, page_fragments[match_index], booth.booth_number or ""))

    matched_ids = {booth.id for booth, _, _ in matches}

    # A label such as "Sealy/Sherwood" denotes a shared footprint. If one
    # company has a numbered anchor, place its paired company at that same
    # anchor so the UI can render one combined booth card.
    for booth in booths:
        if booth.id in matched_ids or not booth.booth_name:
            continue
        for anchor_booth, anchor_fragment, _ in list(matches):
            if not anchor_booth.booth_name:
                continue
            names = (re.escape(anchor_booth.booth_name), re.escape(booth.booth_name))
            page_text = " ".join(str(fragment["text"]) for fragment in page_fragments)
            is_paired = re.search(
                rf"(?:{names[0]}\s*/\s*{names[1]}|{names[1]}\s*/\s*{names[0]})",
                page_text,
                re.IGNORECASE,
            )
            if is_paired:
                matches.append(
                    (booth, anchor_fragment, f"{anchor_booth.booth_name}/{booth.booth_name}")
                )
                matched_ids.add(booth.id)
                break

    # Fall back to names/codes only when no numbered or shared-booth anchor is
    # available.
    for booth in booths:
        if booth.id in matched_ids:
            continue
        for candidate in (booth.booth_name, booth.vendor_code):
            match_index = next(
                (
                    index
                    for index, fragment in enumerate(page_fragments)
                    if index not in used_fragments
                    and _map_label_matches(candidate, str(fragment["text"]))
                ),
                None,
            )
            if match_index is not None:
                used_fragments.add(match_index)
                matches.append((booth, page_fragments[match_index], candidate))
                matched_ids.add(booth.id)
                break
    return matches


def _scan_floor_map_pdf_against_booths(
    db: Session,
    event_id: str,
    content: bytes,
    preserve_manual: bool = True,
) -> dict[str, Any]:
    page_count, fragments, page_dimensions = _pdf_text_positions(content)
    booths = list(
        db.scalars(
            select(VendorHallBooth)
            .where(VendorHallBooth.event_id == event_id)
            .order_by(VendorHallBooth.booth_number)
        ).all()
    )
    manual_raw: dict[str, tuple[float, float, float, float]] = {}
    preserved_crop: tuple[float, float, float, float] | None = None
    if preserve_manual:
        existing_map = db.scalar(
            select(VendorHallFloorMap)
            .join(VendorHallEvent, VendorHallEvent.id == VendorHallFloorMap.vendor_hall_event_id)
            .where(
                VendorHallEvent.event_id == event_id,
                VendorHallFloorMap.is_active.is_(True),
            )
            .order_by(VendorHallFloorMap.uploaded_at.desc())
        )
        old_crop = (existing_map.layout_json or {}).get("crop_box") if existing_map else None
        if isinstance(old_crop, list) and len(old_crop) == 4:
            old_left, old_top, old_right, old_bottom = map(float, old_crop)
            old_width, old_height = old_right - old_left, old_bottom - old_top
            for booth in booths:
                if (
                    booth.map_manually_adjusted
                    and booth.map_x is not None
                    and booth.map_y is not None
                    and booth.map_width is not None
                    and booth.map_height is not None
                ):
                    manual_raw[booth.id] = (
                        old_left + float(booth.map_x) / 100 * old_width,
                        old_top + float(booth.map_y) / 100 * old_height,
                        float(booth.map_width) / 100 * old_width,
                        float(booth.map_height) / 100 * old_height,
                    )
            if manual_raw:
                preserved_crop = (old_left, old_top, old_right, old_bottom)
    page_matches = {
        page_number: _page_booth_matches(booths, fragments, page_number)
        for page_number in range(1, page_count + 1)
    }
    scanned_page = max(page_matches, key=lambda page: (len(page_matches[page]), -page))
    page_width, page_height = page_dimensions[scanned_page]
    matches = page_matches[scanned_page]
    geometry_scan_failed = False
    try:
        geometry = _pdf_booth_geometry(content, scanned_page, booths)
    except VendorHallError:
        geometry = {}
        geometry_scan_failed = True

    # Normalize adjoining booths in the leftmost and rightmost columns by
    # geometry, not company name. This makes stacked edge booths share exact
    # boundaries on any similarly structured floor plan.
    if geometry:
        extreme_x = (
            min(item[0] for item in geometry.values()),
            max(item[0] for item in geometry.values()),
        )
        room_bottom = max(y + height / 2 for _, y, _, height in geometry.values())
        for edge_x in extreme_x:
            column = sorted(
                (
                    (booth_id, measured)
                    for booth_id, measured in geometry.items()
                    if abs(measured[0] - edge_x) <= 1.0
                ),
                key=lambda item: item[1][1],
            )
            for index in range(1, len(column)):
                upper = column[index - 1][1]
                lower_id, lower = column[index]
                upper_bottom = upper[1] + upper[3] / 2
                lower_top = lower[1] - lower[3] / 2
                if abs(lower_top - upper_bottom) <= 1.0:
                    lower_bottom = lower[1] + lower[3] / 2
                    geometry[lower_id] = (
                        lower[0],
                        upper_bottom + (lower_bottom - upper_bottom) / 2,
                        lower[2],
                        lower_bottom - upper_bottom,
                    )
            if column:
                lowest_id = column[-1][0]
                lowest = geometry[lowest_id]
                lowest_bottom = lowest[1] + lowest[3] / 2
                if abs(lowest_bottom - room_bottom) <= 2.0:
                    top = lowest[1] - lowest[3] / 2
                    geometry[lowest_id] = (
                        lowest[0],
                        top + (room_bottom - top) / 2,
                        lowest[2],
                        room_bottom - top,
                    )
    match_geometry: dict[tuple[float, float], tuple[float, float, float, float]] = {}
    for booth, fragment, _ in matches:
        if booth.id in geometry:
            match_geometry[(float(fragment["x"]), float(fragment["y"]))] = geometry[booth.id]
    resolved: list[
        tuple[
            VendorHallBooth,
            dict[str, float | int | str],
            str,
            float,
            float,
            float,
            float,
        ]
    ] = []
    measured_booth_ids: set[str] = set()
    for booth, fragment, matched_label in matches:
        measured = (
            manual_raw.get(booth.id)
            or geometry.get(booth.id)
            or match_geometry.get((float(fragment["x"]), float(fragment["y"])))
        )
        if measured is not None and booth.id not in manual_raw:
            measured_booth_ids.add(booth.id)
        x, y, width, height = measured or (float(fragment["x"]), float(fragment["y"]), 6.5, 8.5)
        resolved.append((booth, fragment, matched_label, x, y, width, height))

    # Crop the operational map to the occupied ballroom, retaining extra room
    # below the booths for the front wall and entry doors.
    if preserved_crop:
        crop_left, crop_top, crop_right, crop_bottom = preserved_crop
    else:
        crop_left = (
            max(0.0, min(x - width / 2 for _, _, _, x, _, width, _ in resolved) - 1.5)
            if resolved
            else 0.0
        )
        crop_right = (
            min(100.0, max(x + width / 2 for _, _, _, x, _, width, _ in resolved) + 1.5)
            if resolved
            else 100.0
        )
        crop_top = (
            max(0.0, min(y - height / 2 for _, _, _, _, y, _, height in resolved) - 3.0)
            if resolved
            else 0.0
        )
        crop_bottom = (
            min(
                100.0,
                max(y + height / 2 for _, _, _, _, y, _, height in resolved) + 8.0,
            )
            if resolved
            else 100.0
        )
    crop_width = max(1.0, crop_right - crop_left)
    crop_height = max(1.0, crop_bottom - crop_top)
    detected: list[dict[str, float | str]] = []
    for booth, _fragment, matched_label, raw_x, raw_y, raw_width, raw_height in resolved:
        booth.floor_map_zone = "Imported floor plan"
        x = (raw_x - crop_left) / crop_width * 100
        y = (raw_y - crop_top) / crop_height * 100
        width = raw_width / crop_width * 100
        height = raw_height / crop_height * 100
        booth.map_x = Decimal(str(round(x, 4)))
        booth.map_y = Decimal(str(round(y, 4)))
        booth.map_width = Decimal(str(round(width, 4)))
        booth.map_height = Decimal(str(round(height, 4)))
        if not preserve_manual:
            booth.map_manually_adjusted = False
        detected.append(
            {
                "booth_id": booth.id,
                "booth_number": booth.booth_number,
                "matched_label": matched_label,
                "x": x,
                "y": y,
                "width": width,
                "height": height,
            }
        )
    unmatched_count = len(booths) - len(matches)
    fallback_count = len(matches) - len(measured_booth_ids)
    review_reasons = []
    if unmatched_count:
        review_reasons.append(f"{unmatched_count} booth labels did not match the PDF")
    if fallback_count:
        review_reasons.append(f"{fallback_count} matched booths used fallback dimensions")
    if geometry_scan_failed:
        review_reasons.append("Printed booth dimensions could not be analyzed")
    return {
        "analysis_version": "vendor-hall-map-v2",
        "analysis_status": "review_required" if review_reasons else "analyzed",
        "geometry_method": "pdf_words_printed_dimensions_and_spatial_framing",
        "source_type": "pdf",
        "page_count": page_count,
        "scanned_page": scanned_page,
        "page_aspect_ratio": round(
            (page_width * crop_width / 100) / (page_height * crop_height / 100), 4
        ),
        "crop_box": [
            round(crop_left, 4),
            round(crop_top, 4),
            round(crop_right, 4),
            round(crop_bottom, 4),
        ],
        "entryways": [
            {"label": "ENTRY", "x": 18},
            {"label": "ENTRY", "x": 37},
            {"label": "ENTRY", "x": 58},
            {"label": "ENTRY", "x": 73},
            {"label": "ENTRY", "x": 90},
        ],
        "scan_method": "pdf_text_coordinates",
        "text_fragment_count": len(fragments),
        "detected_booth_count": len(matches),
        "unmatched_booth_count": unmatched_count,
        "measured_geometry_count": len(measured_booth_ids),
        "fallback_geometry_count": fallback_count,
        "manual_override_count": len(manual_raw),
        "shared_footprint_count": sum("/" in matched_label for _, _, matched_label in matches),
        "detected_booths": detected,
        "review_required": bool(review_reasons),
        "review_reasons": review_reasons,
    }


def import_vendor_hall_floor_map_pdf(
    db: Session,
    event_id: str,
    name: str,
    filename: str,
    content_type: str,
    content: bytes,
    user: User,
) -> VendorHallFloorMapResponse | None:
    _assert_event_open(db, event_id)
    if content_type not in {"application/pdf", "application/x-pdf"}:
        raise VendorHallError("Floor plan must be uploaded as a PDF")
    if not content or len(content) > MAX_FLOOR_MAP_BYTES:
        raise VendorHallError("Floor plan PDF must be between 1 byte and 20 MB")
    if not content_matches_declared_type(content, content_type):
        raise VendorHallError("Uploaded floor plan is not a valid PDF")
    hall = _floor_map_hall(db, event_id, user)
    if hall is None:
        return None
    layout_json = _scan_floor_map_pdf_against_booths(db, event_id, content, preserve_manual=False)
    floor_map = db.scalar(
        select(VendorHallFloorMap).where(
            VendorHallFloorMap.vendor_hall_event_id == hall.id,
            VendorHallFloorMap.is_active.is_(True),
        )
    )
    if floor_map is None:
        floor_map = VendorHallFloorMap(
            vendor_hall_event_id=hall.id,
            name=name,
            uploaded_by=user.email,
        )
        db.add(floor_map)
    floor_map.name = name.strip()
    floor_map.image_filename = (filename or "floor-plan.pdf")[:255]
    floor_map.image_content_type = "application/pdf"
    floor_map.image_content = content
    floor_map.layout_json = layout_json
    floor_map.is_active = True
    floor_map.uploaded_by = user.email
    db.commit()
    db.refresh(floor_map)
    _audit(
        db,
        event_id,
        hall.id,
        None,
        "vendor_hall.floor_map.pdf_imported",
        user.email,
        {
            "floor_map_id": floor_map.id,
            "filename": floor_map.image_filename,
            "scanned_page": layout_json["scanned_page"],
            "detected_booth_count": layout_json["detected_booth_count"],
            "unmatched_booth_count": layout_json["unmatched_booth_count"],
        },
    )
    return _floor_map_response(floor_map)


def vendor_hall_floor_map_content(db: Session, event_id: str) -> tuple[str, str, bytes] | None:
    hall = get_vendor_hall(db, event_id)
    if hall is None:
        return None
    floor_map = db.scalar(
        select(VendorHallFloorMap)
        .where(
            VendorHallFloorMap.vendor_hall_event_id == hall.id,
            VendorHallFloorMap.is_active.is_(True),
            VendorHallFloorMap.image_content.is_not(None),
        )
        .order_by(VendorHallFloorMap.uploaded_at.desc())
    )
    if floor_map is None or floor_map.image_content is None:
        return None
    return (
        floor_map.image_filename or "floor-plan.pdf",
        floor_map.image_content_type or "application/pdf",
        floor_map.image_content,
    )


def vendor_hall_floor_map_preview(db: Session, event_id: str) -> tuple[str, str, bytes] | None:
    source = vendor_hall_floor_map_content(db, event_id)
    if source is None:
        return None
    filename, content_type, content = source
    if content_type not in {"application/pdf", "application/x-pdf"}:
        return filename, content_type, content
    document = pymupdf.open(stream=content, filetype="pdf")
    try:
        if not document.page_count:
            return None
        page = document[0]
        hall = get_vendor_hall(db, event_id)
        floor_map = (
            db.scalar(
                select(VendorHallFloorMap).where(
                    VendorHallFloorMap.vendor_hall_event_id == hall.id,
                    VendorHallFloorMap.is_active.is_(True),
                )
            )
            if hall
            else None
        )
        crop = (floor_map.layout_json or {}).get("crop_box") if floor_map else None
        clip = None
        if isinstance(crop, list) and len(crop) == 4:
            clip = pymupdf.Rect(
                float(crop[0]) / 100 * page.rect.width,
                float(crop[1]) / 100 * page.rect.height,
                float(crop[2]) / 100 * page.rect.width,
                float(crop[3]) / 100 * page.rect.height,
            )
        pixmap = page.get_pixmap(matrix=pymupdf.Matrix(1.8, 1.8), alpha=False, clip=clip)
        return f"{filename.rsplit('.', 1)[0]}.png", "image/png", pixmap.tobytes("png")
    finally:
        document.close()


def vendor_hall_floor_map_status(
    db: Session,
    event_id: str,
) -> VendorHallFloorMapStatusResponse | None:
    event = db.get(ManagedEvent, event_id)
    if event is None:
        return None
    hall = get_vendor_hall(db, event_id)
    floor_map = None
    if hall is not None:
        floor_map = db.scalar(
            select(VendorHallFloorMap)
            .where(
                VendorHallFloorMap.vendor_hall_event_id == hall.id,
                VendorHallFloorMap.is_active.is_(True),
            )
            .order_by(VendorHallFloorMap.uploaded_at.desc())
        )
    booths = list_vendor_hall_booths(db, event_id) or []
    return VendorHallFloorMapStatusResponse(
        event_id=event_id,
        event_name=event.name,
        floor_map=_floor_map_response(floor_map) if floor_map else None,
        booths=booths,
    )


def update_booth_map_position(
    db: Session,
    event_id: str,
    booth_id: str,
    payload: VendorHallBoothMapPositionWrite,
    user: User,
) -> VendorHallBoothResponse | None:
    booth = _get_booth(db, booth_id)
    if booth is None or booth.event_id != event_id:
        return None
    _assert_event_open(db, event_id)
    booth.floor_map_zone = payload.floor_map_zone
    booth.map_x = payload.map_x
    booth.map_y = payload.map_y
    booth.map_width = payload.map_width
    booth.map_height = payload.map_height
    booth.map_manually_adjusted = True
    db.commit()
    db.refresh(booth)
    _audit(
        db,
        booth.event_id,
        booth.vendor_hall_event_id,
        booth.id,
        "vendor_hall.booth.map_position_updated",
        user.email,
        payload.model_dump(mode="json"),
    )
    return _booth_response(db, booth)


def assign_booth_staff(
    db: Session,
    event_id: str,
    booth_id: str,
    payload: VendorHallBoothStaffAssignmentWrite,
    user: User,
) -> VendorHallBoothResponse | None:
    booth = _get_booth(db, booth_id)
    if booth is None or booth.event_id != event_id:
        return None
    _assert_event_open(db, event_id)
    membership = None
    if payload.membership_id:
        membership = db.get(EventMembership, payload.membership_id)
        if (
            membership is None
            or membership.event_id != event_id
            or not membership.is_active
            or membership.membership_type not in {"staff", "admin"}
        ):
            raise VendorHallError("Booth inspection must be assigned to event staff or admin")
    booth.assigned_staff_membership_id = payload.membership_id
    if payload.membership_id and booth.status == "ready_for_inspection":
        existing_task = db.scalar(
            select(EventStaffTask).where(
                EventStaffTask.vendor_hall_booth_id == booth.id,
                EventStaffTask.status.in_(("open", "in_progress", "blocked")),
            )
        )
        if existing_task is None:
            hall = db.get(VendorHallEvent, booth.vendor_hall_event_id)
            db.add(
                EventStaffTask(
                    event_id=booth.event_id,
                    sub_event_id=hall.sub_event_id if hall else None,
                    vendor_hall_booth_id=booth.id,
                    assigned_membership_id=payload.membership_id,
                    title=f"Validate inventory: {booth.booth_name}",
                    description=(
                        f"Validate the vendor inventory for booth {booth.booth_number or 'TBD'}, "
                        "record discrepancies, and complete the booth inspection."
                    ),
                    priority="high",
                    status="open",
                    task_phase="pre_event",
                    created_by=user.email,
                )
            )
            _notify_booth_inspection_assignment(
                db, booth, booth.assigned_staff_membership_id, user.email
            )
    db.commit()
    db.refresh(booth)
    _audit(
        db,
        event_id,
        booth.vendor_hall_event_id,
        booth.id,
        "vendor_hall.booth.staff_assigned",
        user.email,
        {"membership_id": payload.membership_id},
    )
    return _booth_response(db, booth)


def export_vendor_hall_report(
    db: Session,
    event_id: str,
    report_type: str,
) -> tuple[str, str] | None:
    event = db.get(ManagedEvent, event_id)
    if event is None:
        return None
    if report_type not in VENDOR_HALL_EXPORT_REPORTS:
        raise VendorHallError("Unsupported vendor hall export report")
    if report_type == "full-inventory":
        rows = _inventory_export_rows(db, event_id)
    elif report_type == "available-for-sale":
        rows = _inventory_export_rows(db, event_id, available_for_sale=True)
    elif report_type == "damaged-items":
        rows = _inventory_export_rows(db, event_id, statuses={"damaged"})
    elif report_type == "missing-items":
        rows = _inventory_export_rows(db, event_id, statuses={"not_in_booth"})
    elif report_type == "vendor-summary":
        rows = _vendor_summary_export_rows(db, event_id)
    elif report_type == "booth-completion":
        rows = _booth_completion_export_rows(db, event_id)
    else:
        rows = _staff_checkin_export_rows(db, event_id)
    filename = f"vendor-hall-{event.slug}-{report_type}.csv"
    return filename, _rows_to_csv(rows)


def _inventory_export_rows(
    db: Session,
    event_id: str,
    statuses: set[str] | None = None,
    available_for_sale: bool | None = None,
) -> list[dict[str, object]]:
    query = select(VendorHallInventoryItem).where(VendorHallInventoryItem.event_id == event_id)
    if statuses is not None:
        query = query.where(VendorHallInventoryItem.status.in_(statuses))
    if available_for_sale is not None:
        query = query.where(VendorHallInventoryItem.available_for_sale.is_(available_for_sale))
    items = db.scalars(query.order_by(VendorHallInventoryItem.vendor_code)).all()
    booths = {
        booth.id: booth
        for booth in db.scalars(
            select(VendorHallBooth).where(
                VendorHallBooth.id.in_({item.vendor_hall_booth_id for item in items})
            )
        ).all()
    }
    vendor_names = _vendor_names(db, {item.vendor_code for item in items})
    rows: list[dict[str, object]] = []
    for item in items:
        booth = booths.get(item.vendor_hall_booth_id)
        rows.append(
            {
                "event_id": item.event_id,
                "vendor_code": item.vendor_code,
                "vendor_name": vendor_names.get(item.vendor_code, ""),
                "booth_number": booth.booth_number if booth else "",
                "booth_name": booth.booth_name if booth else "",
                "booth_status": booth.status if booth else "",
                "item_name": item.item_name,
                "model_number": item.model_number,
                "serial_number": item.serial_number,
                "quantity_expected": item.quantity_expected,
                "quantity_checked_in": item.quantity_checked_in,
                "unit_price": item.unit_price,
                "condition": item.condition,
                "item_status": item.status,
                "available_for_sale": item.available_for_sale,
                "sell_to_buddys_price": item.sell_to_buddys_price,
                "notes": item.notes,
                "vendor_notes": item.vendor_notes,
                "staff_notes": item.staff_notes,
                "updated_at": _iso(item.updated_at),
            }
        )
    return rows


def _vendor_summary_export_rows(db: Session, event_id: str) -> list[dict[str, object]]:
    booths = db.scalars(
        select(VendorHallBooth)
        .where(VendorHallBooth.event_id == event_id)
        .order_by(VendorHallBooth.vendor_code, VendorHallBooth.booth_number)
    ).all()
    vendor_names = _vendor_names(db, {booth.vendor_code for booth in booths})
    metrics = _booth_export_metrics(db, {booth.id for booth in booths})
    rows: list[dict[str, object]] = []
    for booth in booths:
        booth_metrics = metrics[booth.id]
        rows.append(
            {
                "vendor_code": booth.vendor_code,
                "vendor_name": vendor_names.get(booth.vendor_code, ""),
                "booth_number": booth.booth_number,
                "booth_name": booth.booth_name,
                "booth_status": booth.status,
                "inventory_count": booth_metrics["inventory"],
                "available_for_sale_count": booth_metrics["available"],
                "damaged_count": booth_metrics["damaged"],
                "missing_count": booth_metrics["missing"],
                "exceptions_count": booth_metrics["exceptions"],
                "submitted_at": _iso(booth.submitted_at),
                "checkin_started_at": _iso(booth.checkin_started_at),
                "checkin_completed_at": _iso(booth.checkin_completed_at),
            }
        )
    return rows


def _booth_completion_export_rows(db: Session, event_id: str) -> list[dict[str, object]]:
    booths = db.scalars(
        select(VendorHallBooth)
        .where(VendorHallBooth.event_id == event_id)
        .order_by(VendorHallBooth.booth_number, VendorHallBooth.booth_name)
    ).all()
    vendor_names = _vendor_names(db, {booth.vendor_code for booth in booths})
    metrics = _booth_export_metrics(db, {booth.id for booth in booths})
    rows: list[dict[str, object]] = []
    for booth in booths:
        booth_metrics = metrics[booth.id]
        expected = booth_metrics["inventory"]
        checked = booth_metrics["checked"]
        completion = Decimal("0.00")
        if expected:
            completion = (Decimal(checked) / Decimal(expected) * Decimal(100)).quantize(
                Decimal("0.01")
            )
        rows.append(
            {
                "vendor_code": booth.vendor_code,
                "vendor_name": vendor_names.get(booth.vendor_code, ""),
                "booth_number": booth.booth_number,
                "booth_name": booth.booth_name,
                "status": booth.status,
                "items_expected": expected,
                "items_checked": checked,
                "completion_percentage": completion,
                "exceptions_count": booth_metrics["exceptions"],
                "submitted_by": booth.submitted_by,
                "checked_in_by": booth.checked_in_by,
                "submitted_at": _iso(booth.submitted_at),
                "checkin_started_at": _iso(booth.checkin_started_at),
                "checkin_completed_at": _iso(booth.checkin_completed_at),
            }
        )
    return rows


def _staff_checkin_export_rows(db: Session, event_id: str) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    booth_checkins = db.scalars(
        select(VendorHallBoothCheckin)
        .join(
            VendorHallBooth,
            VendorHallBooth.id == VendorHallBoothCheckin.vendor_hall_booth_id,
        )
        .where(VendorHallBooth.event_id == event_id)
        .order_by(VendorHallBoothCheckin.started_at)
    ).all()
    item_checkins = db.scalars(
        select(VendorHallItemCheckin)
        .join(
            VendorHallBooth,
            VendorHallBooth.id == VendorHallItemCheckin.vendor_hall_booth_id,
        )
        .where(VendorHallBooth.event_id == event_id)
        .order_by(VendorHallItemCheckin.checked_at)
    ).all()
    booth_ids = {checkin.vendor_hall_booth_id for checkin in [*booth_checkins, *item_checkins]}
    booths = {
        booth.id: booth
        for booth in db.scalars(
            select(VendorHallBooth).where(VendorHallBooth.id.in_(booth_ids))
        ).all()
    }
    items = {
        item.id: item
        for item in db.scalars(
            select(VendorHallInventoryItem).where(
                VendorHallInventoryItem.id.in_(
                    {checkin.inventory_item_id for checkin in item_checkins}
                )
            )
        ).all()
    }
    for checkin in booth_checkins:
        booth = booths.get(checkin.vendor_hall_booth_id)
        rows.append(
            {
                "activity_type": "booth_checkin",
                "vendor_code": booth.vendor_code if booth else "",
                "booth_number": booth.booth_number if booth else "",
                "booth_name": booth.booth_name if booth else "",
                "item_name": "",
                "status": checkin.status,
                "quantity_checked": "",
                "actor": checkin.completed_by or checkin.started_by,
                "activity_at": _iso(checkin.completed_at or checkin.started_at),
                "notes": checkin.notes,
            }
        )
    for checkin in item_checkins:
        booth = booths.get(checkin.vendor_hall_booth_id)
        item = items.get(checkin.inventory_item_id)
        rows.append(
            {
                "activity_type": "item_checkin",
                "vendor_code": booth.vendor_code if booth else "",
                "booth_number": booth.booth_number if booth else "",
                "booth_name": booth.booth_name if booth else "",
                "item_name": item.item_name if item else "",
                "status": checkin.status,
                "quantity_checked": checkin.quantity_checked,
                "actor": checkin.checked_by,
                "activity_at": _iso(checkin.checked_at),
                "notes": checkin.exception_notes or checkin.damage_notes,
            }
        )
    return rows


def _rows_to_csv(rows: list[dict[str, object]]) -> str:
    if not rows:
        return "message\nNo rows matched this report.\n"
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()), extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({key: _csv_value(value) for key, value in row.items()})
    return output.getvalue()


def _csv_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return _iso(value)
    return str(spreadsheet_safe_cell(str(value)))


def _vendor_names(db: Session, vendor_codes: set[str]) -> dict[str, str]:
    return dict(
        db.execute(
            select(CatalogVendor.vendor_code, CatalogVendor.name).where(
                CatalogVendor.vendor_code.in_(vendor_codes)
            )
        ).all()
    )


def _booth_export_metrics(
    db: Session,
    booth_ids: set[str],
) -> dict[str, dict[str, int]]:
    metrics = {
        booth_id: {
            "inventory": 0,
            "available": 0,
            "damaged": 0,
            "missing": 0,
            "checked": 0,
            "exceptions": 0,
        }
        for booth_id in booth_ids
    }
    for row in db.execute(
        select(
            VendorHallInventoryItem.vendor_hall_booth_id,
            func.count(),
            func.sum(case((VendorHallInventoryItem.available_for_sale.is_(True), 1), else_=0)),
            func.sum(case((VendorHallInventoryItem.status == "damaged", 1), else_=0)),
            func.sum(case((VendorHallInventoryItem.status == "not_in_booth", 1), else_=0)),
            func.sum(
                case(
                    (
                        VendorHallInventoryItem.status.in_(
                            ("checked_in", "damaged", "not_in_booth", "quantity_mismatch")
                        ),
                        1,
                    ),
                    else_=0,
                )
            ),
        )
        .where(VendorHallInventoryItem.vendor_hall_booth_id.in_(booth_ids))
        .group_by(VendorHallInventoryItem.vendor_hall_booth_id)
    ).all():
        booth_id, inventory, available, damaged, missing, checked = row
        metrics[booth_id].update(
            inventory=inventory,
            available=available or 0,
            damaged=damaged or 0,
            missing=missing or 0,
            checked=checked or 0,
        )
    for booth_id, exception_count in db.execute(
        select(
            VendorHallException.vendor_hall_booth_id,
            func.count(func.distinct(VendorHallException.inventory_item_id)),
        )
        .where(
            VendorHallException.vendor_hall_booth_id.in_(booth_ids),
            VendorHallException.status == "open",
        )
        .group_by(VendorHallException.vendor_hall_booth_id)
    ).all():
        metrics[booth_id]["exceptions"] = exception_count
    return metrics


def _iso(value: datetime | None) -> str:
    return value.isoformat() if value else ""


def _inventory_payload_from_import_row(row: dict[str, str | None]) -> VendorHallInventoryItemWrite:
    normalized = {str(key).strip().lower(): value for key, value in row.items() if key}
    item_name = str(normalized.get("item_name") or normalized.get("name") or "").strip()
    if not item_name:
        raise ValueError("item_name is required")
    available_for_sale = _truthy(normalized.get("available_for_sale"))
    sale_price = _empty_to_none(
        normalized.get("sell_to_buddys_price") or normalized.get("sale_price")
    )
    return VendorHallInventoryItemWrite(
        item_name=item_name,
        model_number=_empty_to_none(normalized.get("model_number") or normalized.get("model")),
        serial_number=_empty_to_none(normalized.get("serial_number") or normalized.get("serial")),
        description=_empty_to_none(normalized.get("description")),
        quantity_expected=int(normalized.get("quantity_expected") or normalized.get("qty") or 1),
        unit_price=_empty_to_none(normalized.get("unit_price") or normalized.get("price")),
        currency=str(normalized.get("currency") or "USD")[:3].upper(),
        condition=_condition_or_unknown(normalized.get("condition")),
        status="expected",
        available_for_sale=available_for_sale,
        sell_to_buddys_price=sale_price if available_for_sale else None,
        notes=_empty_to_none(normalized.get("notes")),
        vendor_notes=_empty_to_none(normalized.get("vendor_notes")),
    )


def _empty_to_none(value: str | None) -> str | None:
    if value is None:
        return None
    stripped = str(value).strip()
    return stripped or None


def _truthy(value: str | None) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on"}


def _condition_or_unknown(value: str | None) -> str:
    condition = str(value or "unknown").strip().lower().replace(" ", "_")
    if condition in {"new", "floor_model", "open_box", "used", "damaged", "unknown"}:
        return condition
    return "unknown"


def _booth_item_count(db: Session, booth_id: str) -> int:
    return (
        db.scalar(
            select(func.count())
            .select_from(VendorHallInventoryItem)
            .where(VendorHallInventoryItem.vendor_hall_booth_id == booth_id)
        )
        or 0
    )


def _booth_checked_count(db: Session, booth_id: str) -> int:
    return (
        db.scalar(
            select(func.count(func.distinct(VendorHallItemCheckin.inventory_item_id)))
            .select_from(VendorHallItemCheckin)
            .where(
                VendorHallItemCheckin.vendor_hall_booth_id == booth_id,
            )
        )
        or 0
    )


def _booth_exception_count(db: Session, booth_id: str) -> int:
    return (
        db.scalar(
            select(func.count(func.distinct(VendorHallException.inventory_item_id)))
            .select_from(VendorHallException)
            .where(
                VendorHallException.vendor_hall_booth_id == booth_id,
                VendorHallException.status == "open",
            )
        )
        or 0
    )


def _create_item_exception(
    db: Session,
    booth: VendorHallBooth,
    item: VendorHallInventoryItem,
    status: str,
    payload: VendorHallItemCheckinWrite,
    actor: str,
) -> None:
    description = payload.exception_notes or payload.damage_notes
    if not description:
        description = f"{item.item_name} marked {status.replace('_', ' ')} during booth check-in."
    db.add(
        VendorHallException(
            vendor_hall_booth_id=booth.id,
            inventory_item_id=item.id,
            exception_type=status,
            severity="high" if status == "damaged" else "medium",
            description=description,
            created_by=actor,
        )
    )


def _audit(
    db: Session,
    event_id: str,
    hall_id: str,
    booth_id: str | None,
    action: str,
    actor: str,
    payload: dict,
    inventory_item_id: str | None = None,
) -> None:
    db.add(
        VendorHallAuditLog(
            event_id=event_id,
            vendor_hall_event_id=hall_id,
            vendor_hall_booth_id=booth_id,
            inventory_item_id=inventory_item_id,
            action=action,
            actor=actor,
            payload=payload,
        )
    )
    db.commit()
