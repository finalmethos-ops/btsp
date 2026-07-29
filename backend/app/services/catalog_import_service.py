from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.catalog import (
    CatalogImportRun,
    CatalogProduct,
    CatalogProductCostHistory,
    CatalogVendor,
    VendorMOQRule,
)
from app.schemas.event_snapshot import EventSnapshotCreate
from app.services.snapshot_service import append_snapshot

MAX_CATALOG_BYTES = 10 * 1024 * 1024
VENDOR_REQUIRED = {"vendor_code", "name"}
PRODUCT_REQUIRED = {
    "model_number",
    "vendor_code",
    "name",
    "unit_price",
    "department",
    "product_category_code",
}
MAX_SHEET_ROWS = 50_000


class CatalogImportError(ValueError):
    pass


def _record_import_cost(
    db: Session,
    product: CatalogProduct,
    actor: str,
    previous_price: Decimal | None = None,
    previous_currency: str | None = None,
) -> None:
    now = datetime.now(UTC)
    current = db.scalar(
        select(CatalogProductCostHistory).where(
            CatalogProductCostHistory.product_code == product.product_code,
            CatalogProductCostHistory.effective_to.is_(None),
        )
    )
    if current is None and previous_price is not None and previous_currency is not None:
        current = CatalogProductCostHistory(
            product_code=product.product_code,
            vendor_code=product.vendor_code,
            unit_price=previous_price,
            currency=previous_currency,
            effective_from=product.created_at or now,
            changed_by=actor,
            source="history-bootstrap",
        )
        db.add(current)
        db.flush()
    if current is not None:
        current.effective_to = now
    db.add(
        CatalogProductCostHistory(
            product_code=product.product_code,
            vendor_code=product.vendor_code,
            unit_price=product.unit_price,
            currency=product.currency,
            effective_from=now,
            changed_by=actor,
            source="admin-import",
        )
    )


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _boolean(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"true", "yes", "y", "1"}:
        return True
    if normalized in {"false", "no", "n", "0"}:
        return False
    raise CatalogImportError(f"Invalid boolean value: {value}")


def _decimal(value: Any, field: str, minimum: Decimal = Decimal("0")) -> Decimal:
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, TypeError) as exc:
        raise CatalogImportError(f"{field} must be numeric") from exc
    if parsed < minimum:
        raise CatalogImportError(f"{field} must be at least {minimum}")
    return parsed


def _currency(value: Any, field: str) -> Decimal:
    parsed = _decimal(value, field)
    if parsed.as_tuple().exponent < -2:
        raise CatalogImportError(f"{field} may have no more than two decimal places")
    return parsed


def _whole_quantity(value: Any, field: str) -> Decimal:
    parsed = _decimal(value, field, Decimal("1"))
    if parsed != parsed.to_integral_value():
        raise CatalogImportError(f"{field} must be a whole number")
    return parsed


def _sheet_rows(workbook: Any, name: str, required: set[str]) -> list[dict[str, Any]]:
    if name not in workbook.sheetnames:
        raise CatalogImportError(f"Workbook is missing required sheet: {name}")
    rows = workbook[name].iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise CatalogImportError(f"Sheet {name} is empty") from exc
    headers = [_text(value).lower() for value in raw_headers]
    missing = required - set(headers)
    if missing:
        raise CatalogImportError(f"Sheet {name} is missing columns: {', '.join(sorted(missing))}")
    result: list[dict[str, Any]] = []
    for number, row in enumerate(rows, start=1):
        if number > MAX_SHEET_ROWS:
            raise CatalogImportError(f"Sheet {name} must not exceed {MAX_SHEET_ROWS} data rows")
        if any(value is not None for value in row):
            result.append(dict(zip(headers, row, strict=False)))
    return result


def _parse(content: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise CatalogImportError("File is not a readable Excel workbook") from exc
    vendors = _sheet_rows(workbook, "Vendors", VENDOR_REQUIRED)
    products = _sheet_rows(workbook, "Products", PRODUCT_REQUIRED)
    parsed_vendors: list[dict[str, Any]] = []
    vendor_codes: set[str] = set()
    for number, row in enumerate(vendors, start=2):
        code = _text(row.get("vendor_code"))
        name = _text(row.get("name"))
        if not code or not name:
            raise CatalogImportError(f"Vendors row {number}: vendor_code and name are required")
        if code in vendor_codes:
            raise CatalogImportError(f"Vendors row {number}: duplicate vendor_code {code}")
        vendor_codes.add(code)
        parsed_vendors.append(
            {"vendor_code": code, "name": name, "is_active": _boolean(row.get("is_active"))}
        )
    parsed_products: list[dict[str, Any]] = []
    product_codes: set[str] = set()
    for number, row in enumerate(products, start=2):
        code = _text(row.get("model_number"))
        vendor_code = _text(row.get("vendor_code"))
        name = _text(row.get("name"))
        if not code or not vendor_code or not name:
            raise CatalogImportError(
                f"Products row {number}: model_number, vendor_code, and name are required"
            )
        if len(code) > 64:
            raise CatalogImportError(f"Products row {number}: model_number exceeds 64 characters")
        if code in product_codes:
            raise CatalogImportError(f"Products row {number}: duplicate model_number {code}")
        if vendor_code not in vendor_codes:
            raise CatalogImportError(f"Products row {number}: unknown vendor_code {vendor_code}")
        product_codes.add(code)
        parsed_products.append(
            {
                "product_code": code,
                "vendor_code": vendor_code,
                "name": name,
                "model_number": code,
                "department": _text(row.get("department")).upper() or None,
                "product_category_code": _text(row.get("product_category_code")).upper() or None,
                "brand": _text(row.get("brand")) or None,
                "unit_price": _currency(row.get("unit_price"), "unit_price"),
                "currency": (_text(row.get("currency")) or "USD").upper(),
                "minimum_order_quantity": _whole_quantity(
                    row.get("minimum_order_quantity") or 1,
                    "minimum_order_quantity",
                ),
                "is_available": _boolean(row.get("is_available")),
                "is_active": _boolean(row.get("is_active")),
            }
        )
    return parsed_vendors, parsed_products


def import_catalog(db: Session, filename: str, content: bytes, actor: str) -> CatalogImportRun:
    if not filename.lower().endswith(".xlsx"):
        raise CatalogImportError("Catalog file must use the .xlsx format")
    if not content or len(content) > MAX_CATALOG_BYTES:
        raise CatalogImportError("Catalog file must be non-empty and no larger than 10 MB")
    vendors, products = _parse(content)
    from app.services.model_category_service import (
        ModelCategoryError,
        validate_model_category,
    )

    for product in products:
        try:
            product["department"], product["product_category_code"] = validate_model_category(
                db, product["department"], product["product_category_code"]
            )
        except ModelCategoryError as exc:
            raise CatalogImportError(str(exc)) from exc
    run = CatalogImportRun(filename=filename, status="processing", imported_by=actor)
    db.add(run)
    for values in vendors:
        item = db.scalar(
            select(CatalogVendor).where(CatalogVendor.vendor_code == values["vendor_code"])
        )
        if item is None:
            item = CatalogVendor(**values, source_file=filename)
            db.add(item)
        else:
            for key, value in values.items():
                setattr(item, key, value)
            item.source_file = filename
    db.flush()
    for values in products:
        item = db.scalar(
            select(CatalogProduct).where(CatalogProduct.product_code == values["product_code"])
        )
        if item is None:
            active_moqs = list(
                db.scalars(
                    select(VendorMOQRule).where(
                        VendorMOQRule.vendor_code == values["vendor_code"],
                        VendorMOQRule.is_active.is_(True),
                    )
                ).all()
            )
            default_moq = (
                active_moqs[0].id
                if len(active_moqs) == 1
                else next(
                    (rule.id for rule in active_moqs if rule.code == "STANDARD"),
                    None,
                )
            )
            item = CatalogProduct(**values, source_file=filename)
            item.moq_rule_id = default_moq
            db.add(item)
            db.flush()
            _record_import_cost(db, item, actor)
        else:
            previous_price = item.unit_price
            previous_currency = item.currency
            for key, value in values.items():
                setattr(item, key, value)
            item.source_file = filename
            if item.unit_price != previous_price or item.currency != previous_currency:
                _record_import_cost(
                    db,
                    item,
                    actor,
                    previous_price,
                    previous_currency,
                )
    run.status = "completed"
    run.vendor_rows = len(vendors)
    run.product_rows = len(products)
    run.completed_at = datetime.now(UTC)
    db.commit()
    db.refresh(run)
    append_snapshot(
        db,
        EventSnapshotCreate(
            event_type="catalog.imported",
            entity_type="catalog_import",
            entity_id=str(run.id),
            actor=actor,
            payload={
                "filename": filename,
                "vendor_rows": len(vendors),
                "product_rows": len(products),
            },
        ),
    )
    return run
