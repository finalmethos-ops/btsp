from datetime import UTC, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.event_management import ManagedEvent
from app.services.event_staff_task_service import list_event_tasks
from app.services.spreadsheet_security import spreadsheet_safe_cell

EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
NAVY = "07142C"
YELLOW = "FFD400"
WHITE = "FFFFFF"
LIGHT_BLUE = "EAF2FF"


def _style_header(row) -> None:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(vertical="center")


def _fit_columns(worksheet, maximum: int = 48) -> None:
    for index, column in enumerate(worksheet.columns, start=1):
        letter = get_column_letter(index)
        width = max((len(str(cell.value or "")) for cell in column), default=0)
        worksheet.column_dimensions[letter].width = min(max(width + 2, 12), maximum)


def export_event_staff_tasks(
    db: Session,
    event_id: str,
) -> tuple[str, bytes] | None:
    event = db.get(ManagedEvent, event_id)
    tasks = list_event_tasks(db, event_id)
    if event is None or tasks is None:
        return None

    now = datetime.now(UTC)
    terminal_statuses = {"done", "cancelled"}
    active = [task for task in tasks if task.status not in terminal_statuses]
    overdue = [task for task in active if task.due_at is not None and task.due_at < now]
    completed = [task for task in tasks if task.status == "done"]
    cancelled_count = len([task for task in tasks if task.status == "cancelled"])
    eligible_count = len(tasks) - cancelled_count
    completion_rate = round((len(completed) / eligible_count) * 100, 1) if eligible_count else 0

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Dashboard"
    overview.sheet_view.showGridLines = False
    overview.merge_cells("A1:D1")
    overview["A1"] = f"{spreadsheet_safe_cell(event.name)} — Staff Task Dashboard"
    overview["A1"].fill = PatternFill("solid", fgColor=NAVY)
    overview["A1"].font = Font(color=WHITE, bold=True, size=18)
    overview["A1"].alignment = Alignment(vertical="center")
    overview.row_dimensions[1].height = 30
    overview["A3"] = "Generated"
    overview["B3"] = now.isoformat()
    overview["A4"] = "Event status"
    overview["B4"] = event.status

    metrics = [
        ("Total tasks", len(tasks)),
        ("Active tasks", len(active)),
        ("In progress", len([task for task in tasks if task.status == "in_progress"])),
        ("Blocked", len([task for task in tasks if task.status == "blocked"])),
        ("Overdue", len(overdue)),
        ("Completed", len(completed)),
        ("Cancelled", cancelled_count),
        ("Completion rate", completion_rate / 100),
        ("Evidence photos", sum(len(task.attachments) for task in tasks)),
    ]
    overview.append([])
    overview.append(["Metric", "Value"])
    _style_header(overview[6])
    for label, value in metrics:
        overview.append([label, value])
    overview["B14"].number_format = "0.0%"

    overview["D3"] = "Immediate attention"
    overview["D3"].fill = PatternFill("solid", fgColor=YELLOW)
    overview["D3"].font = Font(color=NAVY, bold=True)
    attention = sorted(
        [task for task in active if task in overdue or task.status == "blocked"],
        key=lambda task: (task.status != "blocked", task.due_at or now),
    )
    for index, task in enumerate(attention[:10], start=4):
        overview.cell(index, 4, spreadsheet_safe_cell(task.title))
        overview.cell(
            index,
            5,
            spreadsheet_safe_cell(
                f"{task.assigned_display_name} · {task.status.replace('_', ' ')}"
            ),
        )

    details = workbook.create_sheet("Task Register")
    details.sheet_view.showGridLines = False
    headers = [
        "Phase",
        "Sub-event",
        "Task",
        "Instructions",
        "Assigned Staff",
        "Assigned Email",
        "Booth",
        "Priority",
        "Status",
        "Due At",
        "Completed At",
        "Completed By",
        "Staff Note",
        "Evidence Photos",
        "Last Updated",
    ]
    details.append(headers)
    _style_header(details[1])
    for task in tasks:
        details.append(
            [
                spreadsheet_safe_cell(task.task_phase.replace("_", " ").title()),
                spreadsheet_safe_cell(task.sub_event_name or "Entire event"),
                spreadsheet_safe_cell(task.title),
                spreadsheet_safe_cell(task.description or ""),
                spreadsheet_safe_cell(task.assigned_display_name),
                spreadsheet_safe_cell(task.assigned_email),
                spreadsheet_safe_cell(task.vendor_hall_booth_name or ""),
                task.priority,
                task.status.replace("_", " "),
                task.due_at.isoformat() if task.due_at else "",
                task.completed_at.isoformat() if task.completed_at else "",
                spreadsheet_safe_cell(task.completed_by or ""),
                spreadsheet_safe_cell(task.status_note or ""),
                len(task.attachments),
                task.updated_at.isoformat(),
            ]
        )
    for row_index in range(2, details.max_row + 1):
        if row_index % 2 == 0:
            for cell in details[row_index]:
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        for cell in details[row_index]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    details.freeze_panes = "A2"
    details.auto_filter.ref = details.dimensions

    workload = workbook.create_sheet("Staff Workload")
    workload.sheet_view.showGridLines = False
    workload_headers = [
        "Assigned Staff",
        "Assigned Email",
        "Total",
        "Active",
        "In Progress",
        "Blocked",
        "Overdue",
        "Completed",
        "Cancelled",
        "Completion Rate",
    ]
    workload.append(workload_headers)
    _style_header(workload[1])
    assignees: dict[tuple[str, str], list] = {}
    for task in tasks:
        assignees.setdefault(
            (task.assigned_display_name, task.assigned_email),
            [],
        ).append(task)
    for (display_name, email), assigned_tasks in sorted(
        assignees.items(),
        key=lambda item: (item[0][0].casefold(), item[0][1].casefold()),
    ):
        assigned_cancelled = sum(task.status == "cancelled" for task in assigned_tasks)
        assigned_completed = sum(task.status == "done" for task in assigned_tasks)
        assigned_eligible = len(assigned_tasks) - assigned_cancelled
        assigned_active = sum(task.status not in terminal_statuses for task in assigned_tasks)
        assigned_overdue = sum(
            task.status not in terminal_statuses and task.due_at is not None and task.due_at < now
            for task in assigned_tasks
        )
        workload.append(
            [
                spreadsheet_safe_cell(display_name),
                spreadsheet_safe_cell(email),
                len(assigned_tasks),
                assigned_active,
                sum(task.status == "in_progress" for task in assigned_tasks),
                sum(task.status == "blocked" for task in assigned_tasks),
                assigned_overdue,
                assigned_completed,
                assigned_cancelled,
                assigned_completed / assigned_eligible if assigned_eligible else 0,
            ]
        )
    for row_index in range(2, workload.max_row + 1):
        workload.cell(row_index, 10).number_format = "0.0%"
        if row_index % 2 == 0:
            for cell in workload[row_index]:
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    workload.freeze_panes = "A2"
    workload.auto_filter.ref = workload.dimensions

    _fit_columns(overview)
    _fit_columns(details)
    _fit_columns(workload)

    buffer = BytesIO()
    workbook.save(buffer)
    safe_name = "".join(
        character if character.isalnum() or character in {"-", "_"} else "-"
        for character in event.slug
    ).strip("-")
    return f"{safe_name or 'event'}-staff-tasks.xlsx", buffer.getvalue()
