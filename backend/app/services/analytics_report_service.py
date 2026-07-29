import hashlib
import io
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.models.analytics import AnalyticsReportRun, AnalyticsReportSchedule
from app.schemas.analytics import (
    AnalyticsReportScheduleCreate,
    AnalyticsReportType,
    SpendDimension,
)
from app.services.analytics_service import (
    inventory_positions,
    spend_analysis,
    vendor_scorecards,
    workflow_analytics,
)
from app.services.spreadsheet_security import spreadsheet_safe_row


class AnalyticsReportError(ValueError):
    pass


EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
NAVY = "07142C"
NAVY_LIGHT = "0B2A57"
YELLOW = "FFD400"
WHITE = "FFFFFF"
SLATE = "64748B"
LIGHT_BLUE = "EAF2FF"
LIGHT_YELLOW = "FFF4B8"
THIN_BLUE = Side(style="thin", color="9CB8DC")


def _date_parameter(parameters: dict[str, str], key: str) -> datetime | None:
    value = parameters.get(key)
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError as exc:
        raise AnalyticsReportError(f"Report parameter {key} must be an ISO datetime") from exc


def _safe_rows(rows: list[list[object]]) -> list[list[object]]:
    return [spreadsheet_safe_row(row) for row in rows]


def _format_parameter_name(value: str) -> str:
    return value.replace("_", " ").title()


def _format_parameter_value(value: str) -> str:
    return value.replace("T00:00:00Z", "").replace("T23:59:59Z", "")


def _sum_decimal(rows: list[list[object]], index: int) -> Decimal:
    return sum((Decimal(str(row[index] or 0)) for row in rows), Decimal("0"))


def _average(rows: list[list[object]], index: int) -> Decimal | None:
    values = [Decimal(str(row[index])) for row in rows if row[index] is not None]
    if not values:
        return None
    return (sum(values, Decimal("0")) / len(values)).quantize(Decimal("0.01"))


def _report_dataset(
    db: Session,
    report_type: AnalyticsReportType,
    parameters: dict[str, str],
) -> tuple[
    str,
    str,
    list[str],
    list[list[object]],
    list[tuple[str, object, str | None]],
    list[tuple[str, object, str]],
]:
    if report_type is AnalyticsReportType.INVENTORY_POSITION:
        result = inventory_positions(
            db,
            store_number=parameters.get("store_number"),
            product_code=parameters.get("product_code"),
        )
        headers = [
            "Store Number",
            "Product Code",
            "Product Name",
            "Accepted Quantity",
            "Rejected Quantity",
            "Outstanding Backorder Quantity",
        ]
        rows = [
            [
                item.store_number,
                item.product_code,
                item.product_name,
                item.accepted_quantity,
                item.rejected_quantity,
                item.outstanding_backorder_quantity,
            ]
            for item in result.positions
        ]
        metrics = [
            ("Detail Rows", len(rows), "0"),
            ("Stores", len({str(row[0]) for row in rows}), "0"),
            ("Accepted Units", _sum_decimal(rows, 3), "#,##0.00"),
            ("Rejected Units", _sum_decimal(rows, 4), "#,##0.00"),
            ("Backorder Units", _sum_decimal(rows, 5), "#,##0.00"),
        ]
        highlights = [
            (str(row[1]), row[3], f"Store {row[0]}")
            for row in sorted(rows, key=lambda item: Decimal(str(item[3] or 0)), reverse=True)[:10]
        ]
        return (
            "Inventory Position",
            "Inventory Detail",
            headers,
            rows,
            metrics,
            highlights,
        )

    if report_type is AnalyticsReportType.SPEND:
        try:
            dimension = SpendDimension(parameters.get("group_by", "vendor"))
        except ValueError as exc:
            raise AnalyticsReportError("Spend report group_by parameter is invalid") from exc
        result = spend_analysis(
            db,
            dimension,
            date_from=_date_parameter(parameters, "date_from"),
            date_to=_date_parameter(parameters, "date_to"),
            vendor_code=parameters.get("vendor_code"),
            store_number=parameters.get("store_number"),
            workflow_code=parameters.get("workflow_code"),
        )
        headers = [
            _format_parameter_name(dimension.value),
            "Currency",
            "Purchase Order Count",
            "Line Count",
            "Quantity",
            "Landed Spend",
        ]
        rows = [
            [
                item.dimension_key,
                item.currency,
                item.purchase_order_count,
                item.line_count,
                item.quantity,
                item.amount,
            ]
            for item in result.metrics
        ]
        metrics = [
            ("Ranked Groups", len({str(row[0]) for row in rows}), "0"),
            ("Detail Rows", len(rows), "0"),
            ("Ordered Units", _sum_decimal(rows, 4), "#,##0.00"),
        ]
        for currency in sorted({str(row[1]) for row in rows}):
            currency_total = sum(
                (Decimal(str(row[5] or 0)) for row in rows if str(row[1]) == currency),
                Decimal("0"),
            )
            metrics.append((f"{currency} Landed Spend", currency_total, "#,##0.00"))
        highlights = [
            (str(row[0]), row[5], str(row[1]))
            for row in sorted(rows, key=lambda item: Decimal(str(item[5] or 0)), reverse=True)[:10]
        ]
        return (
            f"Spend by {_format_parameter_name(dimension.value)}",
            "Spend Detail",
            headers,
            rows,
            metrics,
            highlights,
        )

    if report_type is AnalyticsReportType.VENDOR_SCORECARDS:
        result = vendor_scorecards(
            db,
            date_from=_date_parameter(parameters, "date_from"),
            date_to=_date_parameter(parameters, "date_to"),
            minimum_orders=int(parameters.get("minimum_orders", "1")),
        )
        headers = [
            "Vendor Code",
            "Vendor Name",
            "Purchase Order Count",
            "Acknowledgement Coverage %",
            "On-Time Delivery %",
            "Receiving Acceptance %",
            "Invoice Match %",
            "Delay Events",
            "Backorder Events",
            "Out-of-Stock Events",
            "Substitution Events",
            "Confirmed PO Changes",
            "Approved Reconciliations",
            "Rejected Reconciliations",
        ]
        rows = [
            [
                item.vendor_code,
                item.vendor_name,
                item.purchase_order_count,
                item.acknowledgement_coverage_rate,
                item.on_time_delivery_rate,
                item.receiving_acceptance_rate,
                item.invoice_match_rate,
                item.delay_event_count,
                item.backorder_event_count,
                item.out_of_stock_event_count,
                item.substitution_event_count,
                item.confirmed_po_change_count,
                item.approved_reconciliation_count,
                item.rejected_reconciliation_count,
            ]
            for item in result.scorecards
        ]
        metrics: list[tuple[str, object, str | None]] = [
            ("Vendors", len(rows), "0"),
            ("Purchase Orders", sum(int(row[2]) for row in rows), "0"),
            ("Average Ack Coverage", _average(rows, 3) or 0, '0.00"%"'),
            ("Average On-Time", _average(rows, 4) or 0, '0.00"%"'),
            ("Average Acceptance", _average(rows, 5) or 0, '0.00"%"'),
            ("Average Invoice Match", _average(rows, 6) or 0, '0.00"%"'),
        ]
        highlights = [
            (str(row[1]), row[2], str(row[0]))
            for row in sorted(rows, key=lambda item: int(item[2]), reverse=True)[:10]
        ]
        return (
            "Vendor Performance Scorecards",
            "Vendor Detail",
            headers,
            rows,
            metrics,
            highlights,
        )

    result = workflow_analytics(
        db,
        date_from=_date_parameter(parameters, "date_from"),
        date_to=_date_parameter(parameters, "date_to"),
        workflow_code=parameters.get("workflow_code"),
    )
    headers = [
        "Workflow Code",
        "Instance Count",
        "Active Count",
        "Completed Count",
        "Transition Count",
        "Approval Count",
        "Rejection Count",
        "Average Completion Seconds",
        "Median Completion Seconds",
        "P90 Completion Seconds",
    ]
    rows = [
        [
            item.workflow_code,
            item.instance_count,
            item.active_count,
            item.completed_count,
            item.transition_count,
            item.approval_count,
            item.rejection_count,
            item.average_completion_seconds,
            item.median_completion_seconds,
            item.p90_completion_seconds,
        ]
        for item in result.workflows
    ]
    metrics = [
        ("Workflows", len(rows), "0"),
        ("Instances", sum(int(row[1]) for row in rows), "0"),
        ("Active", sum(int(row[2]) for row in rows), "0"),
        ("Completed", sum(int(row[3]) for row in rows), "0"),
        ("Approvals", sum(int(row[5]) for row in rows), "0"),
        ("Rejections", sum(int(row[6]) for row in rows), "0"),
    ]
    highlights = [
        (str(row[0]), row[3], f"{row[2]} active")
        for row in sorted(rows, key=lambda item: int(item[3]), reverse=True)[:10]
    ]
    return (
        "Workflow and Approval Performance",
        "Workflow Detail",
        headers,
        rows,
        metrics,
        highlights,
    )


def _style_detail_sheet(sheet, headers: list[str], rows: list[list[object]]) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).coordinate}"
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=YELLOW, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=THIN_BLUE)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, Decimal):
                cell.number_format = "#,##0.00"
    for column_index, header in enumerate(headers, start=1):
        values = [str(header), *(str(row[column_index - 1] or "") for row in rows[:200])]
        width = min(42, max(12, max(len(value) for value in values) + 2))
        sheet.column_dimensions[sheet.cell(1, column_index).column_letter].width = width


def _add_dashboard(
    sheet,
    title: str,
    report_type: AnalyticsReportType,
    parameters: dict[str, str],
    metrics: list[tuple[str, object, str | None]],
    highlights: list[tuple[str, object, str]],
) -> None:
    sheet.sheet_view.showGridLines = False
    for column in "ABCDEFGH":
        sheet.column_dimensions[column].width = 18
    sheet.merge_cells("A1:H2")
    title_cell = sheet["A1"]
    title_cell.value = f"BTSP Executive Analytics — {title}"
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.font = Font(color=YELLOW, bold=True, size=22)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30

    sheet["A4"] = "Generated UTC"
    sheet["B4"] = datetime.now(UTC).replace(microsecond=0).isoformat()
    sheet["E4"] = "Report"
    sheet["F4"] = _format_parameter_name(report_type.value)
    sheet["A5"] = "Filters"
    sheet.merge_cells("B5:H5")
    sheet["B5"] = (
        " · ".join(
            f"{_format_parameter_name(key)}: {_format_parameter_value(value)}"
            for key, value in sorted(parameters.items())
        )
        or "All available records"
    )
    for coordinate in ("A4", "E4", "A5"):
        sheet[coordinate].font = Font(color=SLATE, bold=True)
    sheet["B5"].alignment = Alignment(wrap_text=True)

    for index, (label, value, number_format) in enumerate(metrics[:8]):
        card_row = 8 + (index // 4) * 3
        start_column = 1 + (index % 4) * 2
        end_column = start_column + 1
        sheet.merge_cells(
            start_row=card_row,
            start_column=start_column,
            end_row=card_row,
            end_column=end_column,
        )
        sheet.merge_cells(
            start_row=card_row + 1,
            start_column=start_column,
            end_row=card_row + 1,
            end_column=end_column,
        )
        label_cell = sheet.cell(card_row, start_column, label)
        value_cell = sheet.cell(card_row + 1, start_column, value)
        for row in (card_row, card_row + 1):
            for column in range(start_column, end_column + 1):
                sheet.cell(row, column).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
                sheet.cell(row, column).border = Border(
                    left=THIN_BLUE,
                    right=THIN_BLUE,
                    top=THIN_BLUE,
                    bottom=THIN_BLUE,
                )
        label_cell.font = Font(color=NAVY, bold=True, size=10)
        value_cell.font = Font(color=NAVY, bold=True, size=18)
        value_cell.alignment = Alignment(vertical="center")
        if number_format:
            value_cell.number_format = number_format

    highlight_row = 15
    sheet.merge_cells(start_row=highlight_row, start_column=1, end_row=highlight_row, end_column=8)
    section_cell = sheet.cell(highlight_row, 1, "Ranked Highlights")
    section_cell.fill = PatternFill("solid", fgColor=NAVY_LIGHT)
    section_cell.font = Font(color=YELLOW, bold=True, size=13)
    section_cell.alignment = Alignment(vertical="center")
    header_row = highlight_row + 1
    for column, value in enumerate(["Rank", "Item", "Value", "Context"], start=1):
        cell = sheet.cell(header_row, column, value)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.font = Font(color=NAVY, bold=True)
        cell.border = Border(bottom=THIN_BLUE)
    for rank, (label, value, context) in enumerate(highlights, start=1):
        row = header_row + rank
        safe_values = spreadsheet_safe_row([rank, label, value, context])
        for column, item in enumerate(safe_values, start=1):
            cell = sheet.cell(row, column, item)
            cell.border = Border(bottom=THIN_BLUE)
            if isinstance(item, Decimal):
                cell.number_format = "#,##0.00"

    if highlights:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Top Results"
        chart.height = 7
        chart.width = 13
        chart.legend = None
        chart.y_axis.title = ""
        chart.x_axis.title = "Value"
        data = Reference(
            sheet,
            min_col=3,
            min_row=header_row,
            max_row=header_row + len(highlights),
        )
        categories = Reference(
            sheet,
            min_col=2,
            min_row=header_row + 1,
            max_row=header_row + len(highlights),
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        sheet.add_chart(chart, "E16")

    sheet.freeze_panes = "A8"
    sheet.auto_filter.ref = f"A{header_row}:D{header_row + len(highlights)}"


def render_analytics_report(
    db: Session, report_type: AnalyticsReportType, parameters: dict[str, str]
) -> bytes:
    if report_type is AnalyticsReportType.EXECUTIVE_PACK:
        workbook = Workbook()
        dashboard = workbook.active
        dashboard.title = "Dashboard"
        datasets = [
            _report_dataset(
                db,
                AnalyticsReportType.SPEND,
                {**parameters, "group_by": "vendor"},
            ),
            _report_dataset(
                db,
                AnalyticsReportType.SPEND,
                {**parameters, "group_by": "department"},
            ),
            _report_dataset(
                db,
                AnalyticsReportType.SPEND,
                {**parameters, "group_by": "product_code"},
            ),
            _report_dataset(
                db,
                AnalyticsReportType.VENDOR_SCORECARDS,
                parameters,
            ),
            _report_dataset(
                db,
                AnalyticsReportType.WORKFLOWS,
                parameters,
            ),
            _report_dataset(
                db,
                AnalyticsReportType.INVENTORY_POSITION,
                parameters,
            ),
        ]
        detail_names = [
            "Spend by Vendor",
            "Spend by Department",
            "Product Performance",
            "Vendor Scorecards",
            "Workflow Performance",
            "Inventory Position",
        ]
        for detail_name, dataset in zip(detail_names, datasets, strict=True):
            _, _, headers, rows, _, _ = dataset
            safe_rows = _safe_rows(rows)
            detail = workbook.create_sheet(detail_name)
            detail.append(spreadsheet_safe_row(headers))
            for row in safe_rows:
                detail.append(row)
            _style_detail_sheet(detail, headers, safe_rows)

        product_rows = datasets[2][3]
        scorecard_rows = datasets[3][3]
        workflow_rows = datasets[4][3]
        inventory_rows = datasets[5][3]
        metrics: list[tuple[str, object, str | None]] = [
            ("Vendors", len(scorecard_rows), "0"),
            (
                "Purchase Orders",
                sum(int(row[2]) for row in scorecard_rows),
                "0",
            ),
            ("Ordered Units", _sum_decimal(product_rows, 4), "#,##0.00"),
            (
                "Active Workflows",
                sum(int(row[2]) for row in workflow_rows),
                "0",
            ),
            (
                "Completed Workflows",
                sum(int(row[3]) for row in workflow_rows),
                "0",
            ),
            ("Accepted Units", _sum_decimal(inventory_rows, 3), "#,##0.00"),
            ("Backorder Units", _sum_decimal(inventory_rows, 5), "#,##0.00"),
        ]
        for currency in sorted({str(row[1]) for row in product_rows}):
            metrics.append(
                (
                    f"{currency} Landed Spend",
                    sum(
                        (
                            Decimal(str(row[5] or 0))
                            for row in product_rows
                            if str(row[1]) == currency
                        ),
                        Decimal("0"),
                    ),
                    "#,##0.00",
                )
            )
        highlights = [
            (str(row[0]), row[5], str(row[1]))
            for row in sorted(
                product_rows,
                key=lambda item: Decimal(str(item[5] or 0)),
                reverse=True,
            )[:10]
        ]
        _add_dashboard(
            dashboard,
            "Complete Executive Report Pack",
            report_type,
            parameters,
            metrics,
            highlights,
        )
        workbook.active = 0
        stream = io.BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    title, detail_name, headers, rows, metrics, highlights = _report_dataset(
        db, report_type, parameters
    )
    safe_rows = _safe_rows(rows)
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    detail = workbook.create_sheet(detail_name)
    detail.append(spreadsheet_safe_row(headers))
    for row in safe_rows:
        detail.append(row)
    _style_detail_sheet(detail, headers, safe_rows)
    _add_dashboard(dashboard, title, report_type, parameters, metrics, highlights)
    workbook.active = 0
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def create_report_schedule(
    db: Session, payload: AnalyticsReportScheduleCreate, actor: str
) -> AnalyticsReportSchedule:
    schedule = AnalyticsReportSchedule(
        **payload.model_dump(exclude={"next_run_at"}, mode="json"),
        next_run_at=payload.next_run_at or datetime.now(UTC),
        created_by=actor,
        updated_by=actor,
    )
    db.add(schedule)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise AnalyticsReportError("A report schedule with this name already exists") from exc
    db.refresh(schedule)
    return schedule


def list_report_schedules(db: Session, limit: int = 100) -> list[AnalyticsReportSchedule]:
    return list(
        db.scalars(
            select(AnalyticsReportSchedule).order_by(AnalyticsReportSchedule.name).limit(limit)
        ).all()
    )


def list_report_runs(db: Session, limit: int = 100) -> list[AnalyticsReportRun]:
    return list(
        db.scalars(
            select(AnalyticsReportRun).order_by(AnalyticsReportRun.created_at.desc()).limit(limit)
        ).all()
    )


def run_due_reports(
    db: Session, actor: str, storage_root: str, now: datetime | None = None
) -> list[AnalyticsReportRun]:
    current = now or datetime.now(UTC)
    schedules = list(
        db.scalars(
            select(AnalyticsReportSchedule)
            .where(
                AnalyticsReportSchedule.is_enabled.is_(True),
                AnalyticsReportSchedule.next_run_at <= current,
            )
            .with_for_update(skip_locked=True)
        ).all()
    )
    root = Path(storage_root)
    root.mkdir(parents=True, exist_ok=True)
    runs: list[AnalyticsReportRun] = []
    for schedule in schedules:
        run = AnalyticsReportRun(
            schedule_id=schedule.id,
            scheduled_for=schedule.next_run_at,
            status="processing",
            created_by=actor,
        )
        db.add(run)
        schedule.next_run_at = current + timedelta(minutes=schedule.interval_minutes)
        schedule.updated_by = actor
        db.flush()
        try:
            content = render_analytics_report(
                db, AnalyticsReportType(schedule.report_type), schedule.parameters
            )
            filename = f"{run.id}.xlsx"
            (root / filename).write_bytes(content)
            run.stored_filename = filename
            run.content_type = EXCEL_CONTENT_TYPE
            run.size_bytes = len(content)
            run.sha256 = hashlib.sha256(content).hexdigest()
            run.status = "completed"
        except (AnalyticsReportError, ValueError, OSError) as exc:
            run.status = "failed"
            run.error_message = str(exc)[:1000]
        run.completed_at = current
        runs.append(run)
    db.commit()
    for run in runs:
        db.refresh(run)
    return runs


def report_run_path(run: AnalyticsReportRun, storage_root: str) -> Path:
    if run.status != "completed" or not run.stored_filename:
        raise AnalyticsReportError("Report run has no completed artifact")
    root = Path(storage_root).resolve()
    path = (root / run.stored_filename).resolve()
    if path.parent != root or not path.is_file():
        raise AnalyticsReportError("Report artifact is unavailable")
    digest = hashlib.sha256()
    size = 0
    with path.open("rb") as artifact:
        for chunk in iter(lambda: artifact.read(1024 * 1024), b""):
            size += len(chunk)
            digest.update(chunk)
    if run.size_bytes != size or run.sha256 != digest.hexdigest():
        raise AnalyticsReportError("Report artifact failed its integrity check")
    return path
