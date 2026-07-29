from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.schemas.store_batch import StoreBatchRequest, StoreBatchRow

EXPECTED_HEADERS = (
    "Store #",
    "Entity",
    "Region",
    "Status",
    "Regional Manager",
    "Owner / Operator Name",
    "General Managers Name",
    "Store Manager Email Address",
    "Store address",
    "City",
    "State",
    "Zipcode",
)
PURCHASING_PROGRAMS = {"BPP", "INDEPENDENT"}
MAX_STORE_ROWS = 10_000


class StoreWorkbookError(ValueError):
    pass


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _store_number(value: Any) -> str:
    text = _text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text.zfill(4) if text.isdigit() else text


def _postal_code(value: Any) -> str:
    text = _text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text.zfill(5) if text.isdigit() and len(text) <= 5 else text


def parse_store_workbook(
    content: bytes,
    submitted_by: str,
    source_system: str = "buddys_store_database",
) -> StoreBatchRequest:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise StoreWorkbookError("File is not a readable Excel workbook") from exc
    if not workbook.sheetnames:
        raise StoreWorkbookError("Workbook contains no worksheets")
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = tuple(_text(value) for value in next(rows))
    except StopIteration as exc:
        raise StoreWorkbookError("Workbook is empty") from exc
    if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise StoreWorkbookError("Workbook headers do not match the Buddy's store database")

    batch_rows: list[StoreBatchRow] = []
    seen_store_numbers: set[str] = set()
    for row_number, values in enumerate(rows, start=2):
        if row_number > MAX_STORE_ROWS + 1:
            raise StoreWorkbookError(f"Workbook must not exceed {MAX_STORE_ROWS} store rows")
        if not any(value not in (None, "") for value in values):
            continue
        padded = tuple(values) + (None,) * (len(EXPECTED_HEADERS) - len(values))
        store_number = _store_number(padded[0])
        if not store_number.isdigit() or len(store_number) != 4:
            raise StoreWorkbookError(
                f"Store number must contain four digits at row {row_number}: {store_number}"
            )
        if store_number in seen_store_numbers:
            raise StoreWorkbookError(f"Duplicate store number at row {row_number}: {store_number}")
        seen_store_numbers.add(store_number)
        entity_code = _text(padded[1]).upper()
        region_code = _text(padded[2]).upper()
        purchasing_program = _text(padded[3]).upper()
        if purchasing_program not in PURCHASING_PROGRAMS:
            raise StoreWorkbookError(
                f"Unsupported purchasing program at row {row_number}: {purchasing_program}"
            )
        state_code = _text(padded[10]).upper()
        if len(state_code) != 2:
            raise StoreWorkbookError(f"Invalid state code at row {row_number}: {state_code}")

        batch_rows.append(
            StoreBatchRow(
                row_number=row_number,
                store_number=store_number,
                name=f"Buddy's Store {store_number}",
                region_code=region_code,
                operating_company=entity_code,
                entity_code=entity_code,
                purchasing_program=purchasing_program,
                regional_manager_name=_text(padded[4]),
                owner_operator_name=_text(padded[5]),
                general_manager_name=_text(padded[6]),
                manager_email=_text(padded[7]).lower(),
                address_line1=_text(padded[8]),
                city=_text(padded[9]),
                state_code=state_code,
                postal_code=_postal_code(padded[11]),
                is_ordering_enabled=True,
                is_active=True,
                source_system=source_system,
            )
        )
    if not batch_rows:
        raise StoreWorkbookError("Workbook contains no store rows")
    return StoreBatchRequest(
        source_system=source_system,
        submitted_by=submitted_by,
        rows=batch_rows,
    )


def load_store_workbook(
    path: str | Path,
    submitted_by: str,
    source_system: str = "buddys_store_database",
) -> StoreBatchRequest:
    return parse_store_workbook(Path(path).read_bytes(), submitted_by, source_system)
