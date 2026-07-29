from datetime import UTC, date, datetime, time
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import MetaData, Table, case, inspect, or_, select, text, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.models.catalog import (
    CatalogProduct,
    CatalogProductCostHistory,
    CatalogVendor,
    VendorMOQRule,
)
from app.schemas.catalog import VendorModelUpdate
from app.schemas.event_snapshot import EventSnapshotCreate
from app.services.catalog_import_service import (
    MAX_CATALOG_BYTES,
    _boolean,
    _currency,
    _text,
)
from app.services.model_category_service import ModelCategoryError, validate_model_category
from app.services.snapshot_service import append_snapshot
from app.services.vendor_moq_service import sole_active_rule

MAX_MODEL_ROWS = 50_000

MODEL_COLUMNS = [
    "product_code",
    "vendor_code",
    "name",
    "model_number",
    "department",
    "product_category_code",
    "brand",
    "unit_price",
    "currency",
    "moq_code",
    "is_available",
    "is_active",
    "is_clump",
    "part_of_clump",
    "cost_effective_start_date",
    "cost_status",
]

VENDOR_EXPORT_HEADERS = {
    "vendor": "vendor_name",
    "brand": "brand",
    "model_number": "model_number",
    "product_name": "name",
    "department": "department",
    "productcode": "product_category_code",
    "active": "is_active",
    "is_a_clump": "is_clump",
    "part_of_a_clump": "part_of_clump",
    "effective_start_date": "cost_effective_start_date",
    "landed_cost": "unit_price",
    "in_stock": "is_available",
    "cost_status": "cost_status",
}
VENDOR_EXPORT_REQUIRED_HEADERS = set(VENDOR_EXPORT_HEADERS) - {"product_name"}


class VendorModelError(ValueError):
    pass


def require_vendor_code(vendor_code: str | None) -> str:
    normalized = (vendor_code or "").strip()
    if not normalized:
        raise VendorModelError("Vendor account is missing its vendor identity")
    return normalized


def list_vendor_models(
    db: Session,
    vendor_code: str,
    search: str | None = None,
    classification: str = "all",
) -> list[CatalogProduct]:
    statement = (
        select(CatalogProduct)
        .where(
            CatalogProduct.vendor_code == vendor_code,
            CatalogProduct.model_number.is_not(None),
        )
        .order_by(
            case(
                (CatalogProduct.is_clump.is_(True), 0),
                (CatalogProduct.part_of_clump.is_(False), 1),
                else_=2,
            ),
            CatalogProduct.model_number,
        )
    )
    if search:
        term = f"%{search.strip()}%"
        statement = statement.where(
            or_(
                CatalogProduct.product_code.ilike(term),
                CatalogProduct.name.ilike(term),
                CatalogProduct.model_number.ilike(term),
                CatalogProduct.department.ilike(term),
                CatalogProduct.product_category_code.ilike(term),
            )
        )
    if classification == "clump":
        statement = statement.where(CatalogProduct.is_clump.is_(True))
    elif classification == "part_of_clump":
        statement = statement.where(CatalogProduct.part_of_clump.is_(True))
    elif classification == "single_item":
        statement = statement.where(
            CatalogProduct.is_clump.is_(False),
            CatalogProduct.part_of_clump.is_(False),
        )
    elif classification != "all":
        raise VendorModelError("Unknown model classification filter")
    return list(db.scalars(statement).all())


def _record_cost(
    db: Session,
    product: CatalogProduct,
    actor: str,
    source: str,
    previous_price: Decimal | None = None,
    previous_currency: str | None = None,
    effective_from: datetime | None = None,
) -> None:
    now = datetime.now(UTC)
    effective_at = effective_from or now
    current = db.scalar(
        select(CatalogProductCostHistory).where(
            CatalogProductCostHistory.product_code == product.product_code,
            CatalogProductCostHistory.effective_to.is_(None),
        )
    )
    if current is None and previous_price is not None and previous_currency is not None:
        current = CatalogProductCostHistory(
            product_code=product.product_code,
            vendor_code=product.vendor_code,
            unit_price=previous_price,
            currency=previous_currency,
            effective_from=product.created_at or now,
            changed_by=actor,
            source="history-bootstrap",
        )
        db.add(current)
        db.flush()
    if current is not None:
        current.effective_to = now
    db.add(
        CatalogProductCostHistory(
            product_code=product.product_code,
            vendor_code=product.vendor_code,
            unit_price=product.unit_price,
            currency=product.currency,
            effective_from=effective_at,
            changed_by=actor,
            source=source,
        )
    )


def update_vendor_model(
    db: Session,
    vendor_code: str,
    product_code: str,
    payload: VendorModelUpdate,
    actor: str,
) -> CatalogProduct | None:
    product = db.scalar(
        select(CatalogProduct).where(
            CatalogProduct.product_code == product_code,
            CatalogProduct.vendor_code == vendor_code,
        )
    )
    if product is None:
        return None
    values = payload.model_dump(exclude_unset=True)
    previous_product_code = product.product_code
    new_model_number = values.pop("model_number", None)
    if new_model_number is not None:
        new_model_number = _text(new_model_number)
        if not new_model_number:
            raise VendorModelError("Model number is required")
        conflict = db.scalar(
            select(CatalogProduct.id).where(
                CatalogProduct.product_code == new_model_number,
                CatalogProduct.id != product.id,
            )
        )
        if conflict is not None:
            raise VendorModelError("Model number already exists")
        values["model_number"] = new_model_number
    automatic_rule = sole_active_rule(db, vendor_code)
    if automatic_rule is not None:
        values["moq_rule_id"] = automatic_rule.id
    if "department" in values or "product_category_code" in values:
        department = values.get("department", product.department)
        category_code = values.get("product_category_code", product.product_category_code)
        if not department or not category_code:
            raise VendorModelError("Department and Product Code must be selected together")
        try:
            values["department"], values["product_category_code"] = validate_model_category(
                db, department, category_code
            )
        except ModelCategoryError as exc:
            raise VendorModelError(str(exc)) from exc
    if "moq_rule_id" in values and values["moq_rule_id"] is not None:
        valid_rule = db.scalar(
            select(VendorMOQRule.id).where(
                VendorMOQRule.id == values["moq_rule_id"], VendorMOQRule.vendor_code == vendor_code
            )
        )
        if valid_rule is None:
            raise VendorModelError("MOQ rule does not belong to this vendor")
    if "currency" in values and values["currency"] is not None:
        values["currency"] = values["currency"].upper()
    previous_price = product.unit_price
    previous_currency = product.currency
    identifier_changed = new_model_number is not None and new_model_number != previous_product_code
    for field, value in values.items():
        if field != "model_number":
            setattr(product, field, value)
    cost_changed = product.unit_price != previous_price or product.currency != previous_currency
    if cost_changed:
        _record_cost(
            db,
            product,
            actor,
            "single-edit",
            previous_price,
            previous_currency,
        )
    db.flush()
    if identifier_changed:
        db.execute(
            text(
                "UPDATE catalog_products "
                "SET product_code = :new_code, model_number = :new_code "
                "WHERE id = :product_id AND vendor_code = :vendor_code"
            ),
            {
                "new_code": new_model_number,
                "product_id": product.id,
                "vendor_code": vendor_code,
            },
        )
        connection = db.connection()
        inspector = inspect(connection)
        metadata = MetaData()
        for table_name in (
            "catalog_product_cost_history",
            "purchase_request_line_items",
            "purchase_order_lines",
            "purchase_receipt_lines",
            "purchase_backorders",
            "vendor_advance_ship_notice_lines",
            "vendor_invoice_lines",
        ):
            if inspector.has_table(table_name):
                related_table = Table(table_name, metadata, autoload_with=connection)
                db.execute(
                    update(related_table)
                    .where(related_table.c.product_code == previous_product_code)
                    .values(product_code=new_model_number)
                )
        db.expire(product)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise VendorModelError("Model number already exists") from exc
    db.refresh(product)
    append_snapshot(
        db,
        EventSnapshotCreate(
            event_type="catalog.vendor_model.updated",
            entity_type="catalog_product",
            entity_id=product.product_code,
            actor=actor,
            payload={
                "vendor_code": vendor_code,
                "fields": sorted(values),
                "previous_model_number": previous_product_code if identifier_changed else None,
            },
        ),
    )
    return product


def list_cost_history(
    db: Session,
    vendor_code: str,
    product_code: str,
) -> list[CatalogProductCostHistory] | None:
    exists = db.scalar(
        select(CatalogProduct.id).where(
            CatalogProduct.product_code == product_code,
            CatalogProduct.vendor_code == vendor_code,
        )
    )
    if exists is None:
        return None
    return list(
        db.scalars(
            select(CatalogProductCostHistory)
            .where(
                CatalogProductCostHistory.product_code == product_code,
                CatalogProductCostHistory.vendor_code == vendor_code,
            )
            .order_by(CatalogProductCostHistory.effective_from.desc())
        ).all()
    )


def export_vendor_models(db: Session, vendor_code: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(MODEL_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="123A73")
    rules = {
        rule.id: rule.code
        for rule in db.scalars(
            select(VendorMOQRule).where(VendorMOQRule.vendor_code == vendor_code)
        ).all()
    }
    for product in list_vendor_models(db, vendor_code):
        sheet.append(
            [
                product.product_code,
                product.vendor_code,
                product.name,
                product.model_number,
                product.department,
                product.product_category_code,
                product.brand,
                product.unit_price,
                product.currency,
                rules.get(product.moq_rule_id, ""),
                product.is_available,
                product.is_active,
                product.is_clump,
                product.part_of_clump,
                product.cost_effective_start_date,
                product.cost_status,
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [20, 18, 36, 20, 20, 22, 22, 20, 14, 12, 24, 16, 12]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _header(value: Any) -> str:
    return _text(value).lower().replace(" ", "_")


def _rows(content: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise VendorModelError("File is not a readable Excel workbook") from exc
    sheet = workbook["Products"] if "Products" in workbook.sheetnames else workbook.active
    values = sheet.iter_rows(values_only=True)
    try:
        raw_headers = [_header(value) for value in next(values)]
    except StopIteration as exc:
        raise VendorModelError("Products sheet is empty") from exc
    # Product name was added to the vendor export later, so accept both the
    # legacy layout and the enriched layout while still requiring every field
    # needed to import a valid model/cost record.
    is_vendor_export = VENDOR_EXPORT_REQUIRED_HEADERS.issubset(raw_headers)
    headers = (
        [VENDOR_EXPORT_HEADERS.get(header, header) for header in raw_headers]
        if is_vendor_export
        else raw_headers
    )
    required = {"model_number", "unit_price", "department", "product_category_code"}
    if not is_vendor_export:
        required.add("name")
    missing = required - set(headers)
    if missing:
        raise VendorModelError(f"Product workbook is missing: {', '.join(sorted(missing))}")
    result: list[dict[str, Any]] = []
    for number, row in enumerate(values, start=1):
        if number > MAX_MODEL_ROWS:
            raise VendorModelError(f"Products sheet must not exceed {MAX_MODEL_ROWS} data rows")
        if any(value is not None for value in row):
            item = dict(zip(headers, row, strict=False))
            item["__vendor_export"] = is_vendor_export
            result.append(item)
    return result


def _vendor_identity(value: Any) -> set[str]:
    ignored = {"GROUP", "INC", "LLC", "LTD", "CORP", "CORPORATION", "COMPANY", "CO"}
    normalized = "".join(
        character if character.isalnum() else " " for character in _text(value).upper()
    )
    return {token for token in normalized.split() if token not in ignored}


def _effective_date(value: Any, number: int) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(_text(value))
    except ValueError as exc:
        raise VendorModelError(
            f"Products row {number}: effective start date must be a date"
        ) from exc


def _parsed_values(row: dict[str, Any], vendor: CatalogVendor, number: int) -> dict[str, Any]:
    model_number = _text(row.get("model_number"))
    name = _text(row.get("name")) or model_number
    row_vendor = _text(row.get("vendor_code"))
    row_vendor_name = _text(row.get("vendor_name"))
    if not model_number or not name:
        raise VendorModelError(f"Products row {number}: model_number and name are required")
    if len(model_number) > 64:
        raise VendorModelError(f"Products row {number}: model_number exceeds 64 characters")
    if row_vendor and row_vendor != vendor.vendor_code:
        raise VendorModelError(f"Products row {number}: vendor_code does not match your account")
    if row_vendor_name and _vendor_identity(row_vendor_name) != _vendor_identity(vendor.name):
        raise VendorModelError(f"Products row {number}: Vendor does not match your account")
    try:
        return {
            "product_code": model_number,
            "vendor_code": vendor.vendor_code,
            "name": name,
            "model_number": model_number,
            "department": _text(row.get("department")).upper(),
            "product_category_code": _text(row.get("product_category_code")).upper(),
            "brand": _text(row.get("brand")) or None,
            "is_clump": _boolean(row.get("is_clump"), False),
            "part_of_clump": _boolean(row.get("part_of_clump"), False),
            "cost_effective_start_date": _effective_date(
                row.get("cost_effective_start_date"), number
            ),
            "unit_price": _currency(row.get("unit_price"), "unit_price"),
            "cost_status": _text(row.get("cost_status")) or "Approved",
            "currency": (_text(row.get("currency")) or "USD").upper(),
            "moq_code": _text(row.get("moq_code")) or "STANDARD",
            "is_available": _boolean(row.get("is_available")),
            "is_active": _boolean(row.get("is_active")),
        }
    except ValueError as exc:
        raise VendorModelError(f"Products row {number}: {exc}") from exc


def import_vendor_models(
    db: Session,
    vendor_code: str,
    filename: str,
    content: bytes,
    actor: str,
) -> tuple[int, int, int, int]:
    if not filename.lower().endswith(".xlsx"):
        raise VendorModelError("Model import must use the .xlsx format")
    if not content or len(content) > MAX_CATALOG_BYTES:
        raise VendorModelError("Workbook must be non-empty and no larger than 10 MB")
    vendor = db.scalar(
        select(CatalogVendor).where(
            CatalogVendor.vendor_code == vendor_code,
            CatalogVendor.is_active.is_(True),
        )
    )
    if vendor is None:
        raise VendorModelError("Vendor identity is not active in the catalog")
    source_rows = _rows(content)
    is_vendor_export = bool(source_rows and source_rows[0].pop("__vendor_export", False))
    parsed = [_parsed_values(row, vendor, number) for number, row in enumerate(source_rows, 2)]
    for values in parsed:
        # Vendor exports may legitimately omit classification for a small set of
        # newly introduced models. Preserve the model/name/cost and let admins
        # classify those records later instead of rejecting the entire batch.
        if not values["department"] or not values["product_category_code"]:
            values["department"] = values["department"] or None
            values["product_category_code"] = values["product_category_code"] or None
            continue
        try:
            values["department"], values["product_category_code"] = validate_model_category(
                db, values["department"], values["product_category_code"]
            )
        except ModelCategoryError as exc:
            if is_vendor_export:
                continue
            raise VendorModelError(str(exc)) from exc
    active_rules = list(
        db.scalars(
            select(VendorMOQRule).where(
                VendorMOQRule.vendor_code == vendor_code,
                VendorMOQRule.is_active.is_(True),
            )
        ).all()
    )
    rules = {rule.code: rule.id for rule in active_rules}
    automatic_rule_id = active_rules[0].id if len(active_rules) == 1 else None
    for values in parsed:
        moq_code = values.pop("moq_code")
        if not active_rules:
            values["moq_rule_id"] = None
        elif automatic_rule_id is not None:
            values["moq_rule_id"] = automatic_rule_id
        elif moq_code not in rules:
            raise VendorModelError(f"Unknown MOQ code {moq_code}")
        else:
            values["moq_rule_id"] = rules[moq_code]
    codes = [item["product_code"] for item in parsed]
    if len(codes) != len(set(codes)):
        raise VendorModelError("Products sheet contains duplicate model_number values")
    existing = {
        product.product_code: product
        for product in db.scalars(
            select(CatalogProduct).where(CatalogProduct.product_code.in_(codes))
        ).all()
    }
    created = updated = unchanged = 0
    for values in parsed:
        product = existing.get(values["product_code"])
        if product is not None and product.vendor_code != vendor_code:
            raise VendorModelError(f"Model number {product.product_code} belongs to another vendor")
        if product is None:
            product = CatalogProduct(**values, source_file=filename)
            db.add(product)
            db.flush()
            effective_date = values.get("cost_effective_start_date")
            _record_cost(
                db,
                product,
                actor,
                "batch-import",
                effective_from=(
                    datetime.combine(effective_date, time.min, tzinfo=UTC)
                    if effective_date
                    else None
                ),
            )
            created += 1
            continue
        changed_fields = [key for key, value in values.items() if getattr(product, key) != value]
        if not changed_fields:
            unchanged += 1
            continue
        previous_price = product.unit_price
        previous_currency = product.currency
        for field in changed_fields:
            setattr(product, field, values[field])
        product.source_file = filename
        if "unit_price" in changed_fields or "currency" in changed_fields:
            _record_cost(
                db,
                product,
                actor,
                "batch-import",
                previous_price,
                previous_currency,
                effective_from=(
                    datetime.combine(values["cost_effective_start_date"], time.min, tzinfo=UTC)
                    if values.get("cost_effective_start_date")
                    else None
                ),
            )
        updated += 1
    db.commit()
    append_snapshot(
        db,
        EventSnapshotCreate(
            event_type="catalog.vendor_models.imported",
            entity_type="catalog_vendor",
            entity_id=vendor_code,
            actor=actor,
            payload={
                "filename": filename,
                "created": created,
                "updated": updated,
                "unchanged": unchanged,
                "total_rows": len(parsed),
            },
        ),
    )
    return created, updated, unchanged, len(parsed)
