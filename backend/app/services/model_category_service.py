from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.catalog import ModelCategory


class ModelCategoryError(ValueError):
    pass


MAX_CATEGORY_ROWS = 50_000


def _text(value: object) -> str:
    return "" if value is None else str(value).strip().upper()


def import_model_categories(db: Session, content: bytes) -> tuple[int, int]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ModelCategoryError("Category workbook could not be read") from exc
    if "Model Categories" not in workbook.sheetnames:
        raise ModelCategoryError("Category workbook requires a Model Categories sheet")
    rows = workbook["Model Categories"].iter_rows(values_only=True)
    try:
        headers = tuple(_text(value) for value in next(rows))
    except StopIteration as exc:
        raise ModelCategoryError("Category workbook is empty") from exc
    if headers[:3] != ("DEPARTMENT", "PRODUCT CODE", "STATUS"):
        raise ModelCategoryError("Category workbook headers are invalid")
    values: dict[tuple[str, str], str] = {}
    for number, row in enumerate(rows, start=1):
        if number > MAX_CATEGORY_ROWS:
            raise ModelCategoryError(
                f"Category workbook must not exceed {MAX_CATEGORY_ROWS} data rows"
            )
        if len(row) >= 2 and _text(row[0]) and _text(row[1]):
            values[(_text(row[0]), _text(row[1]))] = _text(row[2]) or "VALID"
    existing = {
        (item.department, item.product_category_code): item
        for item in db.scalars(select(ModelCategory)).all()
    }
    created = updated = 0
    for (department, code), status in values.items():
        item = existing.get((department, code))
        if item is None:
            db.add(
                ModelCategory(
                    department=department,
                    product_category_code=code,
                    status=status,
                )
            )
            created += 1
        elif item.status != status:
            item.status = status
            updated += 1
    db.commit()
    return created, updated


def list_model_categories(db: Session) -> list[ModelCategory]:
    return list(
        db.scalars(
            select(ModelCategory)
            .where(ModelCategory.status == "VALID")
            .order_by(ModelCategory.department, ModelCategory.product_category_code)
        ).all()
    )


def validate_model_category(db: Session, department: str, product_code: str) -> tuple[str, str]:
    normalized_department = department.strip().upper()
    normalized_code = product_code.strip().upper()
    exists = db.scalar(
        select(ModelCategory.id).where(
            ModelCategory.department == normalized_department,
            ModelCategory.product_category_code == normalized_code,
            ModelCategory.status == "VALID",
        )
    )
    if exists is None:
        raise ModelCategoryError(
            f"Invalid Department/Product Code pair: {normalized_department} / {normalized_code}"
        )
    return normalized_department, normalized_code
