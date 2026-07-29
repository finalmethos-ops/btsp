from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from app.db.session import Base
from app.models.catalog import CatalogProduct, CatalogVendor
from app.models.event_snapshot import EventSnapshot
from app.models.purchase_order import PurchaseOrder, PurchaseOrderLine
from app.models.purchasing import PurchaseRequest  # noqa: F401 - registers FK metadata
from app.models.receiving import (
    InvoiceLineMatch,
    InvoiceReconciliation,
    PurchaseBackorder,
    PurchaseReceipt,
    PurchaseReceiptLine,
    ReceiptVariance,
    ReconciliationException,
    VendorInvoice,
    VendorInvoiceLine,
)
from app.models.store import Store
from app.models.workflow import WorkflowInstance
from app.schemas.analytics import AnalyticsReportScheduleCreate, AnalyticsReportType, SpendDimension
from app.services.analytics_report_service import (
    create_report_schedule,
    render_analytics_report,
    report_run_path,
    run_due_reports,
)
from app.services.analytics_service import (
    executive_spend_dashboard,
    inventory_positions,
    operational_dashboard,
    spend_analysis,
    vendor_scorecards,
    workflow_analytics,
)


@pytest.fixture
def db() -> Session:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        session.add_all(
            [
                CatalogVendor(
                    vendor_code="V-A", name="Vendor A", is_active=True, source_file="test"
                ),
                Store(store_number="001", name="Store", region_code="EAST", is_active=True),
            ]
        )
        session.commit()
        yield session


def _order(currency: str, total: int, number: str, status: str = "received") -> PurchaseOrder:
    return PurchaseOrder(
        po_number=number,
        workflow_code="BPP_PURCHASING",
        vendor_code="V-A",
        status=status,
        currency=currency,
        subtotal=total,
        freight_total=0,
        tax_total=0,
        total=total,
        created_by="buyer@example.com",
    )


def test_operational_dashboard_has_stable_zero_state(db: Session) -> None:
    dashboard = operational_dashboard(db)

    assert dashboard.purchasing.purchase_order_count == 0
    assert dashboard.purchasing.ordered_spend == []
    assert dashboard.receiving.accepted_quantity == 0
    assert dashboard.invoices.invoice_count == 0
    assert dashboard.reconciliation.open_exception_count == 0
    assert vendor_scorecards(db).scorecards == []


def test_executive_spend_dashboard_tracks_entity_periods_and_best_sellers(
    db: Session,
) -> None:
    store = db.scalar(select(Store).where(Store.store_number == "001"))
    assert store is not None
    store.entity_code = "BEBE"
    db.add(
        Store(
            store_number="002",
            name="Second Store",
            region_code="EAST",
            entity_code="CRD",
            is_active=True,
        )
    )
    orders = [
        (
            datetime(2026, 7, 10, tzinfo=UTC),
            "PO-MTD-BEBE",
            "001",
            "SKU-A",
            "Product A",
            3,
            300,
        ),
        (
            datetime(2026, 7, 12, tzinfo=UTC),
            "PO-MTD-CRD",
            "002",
            "SKU-B",
            "Product B",
            2,
            500,
        ),
        (
            datetime(2026, 2, 1, tzinfo=UTC),
            "PO-YTD-BEBE",
            "001",
            "SKU-A",
            "Product A",
            1,
            100,
        ),
        (
            datetime(2025, 12, 20, tzinfo=UTC),
            "PO-PRIOR-YEAR",
            "001",
            "SKU-C",
            "Prior Product",
            10,
            1000,
        ),
    ]
    for created_at, number, store_number, code, name, quantity, total in orders:
        order = _order("USD", total, number)
        order.created_at = created_at
        order.lines.append(
            PurchaseOrderLine(
                source_request_id=f"request-{number}",
                source_line_id=1,
                store_number=store_number,
                product_code=code,
                product_name=name,
                quantity=quantity,
                unit_price=total / quantity,
                freight_amount=0,
                tax_amount=0,
                extended_amount=total,
            )
        )
        db.add(order)
    db.commit()

    dashboard = executive_spend_dashboard(db, datetime(2026, 7, 26, 12, tzinfo=UTC))

    assert [(item.entity_code, item.currency, item.amount) for item in dashboard.mtd_by_entity] == [
        ("CRD", "USD", 500),
        ("BEBE", "USD", 300),
    ]
    assert [(item.entity_code, item.currency, item.amount) for item in dashboard.ytd_by_entity] == [
        ("CRD", "USD", 500),
        ("BEBE", "USD", 400),
    ]
    assert [
        (item.rank, item.product_code, item.quantity, item.amount)
        for item in dashboard.top_sellers_mtd
    ] == [(1, "SKU-A", 3, 300), (2, "SKU-B", 2, 500)]


def test_operational_dashboard_aggregates_without_cross_currency_sums(db: Session) -> None:
    usd_order = _order("USD", 100, "PO-USD", "partially_received")
    eur_order = _order("EUR", 50, "PO-EUR")
    line = PurchaseOrderLine(
        source_request_id="request-1",
        source_line_id=1,
        store_number="001",
        product_code="SKU-1",
        product_name="Product",
        quantity=10,
        unit_price=10,
        freight_amount=0,
        tax_amount=0,
        extended_amount=100,
    )
    usd_order.lines.append(line)
    db.add_all([usd_order, eur_order])
    db.flush()
    receipt = PurchaseReceipt(
        receipt_number="RCV-1",
        purchase_order_id=usd_order.id,
        store_number="001",
        receipt_sha256="a" * 64,
        status="posted_with_exceptions",
        received_at=datetime.now(UTC),
        received_by="receiver@example.com",
    )
    receipt_line = PurchaseReceiptLine(
        purchase_order_line_id=line.id,
        product_code="SKU-1",
        received_quantity=8,
        accepted_quantity=7,
        rejected_quantity=1,
    )
    receipt.lines.append(receipt_line)
    db.add(receipt)
    db.flush()
    variance = ReceiptVariance(
        receipt_id=receipt.id,
        receipt_line_id=receipt_line.id,
        variance_type="rejected_quantity",
        severity="exception",
        expected_quantity=0,
        actual_quantity=1,
        difference_quantity=1,
        status="open",
    )
    db.add(variance)
    db.flush()
    db.add(
        PurchaseBackorder(
            backorder_number="BO-1",
            source_variance_id=variance.id,
            purchase_order_id=usd_order.id,
            purchase_order_line_id=line.id,
            store_number="001",
            product_code="SKU-1",
            original_quantity=1,
            fulfilled_quantity=0,
            outstanding_quantity=1,
            status="open",
            created_by="manager@example.com",
        )
    )
    invoice = VendorInvoice(
        invoice_number="INV-1",
        vendor_code="V-A",
        purchase_order_id=usd_order.id,
        invoice_sha256="b" * 64,
        invoice_date=datetime.now(UTC),
        currency="USD",
        subtotal=100,
        freight_total=0,
        tax_total=0,
        total=100,
        status="match_exception",
        received_by="ap@example.com",
    )
    invoice_line = VendorInvoiceLine(
        line_number=1,
        purchase_order_line_id=line.id,
        product_code="SKU-1",
        quantity=10,
        unit_price=10,
        extended_amount=100,
    )
    invoice_line.match = InvoiceLineMatch(
        ordered_quantity=10,
        accepted_quantity=7,
        invoiced_quantity=10,
        quantity_difference=3,
        ordered_unit_price=10,
        invoiced_unit_price=10,
        price_difference=0,
        status="exception",
    )
    invoice.lines.append(invoice_line)
    db.add(invoice)
    db.flush()
    case = InvoiceReconciliation(
        invoice_id=invoice.id,
        purchase_order_id=usd_order.id,
        status="exception_review",
        created_by="analyst@example.com",
    )
    case.exceptions.append(
        ReconciliationException(
            invoice_line_id=invoice_line.id,
            exception_type="quantity",
            expected_amount=7,
            actual_amount=10,
            difference_amount=3,
            status="open",
        )
    )
    db.add(case)
    db.add_all(
        [
            EventSnapshot(
                event_type="vendor.fulfillment.delay_reported",
                entity_type="purchase_order",
                entity_id=usd_order.id,
                actor="vendor@example.com",
                payload={"vendor_code": "V-A"},
            ),
            EventSnapshot(
                event_type="vendor.fulfillment.backorder_reported",
                entity_type="purchase_order",
                entity_id=usd_order.id,
                actor="vendor@example.com",
                payload={"vendor_code": "V-A"},
            ),
            EventSnapshot(
                event_type="vendor.fulfillment.out_of_stock_reported",
                entity_type="purchase_order",
                entity_id=usd_order.id,
                actor="vendor@example.com",
                payload={"vendor_code": "V-A"},
            ),
            EventSnapshot(
                event_type="vendor.fulfillment.substitution_proposed",
                entity_type="purchase_order",
                entity_id=usd_order.id,
                actor="vendor@example.com",
                payload={"vendor_code": "V-A"},
            ),
            EventSnapshot(
                event_type="purchase_order.vendor_change_confirmed",
                entity_type="purchase_order",
                entity_id=usd_order.id,
                actor="vendor@example.com",
                payload={"vendor_code": "V-A"},
            ),
        ]
    )
    db.commit()

    dashboard = operational_dashboard(db)

    assert dashboard.purchasing.purchase_order_count == 2
    assert [(item.currency, item.amount) for item in dashboard.purchasing.ordered_spend] == [
        ("EUR", 50),
        ("USD", 100),
    ]
    assert dashboard.receiving.accepted_quantity == 7
    assert dashboard.receiving.rejected_quantity == 1
    assert dashboard.receiving.open_variance_count == 1
    assert dashboard.receiving.outstanding_backorder_quantity == 1
    assert dashboard.invoices.line_match_exception_count == 1
    assert dashboard.reconciliation.open_exception_count == 1

    position = inventory_positions(db).positions[0]
    assert position.store_number == "001"
    assert position.product_code == "SKU-1"
    assert position.accepted_quantity == 7
    assert position.rejected_quantity == 1
    assert position.outstanding_backorder_quantity == 1

    scorecard = vendor_scorecards(db).scorecards[0]
    assert scorecard.vendor_code == "V-A"
    assert scorecard.purchase_order_count == 2
    assert [(item.currency, item.amount) for item in scorecard.ordered_spend] == [
        ("EUR", 50),
        ("USD", 100),
    ]
    assert scorecard.acknowledgement_coverage_rate == Decimal("0.00")
    assert scorecard.on_time_delivery_rate is None
    assert scorecard.receiving_acceptance_rate == Decimal("87.50")
    assert scorecard.invoice_match_rate == Decimal("0.00")
    assert scorecard.vendor_fulfillment_event_count == 4
    assert scorecard.delay_event_count == 1
    assert scorecard.backorder_event_count == 1
    assert scorecard.out_of_stock_event_count == 1
    assert scorecard.substitution_event_count == 1
    assert scorecard.confirmed_po_change_count == 1


def test_spend_analysis_groups_landed_line_spend_and_applies_filters(db: Session) -> None:
    db.add(
        CatalogProduct(
            product_code="SKU-SPEND",
            vendor_code="V-A",
            name="Analytics product",
            department="FURNITURE",
            product_category_code="FIXTURES",
            unit_price=10,
            currency="USD",
            minimum_order_quantity=1,
            is_available=True,
            is_active=True,
            source_file="test",
        )
    )
    usd = _order("USD", 107, "PO-SPEND-USD")
    eur = _order("EUR", 20, "PO-SPEND-EUR")
    for order, store, quantity, amount, freight, tax in [
        (usd, "001", 10, 100, 5, 2),
        (eur, "002", 2, 20, 0, 0),
    ]:
        order.lines.append(
            PurchaseOrderLine(
                source_request_id=f"request-{store}",
                source_line_id=1,
                store_number=store,
                product_code="SKU-SPEND",
                product_name="Analytics product",
                quantity=quantity,
                unit_price=10,
                freight_amount=freight,
                tax_amount=tax,
                extended_amount=amount,
            )
        )
    db.add_all([usd, eur])
    db.commit()

    by_vendor = spend_analysis(db, SpendDimension.VENDOR)
    by_store = spend_analysis(db, SpendDimension.STORE, store_number="001")
    by_department = spend_analysis(db, SpendDimension.DEPARTMENT)
    by_product_code = spend_analysis(db, SpendDimension.PRODUCT_CODE)

    assert [(item.currency, item.amount) for item in by_vendor.metrics] == [
        ("USD", 107),
        ("EUR", 20),
    ]
    assert len(by_store.metrics) == 1
    assert by_store.metrics[0].dimension_key == "001"
    assert by_store.metrics[0].quantity == 10
    assert by_store.metrics[0].amount == 107
    assert {item.dimension_key for item in by_department.metrics} == {"FURNITURE"}
    assert {item.dimension_key for item in by_product_code.metrics} == {"SKU-SPEND"}


def test_workflow_analytics_measures_only_completed_cycle_times(db: Session) -> None:
    started = datetime(2026, 6, 1, tzinfo=UTC)
    instances = [
        WorkflowInstance(
            workflow_code="BPP_PURCHASING",
            workflow_version=1,
            entity_type="purchase_request",
            entity_id="request-1",
            current_state="approved",
            status="complete",
            context={},
            started_by="requester@example.com",
            updated_by="approver@example.com",
            started_at=started,
            updated_at=started + timedelta(seconds=100),
        ),
        WorkflowInstance(
            workflow_code="BPP_PURCHASING",
            workflow_version=1,
            entity_type="purchase_request",
            entity_id="request-2",
            current_state="rejected",
            status="complete",
            context={},
            started_by="requester@example.com",
            updated_by="rejector@example.com",
            started_at=started,
            updated_at=started + timedelta(seconds=300),
        ),
        WorkflowInstance(
            workflow_code="BPP_PURCHASING",
            workflow_version=1,
            entity_type="purchase_request",
            entity_id="request-3",
            current_state="department_review",
            status="active",
            context={},
            started_by="requester@example.com",
            updated_by="requester@example.com",
            started_at=started,
            updated_at=started + timedelta(days=10),
        ),
    ]
    db.add_all(instances)
    db.add_all(
        [
            EventSnapshot(
                event_type="workflow.advanced",
                entity_type="purchase_request",
                entity_id="request-1",
                actor="approver@example.com",
                payload={"workflow_code": "BPP_PURCHASING", "action": "department_approve"},
            ),
            EventSnapshot(
                event_type="workflow.advanced",
                entity_type="purchase_request",
                entity_id="request-2",
                actor="rejector@example.com",
                payload={"workflow_code": "BPP_PURCHASING", "action": "reject"},
            ),
        ]
    )
    db.commit()

    metric = workflow_analytics(db).workflows[0]

    assert metric.instance_count == 3
    assert metric.active_count == 1
    assert metric.completed_count == 2
    assert metric.transition_count == 2
    assert metric.approval_count == 1
    assert metric.rejection_count == 1
    assert metric.average_completion_seconds == Decimal("200.00")
    assert metric.median_completion_seconds == Decimal("200.00")
    assert metric.p90_completion_seconds == Decimal("300.00")
    assert [
        (item.actor, item.approval_count, item.rejection_count) for item in metric.approval_actors
    ] == [
        ("approver@example.com", 1, 0),
        ("rejector@example.com", 0, 1),
    ]


def test_scheduled_report_generation_is_due_once_and_content_addressed(
    db: Session, tmp_path
) -> None:
    order = _order("USD", 20, "PO-REPORT")
    order.lines.append(
        PurchaseOrderLine(
            source_request_id="request-report",
            source_line_id=1,
            store_number="001",
            product_code="SKU-REPORT",
            product_name="Report product",
            quantity=2,
            unit_price=10,
            freight_amount=0,
            tax_amount=0,
            extended_amount=20,
        )
    )
    db.add(order)
    db.commit()
    now = datetime(2026, 6, 29, 12, tzinfo=UTC)
    schedule = create_report_schedule(
        db,
        AnalyticsReportScheduleCreate(
            name="Daily vendor spend",
            report_type=AnalyticsReportType.SPEND,
            parameters={"group_by": "vendor"},
            interval_minutes=1440,
            next_run_at=now,
        ),
        "analyst@example.com",
    )

    runs = run_due_reports(db, "scheduler@example.com", str(tmp_path), now)
    repeated = run_due_reports(db, "scheduler@example.com", str(tmp_path), now)

    assert len(runs) == 1
    assert repeated == []
    assert runs[0].status == "completed"
    assert runs[0].size_bytes and runs[0].size_bytes > 0
    assert len(runs[0].sha256 or "") == 64
    report_path = report_run_path(runs[0], str(tmp_path))
    assert report_path.suffix == ".xlsx"
    assert runs[0].content_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    spend_workbook = load_workbook(report_path, data_only=True)
    assert spend_workbook.sheetnames == ["Dashboard", "Spend Detail"]
    assert spend_workbook["Dashboard"]["A1"].value.startswith("BTSP Executive Analytics")
    assert spend_workbook["Spend Detail"]["A1"].value == "Vendor"
    assert schedule.next_run_at.replace(tzinfo=UTC) == now + timedelta(days=1)

    inventory_export = render_analytics_report(db, AnalyticsReportType.INVENTORY_POSITION, {})
    inventory_workbook = load_workbook(BytesIO(inventory_export), data_only=True)
    assert inventory_workbook.sheetnames == ["Dashboard", "Inventory Detail"]
    assert inventory_workbook["Inventory Detail"]["A1"].value == "Store Number"

    executive_export = render_analytics_report(db, AnalyticsReportType.EXECUTIVE_PACK, {})
    executive_workbook = load_workbook(BytesIO(executive_export), data_only=True)
    assert executive_workbook.sheetnames == [
        "Dashboard",
        "Spend by Vendor",
        "Spend by Department",
        "Product Performance",
        "Vendor Scorecards",
        "Workflow Performance",
        "Inventory Position",
    ]
    assert len(executive_workbook["Dashboard"]._charts) == 1


def test_report_exports_escape_spreadsheet_formulas_and_verify_artifacts(
    db: Session, tmp_path
) -> None:
    order = _order("USD", 10, "PO-SAFE")
    order.vendor_code = "=DANGEROUS()"
    order.lines.append(
        PurchaseOrderLine(
            source_request_id="request-safe",
            source_line_id=1,
            store_number="001",
            product_code="SKU-SAFE",
            product_name="Safe product",
            quantity=1,
            unit_price=10,
            freight_amount=0,
            tax_amount=0,
            extended_amount=10,
        )
    )
    db.add(order)
    db.commit()

    content = render_analytics_report(db, AnalyticsReportType.SPEND, {"group_by": "vendor"})
    workbook = load_workbook(BytesIO(content), data_only=False)
    assert workbook["Spend Detail"]["A2"].value == "'=DANGEROUS()"

    now = datetime(2026, 6, 29, 12, tzinfo=UTC)
    create_report_schedule(
        db,
        AnalyticsReportScheduleCreate(
            name="Integrity check",
            report_type=AnalyticsReportType.SPEND,
            parameters={"group_by": "vendor"},
            interval_minutes=60,
            next_run_at=now,
        ),
        "analyst@example.com",
    )
    run = run_due_reports(db, "scheduler@example.com", str(tmp_path), now)[0]
    path = report_run_path(run, str(tmp_path))
    path.write_bytes(path.read_bytes() + b"tampered")
    with pytest.raises(ValueError, match="integrity"):
        report_run_path(run, str(tmp_path))


def test_report_schedule_rejects_unknown_or_invalid_parameters() -> None:
    with pytest.raises(ValueError, match="Unsupported report parameters"):
        AnalyticsReportScheduleCreate(
            name="Unknown parameter",
            report_type=AnalyticsReportType.SPEND,
            parameters={"password": "secret"},
            interval_minutes=60,
        )
    with pytest.raises(ValueError, match="positive integer"):
        AnalyticsReportScheduleCreate(
            name="Bad minimum",
            report_type=AnalyticsReportType.VENDOR_SCORECARDS,
            parameters={"minimum_orders": "0"},
            interval_minutes=60,
        )
