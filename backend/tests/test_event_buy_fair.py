from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.db.session import Base
from app.models import (  # noqa: F401
    analytics,
    attachment,
    catalog,
    configuration,
    event_management,
    event_snapshot,
    identity,
    invoice_intake,
    notification,
    purchase_order,
    purchasing,
    receiving,
    store,
    vendor_integration,
    workflow,
)
from app.models.catalog import CatalogProduct, CatalogVendor
from app.models.event_management import (
    EventMembership,
    EventSettlementEvent,
    ManagedEvent,
    ManagedSubEvent,
    VendorHallBooth,
    VendorHallEvent,
    VendorHallInventoryItem,
)
from app.models.identity import Role, User
from app.models.store import Store
from app.schemas.event_buy_fair import EventBuyFairOrderCreate
from app.schemas.order_lifecycle import LifecycleLineWrite
from app.services.event_buy_fair_service import (
    EventBuyFairError,
    buy_fair_workspace,
    cancel_buy_fair_order,
    create_buy_fair_orders,
    event_buy_fair_export_rows,
    event_buy_fair_summary,
    require_buy_fair_order,
    sub_event_buy_fair_summary,
)
from app.services.event_order_backup_service import export_event_order_backup
from app.services.order_lifecycle_service import submit_vendor_request


def test_vendor_buy_fair_prioritizes_booth_models_and_creates_standard_requests() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        vendor = CatalogVendor(
            vendor_code="FAIR-VENDOR",
            name="Fair Vendor",
            is_active=True,
            source_file="test.xlsx",
        )
        user = User(
            email="vendor@example.com",
            display_name="Event Vendor",
            password_hash="test",
            vendor_code="FAIR-VENDOR",
            is_active=True,
        )
        requester = User(
            email="buddys.requester@example.com",
            display_name="Buddy's Requester",
            password_hash="test",
            entity_code="ENTITY-SOUTH",
            region_code="SOUTH",
            is_active=True,
            roles=[Role(code="FRANCHISE_OPERATOR", name="Franchise Operator")],
        )
        event = ManagedEvent(
            name="Summer Buy Fair",
            slug="summer-buy-fair",
            status="published",
            starts_at=datetime(2027, 7, 1, 13, tzinfo=UTC),
            ends_at=datetime(2027, 7, 2, 21, tzinfo=UTC),
            timezone="America/New_York",
            venue_name="Expo Center",
            address_line1="1 Fair Way",
            city="Orlando",
            state_code="FL",
            postal_code="32801",
            country_code="US",
            created_by="admin@example.com",
        )
        db.add_all([vendor, user, requester, event])
        db.flush()
        sub_event = ManagedSubEvent(
            event_id=event.id,
            name="Vendor Buy Fair",
            starts_at=event.starts_at,
            ends_at=event.ends_at,
            location="Vendor Hall",
            status="published",
            module_codes=["vendor-buy-fair"],
        )
        membership = EventMembership(
            event_id=event.id,
            user_id=user.id,
            membership_type="vendor",
            vendor_code=vendor.vendor_code,
            module_codes=["vendor-buy-fair"],
            is_active=True,
        )
        db.add_all(
            [
                sub_event,
                membership,
                Store(
                    store_number="101",
                    name="Orlando Store",
                    region_code="SOUTH",
                    entity_code="ENTITY-SOUTH",
                    state_code="FL",
                    is_active=True,
                    is_ordering_enabled=True,
                ),
                Store(
                    store_number="102",
                    name="Tampa Store",
                    region_code="SOUTH",
                    entity_code="ENTITY-SOUTH",
                    state_code="FL",
                    is_active=True,
                    is_ordering_enabled=True,
                ),
                CatalogProduct(
                    product_code="CAT-1",
                    model_number="CAT-1",
                    vendor_code=vendor.vendor_code,
                    name="Catalog Model",
                    unit_price=Decimal("100.00"),
                    currency="USD",
                    minimum_order_quantity=1,
                    is_available=True,
                    is_active=True,
                    source_file="test.xlsx",
                ),
                CatalogProduct(
                    product_code="BOOTH-1",
                    model_number="BOOTH-1",
                    vendor_code=vendor.vendor_code,
                    name="Booth Model",
                    unit_price=Decimal("250.00"),
                    currency="USD",
                    minimum_order_quantity=1,
                    is_available=True,
                    is_active=True,
                    source_file="test.xlsx",
                ),
            ]
        )
        db.flush()
        hall = VendorHallEvent(
            event_id=event.id,
            sub_event_id=sub_event.id,
            status="active",
            created_by="admin@example.com",
        )
        db.add(hall)
        db.flush()
        booth = VendorHallBooth(
            vendor_hall_event_id=hall.id,
            event_id=event.id,
            vendor_code=vendor.vendor_code,
            booth_number="B-12",
            booth_name="Fair Vendor",
        )
        db.add(booth)
        db.flush()
        db.add(
            VendorHallInventoryItem(
                vendor_hall_booth_id=booth.id,
                event_id=event.id,
                vendor_code=vendor.vendor_code,
                model_number="BOOTH-1",
                item_name="Booth Model",
                quantity_expected=1,
                created_by=user.email,
            )
        )
        db.commit()

        workspace = buy_fair_workspace(db, sub_event.id, user)
        assert [item.model_identifier for item in workspace.models] == ["BOOTH-1", "CAT-1"]
        assert workspace.models[0].is_booth_model is True
        assert workspace.requesters[0].display_name == "Buddy's Requester"
        assert not hasattr(workspace.requesters[0], "email")

        created = create_buy_fair_orders(
            db,
            sub_event.id,
            EventBuyFairOrderCreate(
                requester_id=requester.id,
                store_numbers=["101", "102"],
                expected_delivery_date=date(2027, 8, 1),
                line_items=[LifecycleLineWrite(product_code="BOOTH-1", quantity=2)],
            ),
            user,
        )
        assert [item.order_number for item in created] == [
            "Summer Buy Fair-101-FAIR-VENDOR-001",
            "Summer Buy Fair-102-FAIR-VENDOR-002",
        ]
        assert all(item.workflow_code == "VENDOR_ORDER" for item in created)
        assert created[0].context["event_id"] == event.id

        # “All Stores” is scoped to the requester’s entity, not the whole
        # company. A BEBE/ENTITY-SOUTH requester may use every store in that
        # entity, but cannot submit against another entity.
        requester.region_code = "ALL_STORES"
        db.add(
            Store(
                store_number="999",
                name="Other Entity Store",
                region_code="SOUTH",
                entity_code="ENTITY-NORTH",
                state_code="FL",
                is_active=True,
                is_ordering_enabled=True,
            )
        )
        db.commit()
        with pytest.raises(EventBuyFairError, match="not authorized"):
            create_buy_fair_orders(
                db,
                sub_event.id,
                EventBuyFairOrderCreate(
                    requester_id=requester.id,
                    store_numbers=["999"],
                    expected_delivery_date=date(2027, 8, 1),
                    line_items=[LifecycleLineWrite(product_code="BOOTH-1", quantity=1)],
                ),
                user,
            )

        submit_vendor_request(db, created[0], user.email)
        cancel_buy_fair_order(db, created[1], user.email)
        replacement = create_buy_fair_orders(
            db,
            sub_event.id,
            EventBuyFairOrderCreate(
                requester_id=requester.id,
                store_numbers=["102"],
                expected_delivery_date=date(2027, 8, 1),
                line_items=[LifecycleLineWrite(product_code="BOOTH-1", quantity=2)],
            ),
            user,
        )
        assert replacement[0].order_number == "Summer Buy Fair-102-FAIR-VENDOR-003"
        second_sub_event = ManagedSubEvent(
            event_id=event.id,
            name="Second Vendor Buy Fair",
            starts_at=event.starts_at,
            ends_at=event.ends_at,
            location="Second Hall",
            status="published",
            module_codes=["vendor-buy-fair"],
        )
        db.add(second_sub_event)
        db.commit()
        other_sub_event_orders = create_buy_fair_orders(
            db,
            second_sub_event.id,
            EventBuyFairOrderCreate(
                requester_id=requester.id,
                store_numbers=["101"],
                expected_delivery_date=date(2027, 8, 1),
                line_items=[LifecycleLineWrite(product_code="CAT-1", quantity=1)],
            ),
            user,
        )
        assert other_sub_event_orders[0].order_number == ("Summer Buy Fair-101-FAIR-VENDOR-004")
        workspace = buy_fair_workspace(db, sub_event.id, user)
        assert workspace.order_count == 2
        assert workspace.total_units == 4
        assert workspace.total_volume == Decimal("1000.00")
        assert any(item.status == "submitted_to_purchasing" for item in workspace.orders)
        sub_event_summary = sub_event_buy_fair_summary(db, sub_event.id)
        assert sub_event_summary is not None
        assert sub_event_summary.sub_event_id == sub_event.id
        assert sub_event_summary.order_count == 2
        event_summary_with_other_sub_event = event_buy_fair_summary(db, event.id)
        assert event_summary_with_other_sub_event is not None
        assert event_summary_with_other_sub_event.order_count == 3
        db.delete(other_sub_event_orders[0])
        db.commit()
        summary = event_buy_fair_summary(db, event.id)
        assert summary is not None
        assert summary.vendor_count == 1
        assert summary.order_count == 2
        assert summary.total_volume == Decimal("1000.00")
        assert summary.vendors[0].vendor_code == "FAIR-VENDOR"
        assert summary.vendors[0].submitted_count == 1
        assert [item.order_number for item in summary.orders] == [
            "Summer Buy Fair-102-FAIR-VENDOR-003",
            "Summer Buy Fair-101-FAIR-VENDOR-001",
        ]
        export = event_buy_fair_export_rows(db, event.id)
        assert export is not None
        exported_event, rows = export
        assert exported_event.id == event.id
        assert rows[0][1] == "order_number"
        assert rows[1][1] == "Summer Buy Fair-102-FAIR-VENDOR-003"

        backup = export_event_order_backup(db, event.id)
        assert backup is not None
        _, content = backup
        workbook = load_workbook(BytesIO(content), data_only=True)
        assert "Entity ENTITY-SOUTH" in workbook.sheetnames
        entity_rows = list(workbook["Entity ENTITY-SOUTH"].iter_rows(min_row=2, values_only=True))
        assert len(entity_rows) == 3
        assert [row[6] for row in entity_rows] == ["101", "102", "102"]
        assert {row[5] for row in entity_rows} == {"SOUTH"}
        assert {row[7] for row in entity_rows} == {"FAIR-VENDOR"}
        assert {row[15] for row in entity_rows} == {
            "cancelled_by_vendor",
            "submitted_to_purchasing",
            "vendor_draft",
        }
        db.add(
            EventSettlementEvent(
                event_id=event.id,
                status="closed",
                created_by="admin@example.com",
            )
        )
        db.commit()
        with pytest.raises(EventBuyFairError, match="settlement is closed"):
            create_buy_fair_orders(
                db,
                sub_event.id,
                EventBuyFairOrderCreate(
                    requester_id=requester.id,
                    store_numbers=["101"],
                    expected_delivery_date=date(2027, 8, 1),
                    line_items=[LifecycleLineWrite(product_code="BOOTH-1", quantity=1)],
                ),
                user,
            )
        with pytest.raises(EventBuyFairError, match="settlement is closed"):
            require_buy_fair_order(db, sub_event.id, created[0].id, user)
