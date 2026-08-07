from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy import event as sqlalchemy_event
from sqlalchemy.orm import Session

from app.db.session import Base
from app.models import (  # noqa: F401
    analytics,
    attachment,
    catalog,
    configuration,
    event_management,
    event_snapshot,
    identity,
    invoice_intake,
    notification,
    purchase_order,
    purchasing,
    receiving,
    store,
    vendor_integration,
    workflow,
)
from app.models.catalog import CatalogProduct, CatalogVendor
from app.models.event_management import (
    EventCalendarEntry,
    EventEntityOrder,
    EventEntityOrderRevision,
    EventMembership,
    EventOrderReleaseLine,
    EventPoll,
    EventPresentationState,
    EventProductSlide,
    EventProductSlideImage,
    EventSettlementEvent,
    EventStaffTask,
    ManagedEvent,
    ManagedSubEvent,
)
from app.models.event_snapshot import EventSnapshot
from app.models.identity import User
from app.models.notification import NotificationEvent
from app.models.purchase_order import PurchaseOrder, PurchaseOrderLine, PurchaseOrderSource
from app.models.purchasing import PurchaseRequest
from app.models.store import Store
from app.schemas.event_announcement import EventAnnouncementWrite
from app.schemas.event_calendar import EventCalendarEntryWrite
from app.schemas.event_management import (
    EventCancellationWrite,
    EventMembershipCreate,
    EventMembershipRoleUpdate,
    EventMembershipUpdate,
    EventSubEventRegistrationWrite,
    EventWrite,
    SubEventModulesWrite,
    SubEventWrite,
)
from app.schemas.event_order_review import EventOrderReviewDecision
from app.schemas.event_ordering import EventEntityOrderWrite
from app.schemas.event_poll import EventPollCreate
from app.schemas.event_product_slide import EventProductSlideWrite
from app.schemas.event_staff_task import EventStaffTaskStatusWrite, EventStaffTaskWrite
from app.schemas.event_vendor_booth import EventVendorBoothWrite
from app.services.event_access_service import (
    event_operations_are_locked,
    event_window_open_for_user,
)
from app.services.event_announcement_service import my_announcements, save_announcement
from app.services.event_attendance_service import (
    EventAttendanceError,
    attendance_roster,
    my_attendance_passes,
    update_attendance,
    update_attendance_by_pass_code,
)
from app.services.event_calendar_service import (
    EventCalendarError,
    my_calendar,
    remove_calendar_entry,
    save_calendar_entry,
)
from app.services.event_live_insights_service import live_insights
from app.services.event_management_service import (
    EventManagementError,
    add_membership,
    add_sub_event,
    assign_membership_sub_events,
    cancel_event,
    create_event,
    list_active_events,
    list_archived_events,
    list_member_events,
    publish_event,
    remove_event,
    remove_sub_event,
    update_event,
    update_membership,
    update_membership_role,
    update_sub_event,
    update_sub_event_modules,
)
from app.services.event_order_backup_service import export_event_order_backup
from app.services.event_order_review_service import (
    EventOrderReviewError,
    decide_order,
    release_approved_orders,
    review_summary,
)
from app.services.event_ordering_service import (
    EventOrderingError,
    _order_capacity,
    ordering_workspace,
    submit_entity_order,
)
from app.services.event_poll_service import (
    EventPollError,
    active_poll,
    cast_vote,
    create_poll,
    set_poll_status,
)
from app.services.event_presentation_service import (
    EventPresentationError,
    _ensure_projectable,
    control_presentation,
    get_live_analytics,
    get_presentation,
)
from app.services.event_product_slide_service import (
    EventProductSlideError,
    create_slide,
    list_slides,
    reorder_slides,
    save_slide_image,
    save_slide_vendor_logo,
)
from app.services.event_staff_task_report_service import export_event_staff_tasks
from app.services.event_staff_task_service import (
    EventStaffTaskError,
    attach_task_evidence,
    enqueue_due_task_reminders,
    my_tasks,
    save_event_task,
    task_evidence_content,
    update_task_status,
)
from app.services.event_vendor_booth_service import (
    create_event_only_vendor,
    list_available_vendors,
    my_booths,
    save_booth,
    vendor_update_booth,
)
from app.services.notification_service import list_user_notification_events


def _event() -> EventWrite:
    return EventWrite(
        name="Vendor Fair 2027",
        slug="vendor-fair-2027",
        starts_at=datetime(2027, 4, 5, 12, tzinfo=UTC),
        ends_at=datetime(2027, 4, 7, 22, tzinfo=UTC),
        venue_name="Convention Center",
        address_line1="1 Expo Way",
        city="Orlando",
        state_code="FL",
        postal_code="32801",
    )


def test_full_screen_image_slide_requires_projectable_image() -> None:
    payload = EventProductSlideWrite(
        slide_type="filler",
        filler_category="full_screen_image",
        name="Opening title card",
        status="ready",
    )
    slide = EventProductSlide(**payload.model_dump(), created_by="admin@example.com")

    with pytest.raises(EventPresentationError, match="Upload an image"):
        _ensure_projectable(slide)

    slide.image = EventProductSlideImage(
        filename="opening.webp",
        content_type="image/webp",
        content=b"image-content",
        uploaded_by="admin@example.com",
    )
    _ensure_projectable(slide)


def test_event_publication_opens_attendee_window_and_is_audited() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        created = create_event(db, _event(), "admin@example.com")
        updated = add_sub_event(
            db,
            created.id,
            SubEventWrite(
                name="Buying session",
                starts_at=datetime(2027, 4, 5, 14, tzinfo=UTC),
                ends_at=datetime(2027, 4, 5, 15, tzinfo=UTC),
                location="Main stage",
            ),
        )
        assert updated is not None
        db.add(
            Store(
                store_number="101",
                name="Publication Test Store",
                region_code="R1",
                entity_code="ENTITY-1",
                is_active=True,
                is_ordering_enabled=True,
            )
        )
        db.commit()
        updated = add_membership(
            db,
            created.id,
            EventMembershipCreate(
                email="franchise-publish@example.com",
                display_name="Franchise Attendee",
                password="Event-Publish-Password!",
                membership_type="franchise_representative",
                entity_code="ENTITY-1",
            ),
        )
        assert updated is not None
        attendee = db.scalar(select(User).where(User.email == "franchise-publish@example.com"))
        assert attendee is not None
        during_event = datetime(2027, 4, 6, 12, tzinfo=UTC)
        assert not event_window_open_for_user(db, created.id, attendee.id, during_event)

        published = publish_event(db, created.id, "admin@example.com")
        assert published is not None
        assert published.status == "published"
        assert published.sub_events[0].status == "published"
        assert event_window_open_for_user(db, created.id, attendee.id, during_event)

        replay = publish_event(db, created.id, "admin@example.com")
        assert replay is not None and replay.status == "published"
        snapshots = db.scalars(
            select(EventSnapshot).where(
                EventSnapshot.entity_id == created.id,
                EventSnapshot.event_type == "event.published",
            )
        ).all()
        assert len(snapshots) == 1
        assert snapshots[0].actor == "admin@example.com"
        assert snapshots[0].payload["sub_events_published"] == 1


def test_multiple_franchise_representatives_share_an_entity_and_are_fully_editable() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        event = create_event(db, _event(), "admin@example.com")
        db.add(
            Store(
                store_number="101",
                name="Shared Entity Store",
                region_code="R1",
                entity_code="ENTITY-1",
                is_active=True,
                is_ordering_enabled=True,
            )
        )
        db.commit()

        for index in (1, 2):
            added = add_membership(
                db,
                event.id,
                EventMembershipCreate(
                    email=f"franchise-{index}@example.com",
                    display_name=f"Franchise Representative {index}",
                    password=f"Franchise-Password-{index}!",
                    membership_type="franchise_representative",
                    entity_code="entity-1",
                ),
            )
            assert added is not None

        memberships = db.scalars(
            select(EventMembership).where(
                EventMembership.event_id == event.id,
                EventMembership.entity_code == "ENTITY-1",
            )
        ).all()
        assert len(memberships) == 2

        updated = update_membership(
            db,
            event.id,
            memberships[0].id,
            EventMembershipUpdate(
                email="renamed-franchise@example.com",
                display_name="Renamed Franchise Representative",
                membership_type="franchise_representative",
                entity_code="entity-1",
                task_scope="Ordering and meeting support",
                is_active=False,
            ),
            "admin@example.com",
        )
        assert updated is not None
        edited = db.get(EventMembership, memberships[0].id)
        assert edited is not None
        assert edited.entity_code == "ENTITY-1"
        assert edited.task_scope == "Ordering and meeting support"
        assert not edited.is_active
        assert edited.user.email == "renamed-franchise@example.com"
        assert edited.user.display_name == "Renamed Franchise Representative"


def test_event_cancellation_closes_access_and_permanent_removal_is_guarded() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    sqlalchemy_event.listen(
        engine,
        "connect",
        lambda connection, _record: connection.execute("PRAGMA foreign_keys=ON"),
    )
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        completed = create_event(
            db,
            _event().model_copy(
                update={
                    "name": "Completed Vendor Fair 2026",
                    "slug": "completed-vendor-fair-2026",
                    "status": "completed",
                    "starts_at": datetime(2026, 4, 5, 12, tzinfo=UTC),
                    "ends_at": datetime(2026, 4, 7, 22, tzinfo=UTC),
                }
            ),
            "admin@example.com",
        )
        assert event_operations_are_locked(db, completed.id)
        with pytest.raises(EventManagementError, match="read-only"):
            update_event(
                db,
                completed.id,
                _event().model_copy(
                    update={
                        "name": "Attempted Completed Event Rename",
                        "slug": "attempted-completed-event-rename",
                    }
                ),
            )
        created = create_event(db, _event(), "admin@example.com")
        assert [event.id for event in list_active_events(db)] == [created.id]
        assert [event.id for event in list_archived_events(db)] == [completed.id]
        updated = add_sub_event(
            db,
            created.id,
            SubEventWrite(
                name="Planning session",
                starts_at=datetime(2027, 4, 5, 14, tzinfo=UTC),
                ends_at=datetime(2027, 4, 5, 15, tzinfo=UTC),
                location="Main stage",
                module_codes=["polling"],
            ),
        )
        assert updated is not None
        event_only_vendor = create_event_only_vendor(db, created.id, "Cancellation Services")
        assert event_only_vendor is not None
        event_only_vendor_code = event_only_vendor.vendor_code
        updated = add_membership(
            db,
            created.id,
            EventMembershipCreate(
                email="staff-cancel@example.com",
                display_name="Cancellation Staff",
                password="Event-Only-Password!",
                membership_type="staff",
            ),
        )
        assert updated is not None
        staff = db.scalar(select(User).where(User.email == "staff-cancel@example.com"))
        assert staff is not None
        assert event_window_open_for_user(db, created.id, staff.id)
        staff_task = save_event_task(
            db,
            created.id,
            EventStaffTaskWrite(
                assigned_membership_id=updated.memberships[0].id,
                sub_event_id=updated.sub_events[0].id,
                title="Cancellation test task",
            ),
            "admin@example.com",
        )
        assert staff_task is not None
        assert [task.id for task in my_tasks(db, staff)] == [staff_task.id]
        poll = create_poll(
            db,
            updated.sub_events[0].id,
            EventPollCreate(question="Continue the session?", options=["Yes", "No"]),
            "admin@example.com",
        )
        assert poll is not None
        poll = set_poll_status(db, poll.id, "open")
        assert poll is not None and poll.status == "open"
        db.add(
            EventPresentationState(
                sub_event_id=updated.sub_events[0].id,
                event_id=created.id,
                status="running",
                ordering_status="open",
                updated_by="presenter@example.com",
            )
        )
        db.commit()
        calendar = save_calendar_entry(
            db,
            created.id,
            EventCalendarEntryWrite(
                entry_type="sub_event",
                sub_event_id=updated.sub_events[0].id,
                title="Planning session",
                starts_at=datetime(2027, 4, 5, 14, tzinfo=UTC),
                ends_at=datetime(2027, 4, 5, 15, tzinfo=UTC),
                visibility_categories=["staff"],
            ),
            "admin@example.com",
        )
        assert calendar is not None

        cancelled = cancel_event(
            db,
            created.id,
            EventCancellationWrite(reason="Venue became unavailable"),
            "admin@example.com",
        )
        assert cancelled is not None
        assert cancelled.status == "cancelled"
        assert cancelled.cancelled_by == "admin@example.com"
        assert cancelled.cancelled_at is not None
        assert cancelled.cancellation_reason == "Venue became unavailable"
        assert cancelled.sub_events[0].status == "cancelled"
        assert list_active_events(db) == []
        assert my_tasks(db, staff) == []
        assert {event.id for event in list_archived_events(db)} == {
            completed.id,
            cancelled.id,
        }
        assert not event_window_open_for_user(db, created.id, staff.id)
        assert db.get(EventCalendarEntry, calendar.id).is_active is False
        presentation_state = db.get(EventPresentationState, updated.sub_events[0].id)
        assert presentation_state.status == "ended"
        assert presentation_state.ordering_status == "closed"
        assert db.get(EventPoll, poll.id).status == "closed"
        with pytest.raises(EventPollError, match="archived"):
            set_poll_status(db, poll.id, "open")
        with pytest.raises(EventCalendarError, match="archived"):
            remove_calendar_entry(db, created.id, calendar.id, "admin@example.com")
        with pytest.raises(EventManagementError, match="read-only"):
            update_event(
                db,
                created.id,
                _event().model_copy(update={"name": "Attempted Cancelled Event Rename"}),
            )
        with pytest.raises(EventPollError, match="cancelled"):
            create_poll(
                db,
                cancelled.sub_events[0].id,
                EventPollCreate(question="Should not open?", options=["Yes", "No"]),
                "admin@example.com",
            )
        with pytest.raises(EventCalendarError, match="cancelled"):
            save_calendar_entry(
                db,
                created.id,
                EventCalendarEntryWrite(
                    entry_type="text",
                    title="Should not publish",
                    starts_at=datetime(2027, 4, 5, 16, tzinfo=UTC),
                    ends_at=datetime(2027, 4, 5, 17, tzinfo=UTC),
                    visibility_categories=["staff"],
                ),
                "admin@example.com",
            )

        assert remove_event(db, created.id, "admin@example.com")
        assert db.get(ManagedEvent, created.id) is None
        assert (
            db.scalar(
                select(CatalogVendor).where(CatalogVendor.vendor_code == event_only_vendor_code)
            )
            is None
        )
        lifecycle_events = db.scalars(
            select(EventSnapshot).where(EventSnapshot.entity_id == created.id)
        ).all()
        assert {item.event_type for item in lifecycle_events} == {
            "event.cancelled",
            "event.deleted",
        }
        cancelled_snapshot = next(
            item for item in lifecycle_events if item.event_type == "event.cancelled"
        )
        assert cancelled_snapshot.payload["presentations_ended"] == 1
        assert cancelled_snapshot.payload["polls_closed"] == 1
        assert cancelled_snapshot.payload["calendar_entries_hidden"] == 1
        deleted_snapshot = next(
            item for item in lifecycle_events if item.event_type == "event.deleted"
        )
        assert deleted_snapshot.actor == "admin@example.com"
        assert deleted_snapshot.payload["cancellation_reason"] == "Venue became unavailable"


def test_sub_event_can_be_permanently_removed_from_draft_event() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    sqlalchemy_event.listen(
        engine,
        "connect",
        lambda connection, _record: connection.execute("PRAGMA foreign_keys=ON"),
    )
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        created = create_event(db, _event(), "admin@example.com")
        updated = add_sub_event(
            db,
            created.id,
            SubEventWrite(
                name="Session to remove",
                starts_at=datetime(2027, 4, 5, 14, tzinfo=UTC),
                ends_at=datetime(2027, 4, 5, 15, tzinfo=UTC),
                location="Main stage",
            ),
        )
        assert updated is not None
        calendar = save_calendar_entry(
            db,
            created.id,
            EventCalendarEntryWrite(
                entry_type="sub_event",
                sub_event_id=updated.sub_events[0].id,
                title="Session to remove",
                starts_at=datetime(2027, 4, 5, 14, tzinfo=UTC),
                ends_at=datetime(2027, 4, 5, 15, tzinfo=UTC),
                visibility_categories=["admin"],
            ),
            "admin@example.com",
        )
        assert calendar is not None
        text_entry = save_calendar_entry(
            db,
            created.id,
            EventCalendarEntryWrite(
                entry_type="text",
                title="Temporary note",
                starts_at=datetime(2027, 4, 5, 16, tzinfo=UTC),
                ends_at=datetime(2027, 4, 5, 17, tzinfo=UTC),
                visibility_categories=["admin"],
            ),
            "admin@example.com",
        )
        assert text_entry is not None
        assert remove_calendar_entry(db, created.id, text_entry.id, "admin@example.com")
        removed = remove_sub_event(
            db,
            created.id,
            updated.sub_events[0].id,
            "admin@example.com",
        )
        assert removed is not None
        assert removed.sub_events == []
        assert db.get(EventCalendarEntry, calendar.id) is None
        snapshots = db.scalars(select(EventSnapshot)).all()
        assert {item.event_type for item in snapshots} == {
            "event.calendar_entry.deleted",
            "event.sub_event.deleted",
        }


def test_event_calendar_and_event_only_account_are_scoped() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        db.add(
            Store(
                store_number="101",
                name="Entity Store",
                region_code="R1",
                entity_code="ENTITY-1",
                is_ordering_enabled=True,
                is_active=True,
            )
        )
        db.commit()
        created = create_event(db, _event(), "admin@example.com")
        updated = add_sub_event(
            db,
            created.id,
            SubEventWrite(
                name="Opening session",
                starts_at=datetime(2027, 4, 5, 14, tzinfo=UTC),
                ends_at=datetime(2027, 4, 5, 15, tzinfo=UTC),
                location="Main stage",
                module_codes=["check-in", "polling"],
            ),
        )
        assert updated is not None
        assert updated.sub_events[0].module_codes == ["check-in", "polling"]
        with pytest.raises(EventProductSlideError, match="not enabled"):
            list_slides(db, updated.sub_events[0].id)
        updated = update_sub_event(
            db,
            created.id,
            updated.sub_events[0].id,
            SubEventWrite(
                name="Opening session",
                starts_at=datetime(2027, 4, 5, 14, tzinfo=UTC),
                ends_at=datetime(2027, 4, 5, 15, tzinfo=UTC),
                location="Ballroom A",
                module_codes=updated.sub_events[0].module_codes,
            ),
        )
        assert updated is not None
        assert updated.sub_events[0].location == "Ballroom A"

        updated = add_membership(
            db,
            created.id,
            EventMembershipCreate(
                email="attendee@example.com",
                display_name="Event Attendee",
                password="Event-Only-Password!",
                membership_type="franchise_representative",
                entity_code="ENTITY-1",
                module_codes=["check-in"],
            ),
        )
        assert updated is not None
        attendee = db.scalar(select(User).where(User.email == "attendee@example.com"))
        assert attendee is not None
        assert attendee.roles == []
        assert updated.memberships[0].module_codes == ["check-in"]
        assert updated.memberships[0].entity_code == "ENTITY-1"
        assigned = list_member_events(db, attendee.id)
        assert len(assigned) == 1
        assert len(assigned[0].memberships) == 1
        assert assigned[0].memberships[0].email == "attendee@example.com"
        membership_id = updated.memberships[0].id
        scoped = assign_membership_sub_events(
            db,
            created.id,
            membership_id,
            EventSubEventRegistrationWrite(sub_event_ids=[]),
            "admin@example.com",
        )
        assert scoped is not None
        assert list_member_events(db, attendee.id)[0].sub_events == []
        assert attendance_roster(db, updated.sub_events[0].id).registered_total == 0
        scoped = assign_membership_sub_events(
            db,
            created.id,
            membership_id,
            EventSubEventRegistrationWrite(
                sub_event_ids=[updated.sub_events[0].id],
                roles={updated.sub_events[0].id: "dockmaster"},
            ),
            "admin@example.com",
        )
        assert scoped is not None
        assert len(list_member_events(db, attendee.id)[0].sub_events) == 1
        assert (
            list_member_events(db, attendee.id)[0]
            .memberships[0]
            .sub_event_roles[updated.sub_events[0].id]
            == "dockmaster"
        )
        passes = my_attendance_passes(db, attendee)
        assert len(passes) == 1
        assert passes[0].pass_code.startswith("BTSP-")
        assert passes[0].sub_events[0].check_in_enabled is True
        assert passes[0].sub_events[0].status == "registered"
        scanned = update_attendance_by_pass_code(
            db,
            updated.sub_events[0].id,
            passes[0].pass_code.lower(),
            "checked_in",
            "staff@example.com",
        )
        assert scanned is not None
        assert scanned.member.email == "attendee@example.com"
        assert scanned.member.status == "checked_in"
        text_entry = save_calendar_entry(
            db,
            created.id,
            EventCalendarEntryWrite(
                entry_type="text",
                title="Welcome reception",
                starts_at=datetime(2027, 4, 5, 18, tzinfo=UTC),
                ends_at=datetime(2027, 4, 5, 19, tzinfo=UTC),
                location="Atrium",
                visibility_categories=["franchise_representative"],
            ),
            "admin@example.com",
        )
        linked_entry = save_calendar_entry(
            db,
            created.id,
            EventCalendarEntryWrite(
                entry_type="sub_event",
                sub_event_id=updated.sub_events[0].id,
                title="Linked session",
                starts_at=datetime(2027, 4, 5, 14, tzinfo=UTC),
                ends_at=datetime(2027, 4, 5, 15, tzinfo=UTC),
                visibility_categories=["franchise_representative"],
            ),
            "admin@example.com",
        )
        assert text_entry is not None and linked_entry is not None
        assert len(my_calendar(db, attendee, False)) == 2
        event_alert = save_announcement(
            db,
            created.id,
            EventAnnouncementWrite(
                title="Badge pickup moved",
                body="Use the north registration desk.",
                severity="important",
                visibility_categories=["franchise_representative"],
                publishes_at=datetime(2026, 1, 1, 12, tzinfo=UTC),
            ),
            "admin@example.com",
        )
        sub_event_alert = save_announcement(
            db,
            created.id,
            EventAnnouncementWrite(
                sub_event_id=updated.sub_events[0].id,
                title="Opening room change",
                body="Opening session moved to Ballroom A.",
                severity="urgent",
                visibility_categories=["franchise_representative"],
                publishes_at=datetime(2026, 1, 1, 12, tzinfo=UTC),
            ),
            "admin@example.com",
        )
        executive_alert = save_announcement(
            db,
            created.id,
            EventAnnouncementWrite(
                title="Executive dinner",
                body="Private dinner details.",
                severity="info",
                visibility_categories=["executive"],
                publishes_at=datetime(2026, 1, 1, 12, tzinfo=UTC),
            ),
            "admin@example.com",
        )
        assert event_alert is not None and sub_event_alert is not None
        assert executive_alert is not None
        assert {item.title for item in my_announcements(db, attendee)} == {
            "Badge pickup moved",
            "Opening room change",
        }
        scoped = assign_membership_sub_events(
            db,
            created.id,
            membership_id,
            EventSubEventRegistrationWrite(sub_event_ids=[]),
            "admin@example.com",
        )
        assert scoped is not None
        assert {item.title for item in my_announcements(db, attendee)} == {"Badge pickup moved"}
        with pytest.raises(EventAttendanceError, match="not assigned"):
            update_attendance_by_pass_code(
                db,
                updated.sub_events[0].id,
                passes[0].pass_code,
                "checked_out",
                "staff@example.com",
            )
        scoped = assign_membership_sub_events(
            db,
            created.id,
            membership_id,
            EventSubEventRegistrationWrite(sub_event_ids=[updated.sub_events[0].id]),
            "admin@example.com",
        )
        assert scoped is not None
        changed = update_sub_event_modules(
            db,
            created.id,
            updated.sub_events[0].id,
            SubEventModulesWrite(
                module_codes=[
                    "polling",
                    "presentation",
                    "staff-tasks",
                    "vendor-booths",
                ]
            ),
        )
        assert changed is not None
        assert changed.sub_events[0].module_codes == [
            "polling",
            "presentation",
            "staff-tasks",
            "vendor-booths",
        ]

        with pytest.raises(EventManagementError, match="Unknown event modules"):
            add_sub_event(
                db,
                created.id,
                SubEventWrite(
                    name="Unsupported module",
                    starts_at=datetime(2027, 4, 5, 16, tzinfo=UTC),
                    ends_at=datetime(2027, 4, 5, 17, tzinfo=UTC),
                    location="Main stage",
                    module_codes=["legacy free text"],
                ),
            )
        legacy_sub_event = db.get(ManagedSubEvent, updated.sub_events[0].id)
        assert legacy_sub_event is not None
        legacy_sub_event.module_codes = ["legacy free text"]
        db.commit()
        assert list_member_events(db, attendee.id)[0].sub_events[0].module_codes == [
            "legacy free text"
        ]


def test_vendor_membership_links_main_platform_vendor_and_dates_are_bounded() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        db.add(
            CatalogVendor(
                vendor_code="EXPO",
                name="Expo Vendor",
                is_active=True,
                source_file="test",
            )
        )
        db.add(
            Store(
                store_number="202",
                name="Ordering Entity",
                region_code="R2",
                entity_code="ENTITY-2",
                is_ordering_enabled=True,
                is_active=True,
            )
        )
        db.commit()
        created = create_event(db, _event(), "admin@example.com")

        linked = add_membership(
            db,
            created.id,
            EventMembershipCreate(
                email="vendor@example.com",
                display_name="Expo Vendor",
                password="Vendor-Event-Password!",
                membership_type="vendor",
                vendor_code="EXPO",
            ),
        )
        assert linked is not None
        assert linked.memberships[0].vendor_code == "EXPO"
        vendor_user = db.scalar(select(User).where(User.email == "vendor@example.com"))
        assert vendor_user is not None
        booth = save_booth(
            db,
            created.id,
            EventVendorBoothWrite(
                vendor_code="EXPO",
                booth_name="Expo Vendor Showcase",
                booth_number="A-10",
                location="Main hall",
                status="published",
            ),
            "admin@example.com",
        )
        assert booth is not None
        assert my_booths(db, vendor_user)[0].booth_number == "A-10"
        updated_booth = vendor_update_booth(
            db,
            booth.id,
            EventVendorBoothWrite(
                vendor_code="EXPO",
                booth_name="Expo Vendor Showcase",
                booth_number="A-11",
                location="Main hall",
                description="Updated by vendor.",
                status="published",
            ),
            vendor_user,
        )
        assert updated_booth is not None
        assert updated_booth.booth_number == "A-11"

        with pytest.raises(EventManagementError, match="within the main event"):
            add_sub_event(
                db,
                created.id,
                SubEventWrite(
                    name="Too early",
                    starts_at=datetime(2027, 4, 4, 14, tzinfo=UTC),
                    ends_at=datetime(2027, 4, 5, 15, tzinfo=UTC),
                    location="Elsewhere",
                ),
            )


def test_event_operations_role_update_clears_vendor_scope() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        db.add(
            CatalogVendor(
                vendor_code="ROLE-VENDOR",
                name="Role Vendor",
                is_active=True,
                source_file="test",
            )
        )
        db.commit()
        created = create_event(db, _event(), "admin@example.com")
        linked = add_membership(
            db,
            created.id,
            EventMembershipCreate(
                email="role-user@example.com",
                display_name="Role User",
                password="Event-Role-Password!",
                membership_type="vendor",
                vendor_code="ROLE-VENDOR",
            ),
        )
        assert linked is not None
        membership_id = linked.memberships[0].id

        updated = update_membership_role(
            db,
            created.id,
            membership_id,
            EventMembershipRoleUpdate(membership_type="dockmaster"),
            "admin@example.com",
        )

        assert updated is not None
        membership = db.get(event_management.EventMembership, membership_id)
        assert membership is not None
        assert membership.membership_type == "dockmaster"
        assert membership.vendor_code is None
        assert membership.vendor_codes == []


def test_event_only_service_vendor_is_scoped_and_can_have_a_booth() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        first = create_event(db, _event(), "admin@example.com")
        second_payload = _event().model_copy(update={"name": "Second Fair", "slug": "second-fair"})
        second = create_event(db, second_payload, "admin@example.com")
        vendor = create_event_only_vendor(db, first.id, "Installation Services LLC")
        assert vendor is not None
        assert vendor.is_active is False
        assert vendor in (list_available_vendors(db, first.id) or [])
        assert vendor not in (list_available_vendors(db, second.id) or [])
        for index, name in enumerate(("Alex Installer", "Jordan Installer"), start=1):
            linked = add_membership(
                db,
                first.id,
                EventMembershipCreate(
                    email=f"installer{index}@example.com",
                    display_name=name,
                    password="Vendor-Event-Password!",
                    membership_type="vendor",
                    vendor_code=vendor.vendor_code,
                ),
            )
            assert linked is not None
        vendor_memberships = [
            member for member in linked.memberships if member.vendor_code == vendor.vendor_code
        ]
        assert len(vendor_memberships) == 2
        booth = save_booth(
            db,
            first.id,
            EventVendorBoothWrite(
                vendor_code=vendor.vendor_code,
                booth_name="Installation Services",
                booth_number="S-14",
            ),
            "admin@example.com",
        )
        assert booth is not None
        assert booth.vendor_name == "Installation Services LLC"


def test_event_staff_tasks_are_assigned_and_scoped() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        created = create_event(db, _event(), "admin@example.com")
        updated = add_sub_event(
            db,
            created.id,
            SubEventWrite(
                name="Registration desk",
                starts_at=datetime(2027, 4, 5, 14, tzinfo=UTC),
                ends_at=datetime(2027, 4, 5, 15, tzinfo=UTC),
                location="Lobby",
                module_codes=["check-in"],
            ),
        )
        assert updated is not None
        updated = add_membership(
            db,
            created.id,
            EventMembershipCreate(
                email="staff@example.com",
                display_name="Event Staff",
                password="Staff-Event-Password!",
                membership_type="staff",
            ),
        )
        assert updated is not None
        staff = db.scalar(select(User).where(User.email == "staff@example.com"))
        assert staff is not None
        staff_membership_id = updated.memberships[0].id
        task = save_event_task(
            db,
            created.id,
            EventStaffTaskWrite(
                assigned_membership_id=staff_membership_id,
                sub_event_id=updated.sub_events[0].id,
                title="Open registration desk",
                description="Set up badges and signage.",
                priority="high",
            ),
            "admin@example.com",
        )
        assert task is not None
        assert task.assigned_email == "staff@example.com"
        assert task.sub_event_name == "Registration desk"
        assert my_tasks(db, staff)[0].title == "Open registration desk"
        evidence = attach_task_evidence(
            db,
            task.id,
            "../../registration.png",
            "image/png",
            b"\x89PNG\r\n\x1a\nstaff-task-evidence",
            staff,
        )
        assert evidence is not None
        assert evidence.filename == "registration.png"
        assert my_tasks(db, staff)[0].attachments[0].id == evidence.id
        evidence_source = task_evidence_content(db, task.id, evidence.id, staff)
        assert evidence_source is not None
        assert evidence_source[0] == "registration.png"
        updated = add_membership(
            db,
            created.id,
            EventMembershipCreate(
                email="event-admin@example.com",
                display_name="Event Admin",
                password="Admin-Event-Password!",
                membership_type="admin",
            ),
        )
        assert updated is not None
        task_record = db.get(EventStaffTask, task.id)
        assert task_record is not None
        reminder_now = datetime(2027, 4, 5, 13, tzinfo=UTC)
        task_record.due_at = reminder_now + timedelta(minutes=30)
        db.commit()
        assert enqueue_due_task_reminders(db, reminder_now) == 1
        assert enqueue_due_task_reminders(db, reminder_now) == 0
        reminder = db.scalar(
            select(NotificationEvent).where(
                NotificationEvent.entity_id == task.id,
                NotificationEvent.event_type == "event_staff_task.reminder",
            )
        )
        assert reminder is not None
        assert set(reminder.resolved_recipients) == {
            "staff@example.com",
            "event-admin@example.com",
        }
        assert reminder.action_href == f"/events/sub-event/{updated.sub_events[0].id}"
        changed = update_task_status(
            db, task.id, EventStaffTaskStatusWrite(status="in_progress"), staff
        )
        assert changed is not None and changed.status == "in_progress"
        blocked = update_task_status(
            db,
            task.id,
            EventStaffTaskStatusWrite(
                status="blocked",
                note="Waiting for the registration badge shipment.",
            ),
            staff,
        )
        assert blocked is not None and blocked.status == "blocked"
        assert blocked.status_note == "Waiting for the registration badge shipment."
        resumed = update_task_status(
            db, task.id, EventStaffTaskStatusWrite(status="in_progress"), staff
        )
        assert resumed is not None and resumed.status == "in_progress"
        assert resumed.status_note is None

        outsider = User(
            email="outsider@example.com",
            display_name="Outsider",
            password_hash="x",
            is_active=True,
        )
        db.add(outsider)
        db.commit()
        assert task_evidence_content(db, task.id, evidence.id, outsider) is None
        with pytest.raises(EventStaffTaskError, match="not assigned"):
            update_task_status(db, task.id, EventStaffTaskStatusWrite(status="done"), outsider)
        completed = update_task_status(
            db,
            task.id,
            EventStaffTaskStatusWrite(
                status="done",
                note="Registration desk is open and badges are organized.",
            ),
            staff,
        )
        assert completed is not None and completed.status == "done"
        assert completed.completed_at is not None
        assert completed.completed_by == "staff@example.com"
        assert completed.status_note == "Registration desk is open and badges are organized."
        completion_event = db.scalar(
            select(NotificationEvent).where(
                NotificationEvent.entity_id == task.id,
                NotificationEvent.event_type == "event_staff_task.completed",
            )
        )
        assert completion_event is not None
        assert "event-admin@example.com" in completion_event.resolved_recipients
        assert completion_event.action_href == (f"/events/sub-event/{updated.sub_events[0].id}")
        admin_notifications = list_user_notification_events(db, "event-admin@example.com", limit=20)
        assert completion_event.id in {item.id for item in admin_notifications}
        assert list_user_notification_events(db, "outsider@example.com", limit=20) == []
        with pytest.raises(EventStaffTaskError, match="cannot move from done"):
            update_task_status(db, task.id, EventStaffTaskStatusWrite(status="done"), staff)
        completion_events = db.scalars(
            select(NotificationEvent).where(
                NotificationEvent.entity_id == task.id,
                NotificationEvent.event_type == "event_staff_task.completed",
            )
        ).all()
        assert len(completion_events) == 1
        cancelled = save_event_task(
            db,
            created.id,
            EventStaffTaskWrite(
                assigned_membership_id=task.assigned_membership_id,
                sub_event_id=task.sub_event_id,
                title=task.title,
                description=task.description,
                priority=task.priority,
                status="cancelled",
                task_phase=task.task_phase,
                due_at=task.due_at,
            ),
            "event-admin@example.com",
            task.id,
        )
        assert cancelled is not None and cancelled.status == "cancelled"
        cancellation_event = db.scalar(
            select(NotificationEvent).where(
                NotificationEvent.entity_id == task.id,
                NotificationEvent.event_type == "event_staff_task.cancelled",
            )
        )
        assert cancellation_event is not None
        assert cancellation_event.resolved_recipients == ["staff@example.com"]
        cancellation_audit = db.scalar(
            select(EventSnapshot).where(
                EventSnapshot.entity_id == task.id,
                EventSnapshot.event_type == "event_staff_task.cancelled",
            )
        )
        assert cancellation_audit is not None
        assert cancellation_audit.actor == "event-admin@example.com"
        report = export_event_staff_tasks(db, created.id)
        assert report is not None
        filename, content = report
        assert filename.endswith("-staff-tasks.xlsx")
        workbook = load_workbook(BytesIO(content), data_only=False)
        assert workbook.sheetnames == ["Dashboard", "Task Register", "Staff Workload"]
        assert workbook["Task Register"].max_row == 2
        assert workbook["Task Register"]["N2"].value == 1
        assert workbook["Staff Workload"].max_row == 2
        assert workbook["Staff Workload"]["A2"].value == "Event Staff"
        assert workbook["Staff Workload"]["C2"].value == 1
        assert workbook["Staff Workload"]["I2"].value == 1
        assert workbook["Staff Workload"]["J2"].value == 0


def test_specialized_event_staff_tasks_respect_sub_event_scope() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        created = create_event(db, _event(), "admin@example.com")
        event = add_sub_event(
            db,
            created.id,
            SubEventWrite(
                name="Loadout",
                starts_at=datetime(2027, 4, 7, 18, tzinfo=UTC),
                ends_at=datetime(2027, 4, 7, 21, tzinfo=UTC),
                location="Dock",
                module_codes=["store-loadout", "staff-tasks"],
            ),
        )
        assert event is not None
        event = add_sub_event(
            db,
            created.id,
            SubEventWrite(
                name="Private session",
                starts_at=datetime(2027, 4, 6, 18, tzinfo=UTC),
                ends_at=datetime(2027, 4, 6, 20, tzinfo=UTC),
                location="Boardroom",
                module_codes=["staff-tasks"],
            ),
        )
        assert event is not None
        loadout_id = next(item.id for item in event.sub_events if item.name == "Loadout")
        private_id = next(item.id for item in event.sub_events if item.name == "Private session")
        event = add_membership(
            db,
            created.id,
            EventMembershipCreate(
                email="team-lead@example.com",
                display_name="Loadout Team Lead",
                password="Team-Lead-Password!",
                membership_type="team_lead",
            ),
        )
        assert event is not None
        membership = next(
            item for item in event.memberships if item.email == "team-lead@example.com"
        )
        scoped = assign_membership_sub_events(
            db,
            created.id,
            membership.id,
            EventSubEventRegistrationWrite(
                sub_event_ids=[loadout_id],
                roles={loadout_id: "team_lead"},
            ),
            "admin@example.com",
        )
        assert scoped is not None
        task = save_event_task(
            db,
            created.id,
            EventStaffTaskWrite(
                assigned_membership_id=membership.id,
                sub_event_id=loadout_id,
                title="Validate team manifests",
            ),
            "admin@example.com",
        )
        assert task is not None
        team_lead = db.scalar(select(User).where(User.email == "team-lead@example.com"))
        assert team_lead is not None
        assert [item.id for item in my_tasks(db, team_lead)] == [task.id]
        with pytest.raises(EventStaffTaskError, match="does not have access"):
            save_event_task(
                db,
                created.id,
                EventStaffTaskWrite(
                    assigned_membership_id=membership.id,
                    sub_event_id=private_id,
                    title="Access private session",
                ),
                "admin@example.com",
            )


def test_event_product_lineup_snapshots_catalog_controls_and_reorders() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        db.add_all(
            [
                CatalogVendor(
                    vendor_code="EXPO",
                    name="Expo Vendor",
                    is_active=True,
                    source_file="test",
                ),
                CatalogVendor(
                    vendor_code="PARTNER",
                    name="Partner Vendor",
                    is_active=True,
                    source_file="test",
                ),
            ]
        )
        db.add(
            Store(
                store_number="202",
                name="Ordering Entity",
                region_code="R2",
                entity_code="ENTITY-2",
                is_ordering_enabled=True,
                is_active=True,
            )
        )
        db.flush()
        db.add(
            CatalogProduct(
                product_code="MODEL-1",
                model_number="MODEL-1",
                vendor_code="EXPO",
                name="Catalog name",
                unit_price=Decimal("100.00"),
                currency="USD",
                minimum_order_quantity=1,
                is_available=True,
                is_active=True,
                source_file="test",
            )
        )
        db.commit()
        event = create_event(db, _event(), "admin@example.com")
        event = add_sub_event(
            db,
            event.id,
            SubEventWrite(
                name="Hot Show",
                starts_at=datetime(2027, 4, 5, 14, tzinfo=UTC),
                ends_at=datetime(2027, 4, 5, 18, tzinfo=UTC),
                location="Main stage",
                module_codes=[
                    "check-in",
                    "live-display",
                    "ordering",
                    "polling",
                    "product-slides",
                    "staff-tasks",
                    "vendor-booths",
                ],
            ),
        )
        assert event is not None
        sub_event_id = event.sub_events[0].id
        event = add_membership(
            db,
            event.id,
            EventMembershipCreate(
                email="buyer@example.com",
                display_name="Entity Buyer",
                password="Entity-Buyer-Password!",
                membership_type="franchise_representative",
                entity_code="ENTITY-2",
            ),
        )
        assert event is not None
        buyer = db.scalar(select(User).where(User.email == "buyer@example.com"))
        assert buyer is not None

        roster = attendance_roster(db, sub_event_id)
        assert roster is not None and roster.registered_total == 1
        membership_id = event.memberships[0].id
        roster = update_attendance(
            db, sub_event_id, membership_id, "checked_in", "staff@example.com"
        )
        assert roster is not None and roster.onsite_total == 1
        passes = my_attendance_passes(db, buyer)
        assert passes[0].sub_events[0].status == "checked_in"
        roster = update_attendance(
            db, sub_event_id, membership_id, "checked_out", "staff@example.com"
        )
        assert roster is not None and roster.checked_out_total == 1
        passes = my_attendance_passes(db, buyer)
        assert passes[0].sub_events[0].status == "checked_out"

        poll = create_poll(
            db,
            sub_event_id,
            EventPollCreate(
                question="Which finish should we feature?",
                options=["Natural oak", "Walnut"],
                show_results=True,
            ),
            "admin@example.com",
        )
        assert poll is not None
        poll = set_poll_status(db, poll.id, "open")
        assert poll is not None and poll.status == "open"
        live_poll = active_poll(db, sub_event_id, buyer)
        assert live_poll is not None
        voted = cast_vote(db, poll.id, poll.options[0].id, buyer)
        assert voted is not None
        assert voted.total_votes == 1
        assert voted.selected_option_id == poll.options[0].id
        with pytest.raises(EventPollError, match="already voted"):
            cast_vote(db, poll.id, poll.options[1].id, buyer)

        event = add_membership(
            db,
            event.id,
            EventMembershipCreate(
                email="second-buyer@example.com",
                display_name="Second Entity Buyer",
                password="Second-Entity-Buyer-Password!",
                membership_type="franchise_representative",
                entity_code="ENTITY-2",
            ),
        )
        assert event is not None
        second_buyer = db.scalar(select(User).where(User.email == "second-buyer@example.com"))
        assert second_buyer is not None

        first = create_slide(
            db,
            sub_event_id,
            EventProductSlideWrite(
                catalog_product_code="MODEL-1",
                model_number="MODEL-1",
                name="Event snapshot name",
                vendor_code="EXPO",
                category="Appliances",
                event_unit_cost=Decimal("89.00"),
                standard_cost=Decimal("100.00"),
                available_inventory=500,
                max_event_units=250,
                allow_waitlist=True,
                delivery_window_start=date(2027, 5, 1),
                delivery_window_end=date(2027, 5, 31),
            ),
            "admin@example.com",
        )
        second = create_slide(
            db,
            sub_event_id,
            EventProductSlideWrite(
                model_number="SPECIAL-2",
                name="Manual special",
                vendor_code="EXPO",
                event_unit_cost=Decimal("50.00"),
                product_variants=[
                    {
                        "model_number": "SPECIAL-TWIN",
                        "name": "Twin",
                        "event_unit_cost": "50.00",
                        "standard_cost": "65.00",
                        "minimum_order_quantity": 1,
                    },
                    {
                        "model_number": "SPECIAL-KING",
                        "name": "King",
                        "event_unit_cost": "75.00",
                        "minimum_order_quantity": 1,
                    },
                ],
                delivery_window_start=date(2027, 6, 1),
                delivery_window_end=date(2027, 6, 30),
            ),
            "admin@example.com",
        )
        assert first is not None and second is not None
        first = save_slide_image(
            db,
            first.id,
            "product.png",
            "image/png",
            b"\x89PNG\r\n\x1a\nproduct-image",
            "admin@example.com",
        )
        first = save_slide_vendor_logo(
            db,
            first.id,
            "vendor-logo.png",
            "image/png",
            b"\x89PNG\r\n\x1a\nvendor-logo",
            "admin@example.com",
        )
        assert first is not None
        assert first.has_image is True
        assert first.has_vendor_logo is True
        assert first.name == "Event snapshot name"
        assert first.max_event_units == 250
        assert _order_capacity(first) == 250
        first.available_inventory = 100
        assert _order_capacity(first) == 100
        first.available_inventory = 500
        assert first.standard_cost == Decimal("100.00")
        assert first.category == "Appliances"
        assert second.product_variants[0].standard_cost == Decimal("65.00")
        reordered = reorder_slides(db, sub_event_id, [second.id, first.id])
        assert reordered is not None
        assert [slide.id for slide in reordered] == [second.id, first.id]

        with pytest.raises(EventProductSlideError, match="every slide"):
            reorder_slides(db, sub_event_id, [first.id])

        presentation = control_presentation(db, sub_event_id, "start", "presenter@example.com")
        assert presentation is not None
        assert presentation.current_slide is not None
        assert presentation.current_slide.id == second.id
        assert presentation.projector_image_preload_ids == [first.id]
        presentation = control_presentation(db, sub_event_id, "open", "presenter@example.com")
        assert presentation is not None
        assert presentation.ordering_status == "open"
        workspace = submit_entity_order(
            db,
            sub_event_id,
            EventEntityOrderWrite(
                quantity=5,
                variant_quantities={"SPECIAL-TWIN": 3, "SPECIAL-KING": 2},
            ),
            buyer,
        )
        assert workspace is not None
        assert workspace.existing_order is not None
        assert workspace.existing_order.total_cost == Decimal("300.00")
        assert workspace.existing_order.variant_quantities["SPECIAL-KING"] == 2
        assert workspace.existing_order.requested_delivery_start == date(2027, 6, 1)
        assert workspace.existing_order.requested_delivery_end == date(2027, 6, 30)
        assert "entity_event_spend" not in workspace.model_dump()
        shared_workspace = ordering_workspace(db, sub_event_id, second_buyer)
        assert shared_workspace is not None
        assert shared_workspace.existing_order is not None
        assert shared_workspace.existing_order.id == workspace.existing_order.id
        shared_workspace = submit_entity_order(
            db,
            sub_event_id,
            EventEntityOrderWrite(
                quantity=5,
                variant_quantities={"SPECIAL-TWIN": 3, "SPECIAL-KING": 2},
            ),
            second_buyer,
        )
        assert shared_workspace is not None
        assert shared_workspace.existing_order is not None
        assert shared_workspace.existing_order.id == workspace.existing_order.id
        assert (
            db.scalar(
                select(func.count(EventEntityOrder.id)).where(
                    EventEntityOrder.sub_event_id == sub_event_id,
                    EventEntityOrder.entity_code == "ENTITY-2",
                )
            )
            == 1
        )
        revisions = db.scalars(
            select(EventEntityOrderRevision)
            .where(EventEntityOrderRevision.order_id == workspace.existing_order.id)
            .order_by(EventEntityOrderRevision.revision)
        ).all()
        assert [item.changed_by for item in revisions] == [
            "buyer@example.com",
            "second-buyer@example.com",
        ]
        live_presentation = get_presentation(db, sub_event_id)
        assert live_presentation is not None
        assert live_presentation.variant_units_ordered == {
            "SPECIAL-TWIN": 3,
            "SPECIAL-KING": 2,
        }
        assert live_presentation.presenter_slides == []
        filler = create_slide(
            db,
            sub_event_id,
            EventProductSlideWrite(
                slide_type="filler",
                filler_category="raffle",
                name="Prize drawing",
                description="Keep your badge ready for the raffle.",
                presenter_notes="Confirm the winning badge number before advancing.",
                status="ready",
            ),
            "admin@example.com",
        )
        assert filler is not None
        assert filler.slide_type == "filler"
        assert filler.vendor_code is None
        assert filler.event_unit_cost is None
        reordered = reorder_slides(db, sub_event_id, [second.id, filler.id, first.id])
        assert reordered is not None
        presentation = control_presentation(db, sub_event_id, "next", "presenter@example.com")
        assert presentation is not None
        assert presentation.current_slide is not None
        assert presentation.current_slide.id == filler.id
        assert presentation.ordering_status == "closed"
        public_presentation = get_presentation(db, sub_event_id)
        assert public_presentation is not None
        assert public_presentation.current_slide is not None
        assert public_presentation.current_slide.presenter_notes is None
        assert public_presentation.presenter_notes is None
        assert public_presentation.slide_queue == []
        presenter_presentation = get_presentation(db, sub_event_id, include_presenter_details=True)
        assert presenter_presentation is not None
        assert presenter_presentation.presenter_notes == (
            "Confirm the winning badge number before advancing."
        )
        assert [item.id for item in presenter_presentation.slide_queue] == [
            second.id,
            filler.id,
            first.id,
        ]
        assert [item.id for item in presenter_presentation.presenter_slides] == [
            second.id,
            filler.id,
            first.id,
        ]
        attendee_workspace = ordering_workspace(db, sub_event_id, buyer)
        assert attendee_workspace is not None
        assert attendee_workspace.current_slide is not None
        assert attendee_workspace.current_slide.presenter_notes is None
        with pytest.raises(EventPresentationError, match="filler slide"):
            control_presentation(db, sub_event_id, "open", "presenter@example.com")
        presentation = control_presentation(db, sub_event_id, "previous", "presenter@example.com")
        assert presentation is not None and presentation.current_slide.id == second.id
        partner_slide = create_slide(
            db,
            sub_event_id,
            EventProductSlideWrite(
                model_number="PARTNER-1",
                name="Partner model",
                vendor_code="PARTNER",
                event_unit_cost=Decimal("125.00"),
                delivery_window_start=date(2027, 6, 1),
                delivery_window_end=date(2027, 6, 30),
            ),
            "admin@example.com",
        )
        assert partner_slide is not None
        reordered = reorder_slides(
            db, sub_event_id, [second.id, partner_slide.id, filler.id, first.id]
        )
        assert reordered is not None
        expanded_event = add_sub_event(
            db,
            event.id,
            SubEventWrite(
                name="Separate buying session",
                starts_at=datetime(2027, 4, 6, 14, tzinfo=UTC),
                ends_at=datetime(2027, 4, 6, 18, tzinfo=UTC),
                location="Second stage",
                module_codes=["product-slides"],
            ),
        )
        assert expanded_event is not None
        other_sub_event_id = next(
            item.id for item in expanded_event.sub_events if item.id != sub_event_id
        )
        other_slide = create_slide(
            db,
            other_sub_event_id,
            EventProductSlideWrite(
                model_number="OTHER-SESSION",
                name="Other session product",
                vendor_code="EXPO",
                event_unit_cost=Decimal("700.00"),
                delivery_window_start=date(2027, 6, 1),
                delivery_window_end=date(2027, 6, 30),
            ),
            "admin@example.com",
        )
        assert other_slide is not None
        other_order = EventEntityOrder(
            event_id=event.id,
            sub_event_id=other_sub_event_id,
            slide_id=other_slide.id,
            membership_id=membership_id,
            user_id=buyer.id,
            entity_code="ENTITY-2",
            quantity=1,
            requested_delivery_start=date(2027, 6, 1),
            requested_delivery_end=date(2027, 6, 15),
            unit_cost=Decimal("700.00"),
            total_cost=Decimal("700.00"),
            status="confirmed",
        )
        db.add(other_order)
        db.commit()
        franchise_insights = live_insights(db, sub_event_id, buyer)
        assert franchise_insights is not None
        assert franchise_insights.scope == "franchise"
        assert franchise_insights.entity_code == "ENTITY-2"
        assert franchise_insights.franchise_sub_event_units == 5
        assert franchise_insights.franchise_sub_event_spend == "300.00"
        assert "event_spend" not in franchise_insights.model_dump()
        assert "event_units" not in franchise_insights.model_dump()
        presentation = control_presentation(db, sub_event_id, "open", "presenter@example.com")
        assert presentation is not None
        assert presentation.total_units_ordered == 5
        assert presentation.sub_event_combined_spend == "300.00"
        assert "event_combined_spend" not in presentation.model_dump()
        assert "event_units_ordered" not in presentation.model_dump()
        db.delete(other_order)
        db.commit()
        add_membership(
            db,
            event.id,
            EventMembershipCreate(
                email="vendor-live@example.com",
                display_name="Vendor Live Viewer",
                password="Vendor-Live-Password!",
                membership_type="vendor",
                vendor_code="EXPO",
            ),
        )
        vendor_viewer = db.scalar(select(User).where(User.email == "vendor-live@example.com"))
        assert vendor_viewer is not None
        vendor_membership = db.scalar(
            select(EventMembership).where(
                EventMembership.event_id == event.id,
                EventMembership.user_id == vendor_viewer.id,
            )
        )
        assert vendor_membership is not None
        vendor_membership.vendor_codes = ["EXPO", "PARTNER"]
        db.commit()
        vendor_insights = live_insights(db, sub_event_id, vendor_viewer)
        assert vendor_insights is not None
        assert vendor_insights.scope == "vendor"
        assert vendor_insights.vendor_sub_event_spend == "300.00"
        assert vendor_insights.vendor_code is None
        assert vendor_insights.next_vendor_code == "PARTNER"
        assert vendor_insights.next_vendor_name == "Partner Vendor"
        assert vendor_insights.slides_until_next_product == 1
        vendor_totals = {item.vendor_code: item for item in vendor_insights.vendor_totals}
        assert vendor_totals["EXPO"].committed_spend == "300.00"
        assert vendor_totals["PARTNER"].committed_spend == "0.00"
        variant_metrics = {item.model_number: item for item in vendor_insights.vendor_products}
        assert variant_metrics["SPECIAL-TWIN"].committed_spend == "150.00"
        assert variant_metrics["SPECIAL-KING"].committed_spend == "150.00"
        assert variant_metrics["PARTNER-1"].vendor_name == "Partner Vendor"
        analytics = get_live_analytics(db, sub_event_id)
        assert analytics is not None
        assert analytics.assigned_entities == 1
        assert analytics.responding_entities == 1
        assert analytics.confirmed_spend == "300.00"
        state = db.get(EventPresentationState, sub_event_id)
        assert state is not None and state.ordering_status == "open"
        presentation = control_presentation(db, sub_event_id, "close", "presenter@example.com")
        assert presentation is not None and presentation.ordering_status == "closed"
        with pytest.raises(EventOrderingError, match="Ordering is not open"):
            submit_entity_order(
                db,
                sub_event_id,
                EventEntityOrderWrite(
                    quantity=1,
                    variant_quantities={"SPECIAL-TWIN": 1},
                ),
                buyer,
            )
        order_id = workspace.existing_order.id
        reviewed_event_id = decide_order(
            db,
            order_id,
            EventOrderReviewDecision(decision="revise", revised_quantity=8, reason="Budget"),
            "purchasing@example.com",
        )
        assert reviewed_event_id == event.id
        review = review_summary(db, event.id)
        assert review is not None
        assert review.approved == 1
        assert review.approved_units == 8
        batch = release_approved_orders(db, event.id, "purchasing@example.com")
        assert batch is not None
        assert batch.order_count == 1
        assert batch.total_spend == Decimal("475.00")
        released_lines = list(
            db.scalars(
                select(EventOrderReleaseLine).where(
                    EventOrderReleaseLine.batch_id == batch.batch_id
                )
            ).all()
        )
        assert {line.model_number for line in released_lines} == {
            "SPECIAL-TWIN",
            "SPECIAL-KING",
        }
        assert sum(line.quantity for line in released_lines) == 8
        assert sum((line.total_cost for line in released_lines), Decimal("0")) == Decimal("475.00")
        purchase_requests = list(
            db.scalars(
                select(PurchaseRequest).where(
                    PurchaseRequest.context["release_batch_id"].as_string() == batch.batch_id
                )
            ).all()
        )
        assert len(purchase_requests) == 1
        released_request = purchase_requests[0]
        assert released_request.status == "submitted_to_purchasing"
        assert released_request.order_number.startswith(f"{event.name}-202-EXPO-")
        assert released_request.context["source"] == "event_live_order_release"
        assert {line.product_name for line in released_request.line_items} == {
            "Twin",
            "King",
        }
        assert released_request.total == Decimal("475.00")
        assert {line.purchase_request_id for line in released_lines} == {released_request.id}
        released_review = review_summary(db, event.id)
        assert released_review is not None
        released_item = next(item for item in released_review.items if item.order_id == order_id)
        assert len(released_item.purchasing_requests) == 1
        assert released_item.purchasing_requests[0].purchase_request_id == released_request.id
        assert released_item.purchasing_requests[0].order_number == released_request.order_number
        assert released_item.purchasing_requests[0].status == "submitted_to_purchasing"
        purchase_order = PurchaseOrder(
            po_number="202-202706-000001",
            workflow_code="VENDOR_ORDER",
            vendor_code="EXPO",
            status="created",
            currency="USD",
            subtotal=Decimal("475.00"),
            freight_total=Decimal("0"),
            tax_total=Decimal("0"),
            total=Decimal("475.00"),
            expected_delivery_date=date(2027, 6, 1),
            created_by="purchasing@example.com",
        )
        purchase_order.sources.append(
            PurchaseOrderSource(
                purchase_request_id=released_request.id,
                store_number="202",
            )
        )
        for line in released_request.line_items:
            purchase_order.lines.append(
                PurchaseOrderLine(
                    source_request_id=released_request.id,
                    source_line_id=line.id,
                    store_number="202",
                    product_code=line.product_code,
                    product_name=line.product_name,
                    quantity=line.quantity,
                    received_quantity=0,
                    unit_price=line.unit_price,
                    freight_amount=0,
                    tax_amount=0,
                    extended_amount=line.extended_amount,
                )
            )
        db.add(purchase_order)
        db.commit()
        backup = export_event_order_backup(db, event.id)
        assert backup is not None
        backup_event, backup_content = backup
        assert backup_event.id == event.id
        workbook = load_workbook(BytesIO(backup_content), data_only=True)
        assert workbook.sheetnames[:3] == [
            "Event Summary",
            "All Order Lines",
            "Entity ENTITY-2",
        ]
        assert workbook.sheetnames[-2:] == ["Purchasing Handoff", "Purchase Orders"]
        order_rows = list(workbook["All Order Lines"].iter_rows(min_row=2, values_only=True))
        assert {row[8] for row in order_rows} == {"SPECIAL-TWIN", "SPECIAL-KING"}
        assert sum(row[10] for row in order_rows) == 8
        assert sum((row[12] for row in order_rows), Decimal("0")) == Decimal("475")
        entity_rows = list(workbook["Entity ENTITY-2"].iter_rows(min_row=2, values_only=True))
        assert entity_rows == order_rows
        assert {row[5] for row in entity_rows} == {"R2"}
        assert {row[6] for row in entity_rows} == {"202"}
        handoff_rows = list(workbook["Purchasing Handoff"].iter_rows(min_row=2, values_only=True))
        assert handoff_rows[0][0] == released_request.id
        assert handoff_rows[0][2] == "event_live_order_release"
        po_rows = list(workbook["Purchase Orders"].iter_rows(min_row=2, values_only=True))
        assert {row[7] for row in po_rows} == {"SPECIAL-TWIN", "SPECIAL-KING"}
        assert {row[1] for row in po_rows} == {purchase_order.po_number}
        assert sum(row[9] for row in po_rows) == 8
        assert sum((row[12] for row in po_rows), Decimal("0")) == Decimal("475")
        presentation = control_presentation(db, sub_event_id, "next", "presenter@example.com")
        assert presentation is not None
        assert presentation.current_slide is not None
        assert presentation.current_slide.id == partner_slide.id
        presentation = control_presentation(db, sub_event_id, "next", "presenter@example.com")
        assert presentation is not None
        assert presentation.current_slide is not None
        assert presentation.current_slide.id == filler.id
        presentation = control_presentation(db, sub_event_id, "next", "presenter@example.com")
        assert presentation is not None
        assert presentation.current_slide is not None
        assert presentation.current_slide.id == first.id
        assert presentation.ordering_status == "closed"
        presentation = control_presentation(db, sub_event_id, "end", "presenter@example.com")
        assert presentation is not None and presentation.status == "ended"
        with pytest.raises(EventPresentationError, match="Start the presentation"):
            control_presentation(db, sub_event_id, "next", "presenter@example.com")
        db.add(
            EventSettlementEvent(
                event_id=event.id,
                status="closed",
                created_by="purchasing@example.com",
            )
        )
        db.commit()
        with pytest.raises(EventPresentationError, match="settlement is closed"):
            control_presentation(db, sub_event_id, "start", "presenter@example.com")
        with pytest.raises(EventOrderingError, match="settlement is closed"):
            submit_entity_order(
                db,
                sub_event_id,
                EventEntityOrderWrite(
                    quantity=1,
                ),
                buyer,
            )
        with pytest.raises(EventPollError, match="settlement is closed"):
            create_poll(
                db,
                sub_event_id,
                EventPollCreate(
                    question="Late poll?",
                    options=["Yes", "No"],
                    show_results=True,
                ),
                "admin@example.com",
            )
        with pytest.raises(EventAttendanceError, match="settlement is closed"):
            update_attendance(
                db,
                sub_event_id,
                membership_id,
                "checked_in",
                "staff@example.com",
            )
        slide_ids = [item.id for item in list_slides(db, sub_event_id) or []]
        with pytest.raises(EventProductSlideError, match="settlement is closed"):
            reorder_slides(db, sub_event_id, slide_ids)
        with pytest.raises(EventOrderReviewError, match="settlement is closed"):
            decide_order(
                db,
                order_id,
                EventOrderReviewDecision(decision="reject", reason="Late change"),
                "purchasing@example.com",
            )
        with pytest.raises(EventOrderReviewError, match="settlement is closed"):
            release_approved_orders(db, event.id, "purchasing@example.com")
