from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session

from app.db.session import Base
from app.models.catalog import (
    CatalogProduct,
    CatalogProductCostHistory,
    CatalogVendor,
    ModelCategory,
    VendorMOQRule,
)
from app.models.event_snapshot import EventSnapshot  # noqa: F401
from app.models.store import Store  # noqa: F401
from app.models.workflow import WorkflowInstance  # noqa: F401
from app.schemas.catalog import VendorModelUpdate
from app.services.vendor_model_service import (
    MODEL_COLUMNS,
    VendorModelError,
    export_vendor_models,
    import_vendor_models,
    list_vendor_models,
    update_vendor_model,
)


def _database() -> Session:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    return Session(engine)


def _seed(db: Session) -> None:
    db.add_all(
        [
            CatalogVendor(
                vendor_code="V-ONE",
                name="Vendor One",
                is_active=True,
                source_file="seed.xlsx",
            ),
            CatalogVendor(
                vendor_code="V-TWO",
                name="Vendor Two",
                is_active=True,
                source_file="seed.xlsx",
            ),
        ]
    )
    db.flush()
    db.add(
        ModelCategory(
            department="APPL MISC",
            product_category_code="MISC",
            status="VALID",
        )
    )
    standard = VendorMOQRule(
        vendor_code="V-ONE",
        code="STANDARD",
        name="Standard MOQ",
        threshold_type="unit_quantity",
        threshold_value=1,
        is_active=True,
    )
    db.add(standard)
    db.flush()
    db.add_all(
        [
            CatalogProduct(
                product_code="ONE-1",
                model_number="ONE-1",
                vendor_code="V-ONE",
                name="Original model",
                unit_price=Decimal("10.00"),
                currency="USD",
                minimum_order_quantity=1,
                is_available=True,
                is_active=True,
                source_file="seed.xlsx",
                moq_rule_id=standard.id,
            ),
            CatalogProduct(
                product_code="TWO-1",
                model_number="TWO-1",
                vendor_code="V-TWO",
                name="Other vendor model",
                unit_price=Decimal("20.00"),
                currency="USD",
                minimum_order_quantity=1,
                is_available=True,
                is_active=True,
                source_file="seed.xlsx",
            ),
        ]
    )
    db.commit()


def test_vendor_model_list_and_edit_are_identity_scoped() -> None:
    with _database() as db:
        _seed(db)

        assert [item.product_code for item in list_vendor_models(db, "V-ONE")] == ["ONE-1"]
        assert (
            update_vendor_model(
                db,
                "V-ONE",
                "TWO-1",
                VendorModelUpdate(name="Not allowed"),
                "vendor@example.com",
            )
            is None
        )


def test_vendor_model_list_filters_clumps_components_and_single_items() -> None:
    with _database() as db:
        _seed(db)
        first = db.scalar(select(CatalogProduct).where(CatalogProduct.product_code == "ONE-1"))
        assert first is not None
        first.is_clump = True
        db.add(
            CatalogProduct(
                product_code="COMPONENT-1",
                model_number="COMPONENT-1",
                vendor_code="V-ONE",
                name="Component",
                unit_price=Decimal("5.00"),
                currency="USD",
                minimum_order_quantity=1,
                is_available=True,
                is_active=True,
                is_clump=False,
                part_of_clump=True,
                source_file="seed.xlsx",
            )
        )
        db.commit()

        clumps = list_vendor_models(db, "V-ONE", classification="clump")
        components = list_vendor_models(db, "V-ONE", classification="part_of_clump")
        assert [item.product_code for item in clumps] == ["ONE-1"]
        assert [item.product_code for item in components] == ["COMPONENT-1"]
        assert list_vendor_models(db, "V-ONE", classification="single_item") == []


def test_vendor_model_list_orders_clumps_singles_then_components() -> None:
    with _database() as db:
        _seed(db)
        first = db.scalar(select(CatalogProduct).where(CatalogProduct.product_code == "ONE-1"))
        assert first is not None
        first.is_clump = True
        db.add_all(
            [
                CatalogProduct(
                    product_code="SINGLE-1",
                    model_number="SINGLE-1",
                    vendor_code="V-ONE",
                    name="Single",
                    unit_price=Decimal("5.00"),
                    currency="USD",
                    minimum_order_quantity=1,
                    source_file="seed.xlsx",
                ),
                CatalogProduct(
                    product_code="PART-1",
                    model_number="PART-1",
                    vendor_code="V-ONE",
                    name="Part",
                    unit_price=Decimal("5.00"),
                    currency="USD",
                    minimum_order_quantity=1,
                    part_of_clump=True,
                    source_file="seed.xlsx",
                ),
            ]
        )
        db.commit()
        ordered = list_vendor_models(db, "V-ONE")
        assert [item.product_code for item in ordered] == ["ONE-1", "SINGLE-1", "PART-1"]


def test_cost_edit_creates_history_without_archiving_other_fields() -> None:
    with _database() as db:
        _seed(db)
        updated = update_vendor_model(
            db,
            "V-ONE",
            "ONE-1",
            VendorModelUpdate(name="Renamed model", unit_price=Decimal("12.50")),
            "vendor@example.com",
        )

        assert updated is not None
        assert updated.name == "Renamed model"
        history = list(
            db.scalars(
                select(CatalogProductCostHistory)
                .where(CatalogProductCostHistory.product_code == "ONE-1")
                .order_by(CatalogProductCostHistory.effective_from)
            ).all()
        )
        assert [item.unit_price for item in history] == [Decimal("10.00"), Decimal("12.50")]
        assert history[0].effective_to is not None
        assert history[1].effective_to is None

        update_vendor_model(
            db,
            "V-ONE",
            "ONE-1",
            VendorModelUpdate(brand="Updated brand"),
            "vendor@example.com",
        )
        assert db.scalar(select(func.count(CatalogProductCostHistory.id))) == 2


def test_owning_vendor_can_replace_model_number_without_changing_internal_key() -> None:
    with _database() as db:
        _seed(db)
        update_vendor_model(
            db,
            "V-ONE",
            "ONE-1",
            VendorModelUpdate(unit_price=Decimal("12.50")),
            "vendor@example.com",
        )

        updated = update_vendor_model(
            db,
            "V-ONE",
            "ONE-1",
            VendorModelUpdate(model_number="ONE-REPLACED"),
            "vendor@example.com",
        )

        assert updated is not None
        assert updated.product_code == "ONE-1"
        assert updated.model_number == "ONE-REPLACED"
        assert set(db.scalars(select(CatalogProductCostHistory.product_code))) == {"ONE-1"}


def test_different_vendors_can_use_the_same_model_number() -> None:
    with _database() as db:
        _seed(db)

        updated = update_vendor_model(
            db,
            "V-ONE",
            "ONE-1",
            VendorModelUpdate(model_number="TWO-1"),
            "vendor@example.com",
        )

        assert updated is not None
        matches = list(
            db.scalars(
                select(CatalogProduct)
                .where(CatalogProduct.model_number == "TWO-1")
                .order_by(CatalogProduct.vendor_code)
            ).all()
        )
        assert [(item.vendor_code, item.product_code) for item in matches] == [
            ("V-ONE", "ONE-1"),
            ("V-TWO", "TWO-1"),
        ]


def test_vendor_cannot_duplicate_a_model_number_inside_its_own_catalog() -> None:
    with _database() as db:
        _seed(db)
        db.add(
            CatalogProduct(
                product_code="ONE-2",
                model_number="ONE-2",
                vendor_code="V-ONE",
                name="Second model",
                unit_price=Decimal("15.00"),
                currency="USD",
                minimum_order_quantity=1,
                is_available=True,
                is_active=True,
                source_file="seed.xlsx",
            )
        )
        db.commit()

        try:
            update_vendor_model(
                db,
                "V-ONE",
                "ONE-1",
                VendorModelUpdate(model_number="ONE-2"),
                "vendor@example.com",
            )
        except VendorModelError as exc:
            assert str(exc) == "Model number already exists"
        else:
            raise AssertionError("Expected same-vendor duplicate model number to be rejected")


def test_excel_export_and_import_only_adds_or_changes_rows() -> None:
    with _database() as db:
        _seed(db)
        exported = export_vendor_models(db, "V-ONE")
        workbook = load_workbook(BytesIO(exported), data_only=True)
        assert workbook["Products"].max_row == 2
        assert workbook["Products"]["A2"].value == "ONE-1"
        assert workbook["Approved MOQ Codes"].sheet_state == "veryHidden"
        assert workbook["Approved MOQ Codes"]["A1"].value == "STANDARD"
        validations = workbook["Products"].data_validations.dataValidation
        assert len(validations) == 1
        assert validations[0].formula1 == "=ApprovedMOQCodes"
        assert "J2:J50001" in str(validations[0].sqref)

        incoming = Workbook()
        sheet = incoming.active
        sheet.title = "Products"
        sheet.append(MODEL_COLUMNS)
        sheet.append(
            [
                "ONE-1",
                "V-ONE",
                "Batch renamed",
                "ONE-1",
                "APPL MISC",
                "MISC",
                None,
                11.25,
                "USD",
                "STANDARD",
                True,
                True,
            ]
        )
        sheet.append(
            [
                "ONE-2",
                "V-ONE",
                "New model",
                "M-2",
                "APPL MISC",
                "MISC",
                "Brand",
                30,
                "USD",
                "STANDARD",
                True,
                True,
            ]
        )
        content = BytesIO()
        incoming.save(content)

        result = import_vendor_models(
            db,
            "V-ONE",
            "models.xlsx",
            content.getvalue(),
            "vendor@example.com",
        )

        assert result == (1, 1, 0, 2)
        assert db.get(CatalogProduct, 1).name == "Batch renamed"
        assert (
            db.scalar(select(CatalogProduct).where(CatalogProduct.product_code == "TWO-1"))
            is not None
        )
        created = db.scalar(select(CatalogProduct).where(CatalogProduct.product_code == "M-2"))
        assert created is not None
        assert created.model_number == "M-2"
        assert db.scalar(select(func.count(CatalogProductCostHistory.id))) == 3


def test_vendor_model_import_matches_moq_codes_case_insensitively() -> None:
    with _database() as db:
        _seed(db)
        standard = db.scalar(
            select(VendorMOQRule).where(
                VendorMOQRule.vendor_code == "V-ONE",
                VendorMOQRule.code == "STANDARD",
            )
        )
        assert standard is not None
        standard.code = "Standard"
        db.add(
            VendorMOQRule(
                vendor_code="V-ONE",
                code="Special",
                name="Special MOQ",
                threshold_type="unit_quantity",
                threshold_value=5,
                is_active=True,
            )
        )
        db.commit()

        incoming = Workbook()
        sheet = incoming.active
        sheet.title = "Products"
        sheet.append(MODEL_COLUMNS)
        sheet.append(
            [
                "ignored-external-code",
                "V-ONE",
                "New standard model",
                "CASE-100",
                "APPL MISC",
                "MISC",
                None,
                25,
                "USD",
                "STANDARD",
                True,
                True,
            ]
        )
        content = BytesIO()
        incoming.save(content)

        result = import_vendor_models(
            db,
            "V-ONE",
            "models.xlsx",
            content.getvalue(),
            "vendor@example.com",
        )

        assert result == (1, 0, 0, 1)
        created = db.scalar(select(CatalogProduct).where(CatalogProduct.model_number == "CASE-100"))
        assert created is not None
        assert created.moq_rule_id == standard.id


def test_vendor_import_allows_a_model_number_owned_by_another_vendor() -> None:
    with _database() as db:
        _seed(db)
        incoming = Workbook()
        sheet = incoming.active
        sheet.title = "Products"
        sheet.append(MODEL_COLUMNS)
        sheet.append(
            [
                "ignored-external-code",
                "V-ONE",
                "Vendor One version",
                "TWO-1",
                "APPL MISC",
                "MISC",
                None,
                19.5,
                "USD",
                "STANDARD",
                True,
                True,
            ]
        )
        content = BytesIO()
        incoming.save(content)

        result = import_vendor_models(
            db,
            "V-ONE",
            "models.xlsx",
            content.getvalue(),
            "vendor@example.com",
        )

        assert result == (1, 0, 0, 1)
        matches = list(
            db.scalars(
                select(CatalogProduct)
                .where(CatalogProduct.model_number == "TWO-1")
                .order_by(CatalogProduct.vendor_code)
            ).all()
        )
        assert [item.vendor_code for item in matches] == ["V-ONE", "V-TWO"]
        assert len({item.product_code for item in matches}) == 2


def test_vendor_model_import_accepts_vendor_product_export_layout() -> None:
    with _database() as db:
        _seed(db)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Vendor Product Export_v2"
        sheet.append(
            [
                "Vendor",
                "Brand",
                "Model_Number",
                "Department",
                "ProductCode",
                "Active",
                "Is_a_Clump",
                "Part_of_a_Clump",
                "Effective_Start_Date",
                "Landed_Cost",
                "In_Stock",
                "Cost_Status",
            ]
        )
        sheet.append(
            [
                "Vendor One Group",
                "V1 Brand",
                "EXPORT-100",
                "APPL MISC",
                "MISC",
                "Yes",
                "Yes",
                "No",
                datetime(2026, 7, 1),
                125.75,
                False,
                "Pending",
            ]
        )
        content = BytesIO()
        workbook.save(content)

        result = import_vendor_models(
            db,
            "V-ONE",
            "Vendor Product Export.xlsx",
            content.getvalue(),
            "vendor@example.com",
        )

        assert result == (1, 0, 0, 1)
        product = db.scalar(
            select(CatalogProduct).where(CatalogProduct.product_code == "EXPORT-100")
        )
        assert product is not None
        assert product.name == "EXPORT-100"
        assert product.brand == "V1 Brand"
        assert product.is_clump is True
        assert product.part_of_clump is False
        assert product.cost_effective_start_date == date(2026, 7, 1)
        assert product.unit_price == Decimal("125.75")
        assert product.is_available is False
        assert product.cost_status == "Pending"
