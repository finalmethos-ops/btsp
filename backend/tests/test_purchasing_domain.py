import hashlib
import json
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session

from app.db.session import Base
from app.models.catalog import CatalogProduct, CatalogVendor
from app.models.configuration import ConfigurationEntry  # noqa: F401
from app.models.event_snapshot import EventSnapshot
from app.models.identity import Role, User
from app.models.notification import NotificationEvent, NotificationTemplate  # noqa: F401
from app.models.purchase_order import (
    PurchaseOrder,
    PurchaseOrderArtifact,
    PurchaseOrderSource,
    PurchaseOrderTransmissionEvent,
)
from app.models.purchasing import PurchaseRequest, PurchaseRequestLineItem
from app.models.store import Store
from app.models.workflow import WorkflowDefinition  # noqa: F401
from app.schemas.attachment import AttachmentCategory
from app.schemas.configuration_entry import ConfigEntryWrite
from app.schemas.flow import FlowActionRequest
from app.schemas.purchase_order_artifact import PurchaseOrderArtifactFormat
from app.schemas.purchase_order_transmission import (
    PurchaseOrderTransmissionAction,
    PurchaseOrderTransmissionChannel,
)
from app.schemas.purchasing import PurchaseLineWrite, PurchaseRequestCreate, PurchaseRequestUpdate
from app.services.attachment_service import (
    AttachmentError,
    attachment_path,
    delete_attachment,
    list_attachments,
    store_attachment,
)
from app.services.bpp_purchasing_seed_service import seed_bpp_purchasing
from app.services.catalog_import_service import CatalogImportError, import_catalog
from app.services.configuration_service import upsert_config_entry
from app.services.independent_seed_service import seed_independent_purchasing
from app.services.purchase_order_artifact_service import (
    artifact_path as po_artifact_path,
)
from app.services.purchase_order_artifact_service import (
    generate_artifact,
)
from app.services.purchase_order_service import (
    PurchaseOrderError,
    _partition_lines,
    generate_purchase_orders,
    seed_purchase_order_defaults,
)
from app.services.purchase_order_transmission_service import (
    PurchaseOrderTransmissionError,
    apply_transmission_action,
    create_transmission,
)
from app.services.purchase_request_service import (
    PurchaseRequestError,
    add_line_item,
    clone_purchase_request,
    create_purchase_request,
    expire_stale_drafts,
    submit_purchase_request,
    update_purchase_request,
    validate_purchase_request,
)
from app.services.purchasing_rule_service import seed_purchasing_defaults
from app.services.workflow_engine import advance_workflow


@pytest.fixture
def db() -> Session:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        session.add(Role(code="SYSTEM_ADMIN", name="System Administrator", is_system_role=True))
        session.add(
            Store(
                store_number="1001",
                name="Test Store",
                region_code="EAST",
                is_active=True,
                is_ordering_enabled=True,
            )
        )
        session.commit()
        yield session


def workbook_bytes(price: int = 125) -> bytes:
    workbook = Workbook()
    vendors = workbook.active
    vendors.title = "Vendors"
    vendors.append(["vendor_code", "name", "is_active"])
    vendors.append(["V-100", "Internal Test Vendor", True])
    products = workbook.create_sheet("Products")
    products.append(
        [
            "product_code",
            "vendor_code",
            "name",
            "model_number",
            "category",
            "brand",
            "unit_price",
            "currency",
            "minimum_order_quantity",
            "is_available",
            "is_active",
        ]
    )
    products.append(
        [
            "P-100",
            "V-100",
            "Test Product",
            "MODEL-1",
            "Fixtures",
            "BTSP",
            price,
            "USD",
            2,
            True,
            True,
        ]
    )
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_catalog_excel_import_is_idempotent_and_updates_current_price(db: Session) -> None:
    first = import_catalog(db, "catalog.xlsx", workbook_bytes(125), "admin@example.com")
    second = import_catalog(db, "catalog.xlsx", workbook_bytes(150), "admin@example.com")

    assert first.status == second.status == "completed"
    assert db.scalar(select(func.count()).select_from(CatalogVendor)) == 1
    assert db.scalar(select(func.count()).select_from(CatalogProduct)) == 1
    assert db.scalar(select(CatalogProduct)).unit_price == Decimal("150.0000")


def test_catalog_import_rejects_unknown_vendor_without_partial_rows(db: Session) -> None:
    content = workbook_bytes()
    workbook = Workbook()
    workbook.remove(workbook.active)
    vendors = workbook.create_sheet("Vendors")
    vendors.append(["vendor_code", "name"])
    vendors.append(["V-100", "Vendor"])
    products = workbook.create_sheet("Products")
    products.append(["product_code", "vendor_code", "name", "unit_price"])
    products.append(["P-100", "UNKNOWN", "Product", 10])
    stream = BytesIO()
    workbook.save(stream)

    assert content
    with pytest.raises(CatalogImportError, match="unknown vendor_code"):
        import_catalog(db, "bad.xlsx", stream.getvalue(), "admin@example.com")
    assert db.scalar(select(func.count()).select_from(CatalogVendor)) == 0


def test_purchase_request_captures_price_calculates_totals_and_submits(db: Session) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(125), "admin@example.com")
    seed_bpp_purchasing(db, actor="admin@example.com")
    user = User(
        email="buyer@example.com",
        display_name="Buyer",
        password_hash="unused",
        region_code="EAST",
        is_active=True,
    )
    db.add(user)
    db.commit()
    request = create_purchase_request(
        db,
        PurchaseRequestCreate(
            workflow_code="BPP_PURCHASING",
            store_number="1001",
            vendor_code="V-100",
        ),
        user.email,
    )
    line = add_line_item(
        db,
        request,
        PurchaseLineWrite(
            product_code="P-100",
            quantity=2,
            freight_amount=Decimal("10"),
            tax_amount=Decimal("5"),
        ),
        user.email,
    )

    assert line.unit_price == Decimal("125.0000")
    assert request.subtotal == Decimal("250.0000")
    assert request.total == Decimal("265.0000")

    import_catalog(db, "catalog.xlsx", workbook_bytes(150), "admin@example.com")
    db.refresh(line)
    assert line.unit_price == Decimal("125.0000")

    submitted = submit_purchase_request(db, request, user, {"workflow.bpp.submit"})
    assert submitted.status == "submitted"
    assert submitted.workflow_instance_id is not None
    event_types = set(
        db.scalars(
            select(EventSnapshot.event_type).where(EventSnapshot.entity_id == request.id)
        ).all()
    )
    assert {
        "purchase_request.created",
        "workflow.started",
        "purchase_request.submitted",
    } <= event_types

    advance_workflow(
        db,
        submitted.workflow_instance_id,
        FlowActionRequest(action="submit_for_department_review", actor=user.email),
        {"workflow.bpp.submit"},
    )
    db.refresh(request)
    assert request.status == "department_review"


def test_line_item_requires_catalog_minimum_and_matching_vendor(db: Session) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(), "admin@example.com")
    request = create_purchase_request(
        db,
        PurchaseRequestCreate(
            workflow_code="IND_PURCHASING",
            store_number="1001",
            vendor_code="V-100",
        ),
        "buyer@example.com",
    )
    with pytest.raises(PurchaseRequestError, match="at least"):
        add_line_item(
            db,
            request,
            PurchaseLineWrite(product_code="P-100", quantity=1),
            "buyer@example.com",
        )
    assert db.scalar(select(func.count()).select_from(PurchaseRequestLineItem)) == 0


def test_configured_rules_change_readiness_without_code_changes(db: Session) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(), "admin@example.com")
    seed_bpp_purchasing(db, actor="admin@example.com")
    seed_purchasing_defaults(db, "admin@example.com")
    request = create_purchase_request(
        db,
        PurchaseRequestCreate(
            workflow_code="BPP_PURCHASING", store_number="1001", vendor_code="V-100"
        ),
        "buyer@example.com",
    )
    add_line_item(
        db,
        request,
        PurchaseLineWrite(product_code="P-100", quantity=2),
        "buyer@example.com",
    )
    assert validate_purchase_request(db, request).ready is True

    upsert_config_entry(
        db,
        ConfigEntryWrite(
            scope_type="purchasing",
            scope_key="BPP_PURCHASING",
            key="rules.blocked_categories",
            value={"categories": ["Fixtures"]},
            updated_by="admin@example.com",
        ),
    )
    result = validate_purchase_request(db, request)
    assert result.ready is False
    assert "category.blocked" in {issue.code for issue in result.errors}

    upsert_config_entry(
        db,
        ConfigEntryWrite(
            scope_type="purchasing",
            scope_key="BPP_PURCHASING",
            key="rules.blocked_categories",
            value={"categories": "not-a-list"},
            updated_by="admin@example.com",
        ),
    )
    invalid = validate_purchase_request(db, request)
    assert invalid.ready is False
    assert "configuration.invalid" in {issue.code for issue in invalid.errors}


def test_purchasing_default_seed_is_idempotent(db: Session) -> None:
    first = seed_purchasing_defaults(db, "admin@example.com")
    second = seed_purchasing_defaults(db, "admin@example.com")
    count = db.scalar(
        select(func.count())
        .select_from(ConfigurationEntry)
        .where(ConfigurationEntry.scope_type == "purchasing")
    )

    assert first == second == 24
    assert count == 24


def test_draft_revision_clone_and_expiration(db: Session) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(), "admin@example.com")
    request = create_purchase_request(
        db,
        PurchaseRequestCreate(
            workflow_code="BPP_PURCHASING", store_number="1001", vendor_code="V-100"
        ),
        "buyer@example.com",
    )
    add_line_item(
        db,
        request,
        PurchaseLineWrite(product_code="P-100", quantity=2),
        "buyer@example.com",
    )
    current_revision = request.revision
    update_purchase_request(
        db,
        request,
        PurchaseRequestUpdate(context={"autosaved": True}, expected_revision=current_revision),
        "buyer@example.com",
    )
    with pytest.raises(PurchaseRequestError, match="reload"):
        update_purchase_request(
            db,
            request,
            PurchaseRequestUpdate(context={}, expected_revision=current_revision),
            "buyer@example.com",
        )

    clone = clone_purchase_request(db, request, "buyer@example.com")
    assert clone.cloned_from_id == request.id
    assert clone.status == "draft"
    assert len(clone.line_items) == 1
    assert clone.total == request.total

    request.expires_at = datetime.now(UTC) - timedelta(minutes=1)
    db.commit()
    assert expire_stale_drafts(db, "system@example.com") == 1
    assert request.status == "expired"
    assert (
        db.scalar(
            select(func.count())
            .select_from(EventSnapshot)
            .where(
                EventSnapshot.entity_id == request.id,
                EventSnapshot.event_type == "purchase_request.expired",
            )
        )
        == 1
    )


def test_attachment_storage_checksum_listing_and_soft_delete(db: Session, tmp_path: Path) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(), "admin@example.com")
    request = create_purchase_request(
        db,
        PurchaseRequestCreate(
            workflow_code="BPP_PURCHASING", store_number="1001", vendor_code="V-100"
        ),
        "buyer@example.com",
    )
    content = b"%PDF-1.7\nBTSP quote"
    attachment = store_attachment(
        db,
        request,
        "../vendor-quote.pdf",
        "application/pdf",
        content,
        AttachmentCategory.QUOTE,
        "buyer@example.com",
        str(tmp_path),
        1024,
    )

    assert attachment.original_filename == "vendor-quote.pdf"
    assert attachment.size_bytes == len(content)
    assert attachment_path(attachment, str(tmp_path)).read_bytes() == content
    assert [item.id for item in list_attachments(db, request.id)] == [attachment.id]

    delete_attachment(db, request, attachment, "buyer@example.com", str(tmp_path))
    assert list_attachments(db, request.id) == []
    assert attachment.is_deleted is True
    event_types = set(
        db.scalars(
            select(EventSnapshot.event_type).where(EventSnapshot.entity_id == request.id)
        ).all()
    )
    assert "purchase_request.attachment_added" in event_types
    assert "purchase_request.attachment_deleted" in event_types


def test_attachment_rejects_spoofed_content(db: Session, tmp_path: Path) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(), "admin@example.com")
    request = create_purchase_request(
        db,
        PurchaseRequestCreate(
            workflow_code="BPP_PURCHASING", store_number="1001", vendor_code="V-100"
        ),
        "buyer@example.com",
    )
    with pytest.raises(AttachmentError, match="does not match"):
        store_attachment(
            db,
            request,
            "fake.pdf",
            "application/pdf",
            b"this is not a PDF",
            AttachmentCategory.PDF,
            "buyer@example.com",
            str(tmp_path),
            1024,
        )


def test_required_attachment_category_controls_readiness(db: Session, tmp_path: Path) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(), "admin@example.com")
    seed_bpp_purchasing(db, actor="admin@example.com")
    request = create_purchase_request(
        db,
        PurchaseRequestCreate(
            workflow_code="BPP_PURCHASING", store_number="1001", vendor_code="V-100"
        ),
        "buyer@example.com",
    )
    add_line_item(
        db,
        request,
        PurchaseLineWrite(product_code="P-100", quantity=2),
        "buyer@example.com",
    )
    upsert_config_entry(
        db,
        ConfigEntryWrite(
            scope_type="purchasing",
            scope_key="BPP_PURCHASING",
            key="rules.required_attachment_categories",
            value={"categories": ["quote"]},
            updated_by="admin@example.com",
        ),
    )
    assert "attachment.required" in {
        issue.code for issue in validate_purchase_request(db, request).errors
    }

    store_attachment(
        db,
        request,
        "quote.pdf",
        "application/pdf",
        b"%PDF-1.7\nquote",
        AttachmentCategory.QUOTE,
        "buyer@example.com",
        str(tmp_path),
        1024,
    )
    assert validate_purchase_request(db, request).ready is True


def _request_at_po_created(
    db: Session,
    entity_suffix: str,
    user: User,
    line_count: int = 1,
    store_number: str = "1001",
) -> PurchaseRequest:
    request = create_purchase_request(
        db,
        PurchaseRequestCreate(
            workflow_code="BPP_PURCHASING",
            store_number=store_number,
            vendor_code="V-100",
            context={"validation_suffix": entity_suffix},
        ),
        user.email,
    )
    for _index in range(line_count):
        add_line_item(
            db,
            request,
            PurchaseLineWrite(
                product_code="P-100",
                quantity=2,
                freight_amount=Decimal("10"),
                tax_amount=Decimal("5"),
            ),
            user.email,
        )
    submit_purchase_request(db, request, user, {"workflow.bpp.submit"})
    actions = (
        ("submit_for_department_review", "workflow.bpp.submit"),
        ("department_approve", "workflow.bpp.department_review"),
        ("purchasing_approve", "workflow.bpp.purchasing_review"),
        ("select_vendor", "workflow.bpp.vendor_select"),
        ("verify_cost", "workflow.bpp.cost_verify"),
        ("executive_approve", "workflow.bpp.executive_approve"),
    )
    for action, permission in actions:
        advance_workflow(
            db,
            request.workflow_instance_id,
            FlowActionRequest(action=action, actor=user.email),
            {permission},
        )
    db.refresh(request)
    assert request.status == "po_created"
    return request


def test_purchase_order_generation_numbering_snapshot_and_duplicate_guard(db: Session) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(), "admin@example.com")
    seed_bpp_purchasing(db, actor="admin@example.com")
    seed_purchase_order_defaults(db, "admin@example.com")
    user = User(
        email="buyer@example.com",
        display_name="Buyer",
        password_hash="unused",
        region_code="EAST",
        is_active=True,
    )
    db.add(user)
    db.commit()
    request = _request_at_po_created(db, "one", user)

    with pytest.raises(PermissionError, match="workflow.bpp.po_generate"):
        generate_purchase_orders(db, [request.id], user.email, set())
    orders = generate_purchase_orders(db, [request.id], user.email, {"workflow.bpp.po_generate"})
    assert len(orders) == 1
    order = orders[0]
    assert order.po_number.startswith("PO-")
    assert order.total == request.total == Decimal("265.0000")
    assert len(order.sources) == 1
    assert len(order.lines) == 1
    assert order.lines[0].source_line_id == request.line_items[0].id
    assert (
        db.scalar(
            select(func.count())
            .select_from(EventSnapshot)
            .where(
                EventSnapshot.entity_type == "purchase_order",
                EventSnapshot.entity_id == order.id,
                EventSnapshot.event_type == "purchase_order.created",
            )
        )
        == 1
    )
    with pytest.raises(PurchaseOrderError, match="already exists"):
        generate_purchase_orders(db, [request.id], user.email, {"workflow.bpp.po_generate"})


def test_purchase_order_consolidates_same_workflow_vendor_and_currency(db: Session) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(), "admin@example.com")
    seed_bpp_purchasing(db, actor="admin@example.com")
    user = User(
        email="buyer@example.com",
        display_name="Buyer",
        password_hash="unused",
        region_code="EAST",
        is_active=True,
    )
    db.add(user)
    db.commit()
    first = _request_at_po_created(db, "first", user)
    second = _request_at_po_created(db, "second", user)

    orders = generate_purchase_orders(
        db,
        [first.id, second.id],
        user.email,
        {"workflow.bpp.po_generate"},
    )
    assert len(orders) == 1
    assert orders[0].total == Decimal("530.0000")
    assert len(orders[0].sources) == 2
    assert len(orders[0].lines) == 2
    assert db.scalar(select(func.count()).select_from(PurchaseOrder)) == 1
    assert db.scalar(select(func.count()).select_from(PurchaseOrderSource)) == 2


def test_purchase_order_splits_by_max_lines_with_financial_integrity(db: Session) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(), "admin@example.com")
    seed_bpp_purchasing(db, actor="admin@example.com")
    user = User(
        email="buyer@example.com",
        display_name="Buyer",
        password_hash="unused",
        region_code="EAST",
        is_active=True,
    )
    db.add(user)
    db.commit()
    request = _request_at_po_created(db, "split", user, line_count=2)
    upsert_config_entry(
        db,
        ConfigEntryWrite(
            scope_type="purchase_order",
            scope_key="default",
            key="split.max_lines",
            value={"count": 1},
            updated_by="admin@example.com",
        ),
    )

    orders = generate_purchase_orders(db, [request.id], user.email, {"workflow.bpp.po_generate"})
    assert len(orders) == 2
    assert all(len(order.lines) == 1 for order in orders)
    assert sum((order.total for order in orders), Decimal("0")) == request.total
    assert len({order.po_number for order in orders}) == 2
    assert db.scalar(select(func.count()).select_from(PurchaseOrderSource)) == 2


def test_purchase_order_partition_respects_max_total() -> None:
    source = SimpleNamespace(id="request-1")
    lines = [
        (source, SimpleNamespace(extended_amount=Decimal("200"))),
        (source, SimpleNamespace(extended_amount=Decimal("150"))),
        (source, SimpleNamespace(extended_amount=Decimal("100"))),
    ]

    partitions = _partition_lines(lines, max_lines=0, max_total=Decimal("300"))

    assert [len(partition) for partition in partitions] == [1, 2]


def test_purchase_order_store_grouping_prevents_cross_store_consolidation(db: Session) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(), "admin@example.com")
    seed_bpp_purchasing(db, actor="admin@example.com")
    db.add(
        Store(
            store_number="1002",
            name="Second Store",
            region_code="EAST",
            is_active=True,
            is_ordering_enabled=True,
        )
    )
    user = User(
        email="buyer@example.com",
        display_name="Buyer",
        password_hash="unused",
        region_code="EAST",
        is_active=True,
    )
    db.add(user)
    db.commit()
    first = _request_at_po_created(db, "store-one", user, store_number="1001")
    second = _request_at_po_created(db, "store-two", user, store_number="1002")
    upsert_config_entry(
        db,
        ConfigEntryWrite(
            scope_type="purchase_order",
            scope_key="default",
            key="consolidation.by_store",
            value={"enabled": True},
            updated_by="admin@example.com",
        ),
    )

    orders = generate_purchase_orders(
        db,
        [first.id, second.id],
        user.email,
        {"workflow.bpp.po_generate"},
    )
    assert len(orders) == 2
    assert {order.sources[0].store_number for order in orders} == {"1001", "1002"}


def test_purchase_order_artifacts_are_immutable_idempotent_and_checksummed(
    db: Session, tmp_path: Path
) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(), "admin@example.com")
    seed_bpp_purchasing(db, actor="admin@example.com")
    user = User(
        email="buyer@example.com",
        display_name="Buyer",
        password_hash="unused",
        region_code="EAST",
        is_active=True,
    )
    db.add(user)
    db.commit()
    request = _request_at_po_created(db, "artifacts", user)
    order = generate_purchase_orders(db, [request.id], user.email, {"workflow.bpp.po_generate"})[0]

    generated = {
        artifact_format: generate_artifact(db, order, artifact_format, user.email, str(tmp_path))
        for artifact_format in PurchaseOrderArtifactFormat
    }
    for artifact_format, artifact in generated.items():
        content = po_artifact_path(artifact, str(tmp_path)).read_bytes()
        assert artifact.sha256 == hashlib.sha256(content).hexdigest()
        assert artifact.size_bytes == len(content)
        assert artifact.artifact_format == artifact_format.value
    assert (
        po_artifact_path(generated[PurchaseOrderArtifactFormat.PDF], str(tmp_path))
        .read_bytes()
        .startswith(b"%PDF-")
    )
    json_content = po_artifact_path(
        generated[PurchaseOrderArtifactFormat.JSON], str(tmp_path)
    ).read_text()
    assert json.loads(json_content)["po_number"] == order.po_number
    csv_content = po_artifact_path(
        generated[PurchaseOrderArtifactFormat.CSV], str(tmp_path)
    ).read_text()
    assert "source_request_id" in csv_content
    repeated = generate_artifact(
        db,
        order,
        PurchaseOrderArtifactFormat.PDF,
        user.email,
        str(tmp_path),
    )
    assert repeated.id == generated[PurchaseOrderArtifactFormat.PDF].id
    assert db.scalar(select(func.count()).select_from(PurchaseOrderArtifact)) == 3
    assert (
        db.scalar(
            select(func.count())
            .select_from(EventSnapshot)
            .where(
                EventSnapshot.entity_id == order.id,
                EventSnapshot.event_type == "purchase_order.artifact_created",
            )
        )
        == 3
    )


def test_purchase_order_internal_transmission_lifecycle_is_audited(
    db: Session, tmp_path: Path
) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(), "admin@example.com")
    seed_bpp_purchasing(db, actor="admin@example.com")
    user = User(
        email="buyer@example.com",
        display_name="Buyer",
        password_hash="unused",
        region_code="EAST",
        is_active=True,
    )
    db.add(user)
    db.commit()
    request = _request_at_po_created(db, "transmission", user)
    order = generate_purchase_orders(db, [request.id], user.email, {"workflow.bpp.po_generate"})[0]
    artifact = generate_artifact(
        db,
        order,
        PurchaseOrderArtifactFormat.PDF,
        user.email,
        str(tmp_path),
    )
    with pytest.raises(PurchaseOrderTransmissionError, match="vendor_submission"):
        create_transmission(
            db,
            order,
            artifact.id,
            PurchaseOrderTransmissionChannel.MANUAL,
            None,
            None,
            user.email,
            {"workflow.bpp.vendor_submit"},
        )
    advance_workflow(
        db,
        request.workflow_instance_id,
        FlowActionRequest(action="generate_po", actor=user.email),
        {"workflow.bpp.po_generate"},
    )
    with pytest.raises(PermissionError, match="workflow.bpp.vendor_submit"):
        create_transmission(
            db,
            order,
            artifact.id,
            PurchaseOrderTransmissionChannel.MANUAL,
            None,
            None,
            user.email,
            set(),
        )
    transmission = create_transmission(
        db,
        order,
        artifact.id,
        PurchaseOrderTransmissionChannel.MANUAL,
        "Purchasing shared drive",
        "Operator-controlled delivery",
        user.email,
        {"workflow.bpp.vendor_submit"},
    )
    assert transmission.status == "prepared"
    actions = (
        (PurchaseOrderTransmissionAction.RELEASE, None, "ready"),
        (PurchaseOrderTransmissionAction.MARK_FAILED, "Printer unavailable", "failed"),
        (PurchaseOrderTransmissionAction.RETRY, "Printer restored", "prepared"),
        (PurchaseOrderTransmissionAction.RELEASE, None, "ready"),
        (PurchaseOrderTransmissionAction.MARK_DELIVERED, "Handed to operator", "delivered"),
    )
    for action, reason, expected in actions:
        transmission = apply_transmission_action(
            db,
            order,
            transmission,
            action,
            reason,
            user.email,
            {"workflow.bpp.vendor_submit"},
        )
        assert transmission.status == expected
    db.refresh(order)
    assert order.status == "transmitted"
    assert len(transmission.events) == 6
    assert (
        db.scalar(
            select(func.count())
            .select_from(PurchaseOrderTransmissionEvent)
            .where(PurchaseOrderTransmissionEvent.transmission_id == transmission.id)
        )
        == 6
    )
    with pytest.raises(PurchaseOrderTransmissionError, match="not valid"):
        apply_transmission_action(
            db,
            order,
            transmission,
            PurchaseOrderTransmissionAction.RELEASE,
            None,
            user.email,
            {"workflow.bpp.vendor_submit"},
        )


def test_purchase_order_generation_preserves_independent_workflow_boundary(db: Session) -> None:
    import_catalog(db, "catalog.xlsx", workbook_bytes(), "admin@example.com")
    seed_independent_purchasing(db, actor="admin@example.com")
    user = User(
        email="independent@example.com",
        display_name="Independent Buyer",
        password_hash="unused",
        region_code="EAST",
        is_active=True,
    )
    db.add(user)
    db.commit()
    request = create_purchase_request(
        db,
        PurchaseRequestCreate(
            workflow_code="IND_PURCHASING",
            store_number="1001",
            vendor_code="V-100",
        ),
        user.email,
    )
    add_line_item(
        db,
        request,
        PurchaseLineWrite(product_code="P-100", quantity=2),
        user.email,
    )
    submit_purchase_request(db, request, user, {"workflow.ind.submit"})
    actions = (
        ("submit_for_store_review", "workflow.ind.submit"),
        ("store_approve", "workflow.ind.review"),
        ("franchise_approve", "workflow.ind.franchise_approve"),
        ("select_vendor", "workflow.ind.vendor_select"),
        ("verify_pricing", "workflow.ind.review"),
        ("regional_approve", "workflow.ind.regional_approve"),
    )
    for action, permission in actions:
        advance_workflow(
            db,
            request.workflow_instance_id,
            FlowActionRequest(action=action, actor=user.email),
            {permission},
        )
    db.refresh(request)
    assert request.status == "po_created"

    order = generate_purchase_orders(db, [request.id], user.email, {"workflow.ind.review"})[0]
    assert order.workflow_code == "IND_PURCHASING"
    assert order.vendor_code == request.vendor_code
    assert order.sources[0].purchase_request_id == request.id
